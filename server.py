#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
CRM de Leads do Agro — servidor sem dependencias (so a biblioteca padrao do Python).

Como rodar:
    python3 server.py

Abre em http://localhost:3000
Webhook do Chatwoot: http://localhost:3000/webhook/chatwoot?token=SEU_TOKEN
(o token aparece no terminal quando o servidor inicia)
"""

import csv
import heapq
import io
import json
import math
import os
import re
import base64
import hashlib
import secrets
import shutil
import sqlite3
import gzip
import threading
import time
import unicodedata
from datetime import datetime, timezone, timedelta
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
from urllib.parse import urlparse, parse_qs
import urllib.request
import urllib.error

# ---------------------------------------------------------------------------
# Configuracao
# ---------------------------------------------------------------------------
PORT = int(os.environ.get("PORT", "3000"))

# Token do webhook (protege o endpoint). Prioridade: variavel de ambiente
# WEBHOOK_TOKEN; senao, um token aleatorio persistido no banco (settings) na
# primeira execucao. NUNCA derivado do nome de usuario (seria adivinhavel).
WEBHOOK_TOKEN = os.environ.get("WEBHOOK_TOKEN") or None  # resolvido no boot

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
# DATA_DIR pode ser sobrescrito por variavel de ambiente (ex.: um volume /data
# na nuvem), para os leads nao sumirem quando o servidor reinicia.
DATA_DIR = os.environ.get("DATA_DIR") or os.path.join(BASE_DIR, "data")
DB_FILE = os.path.join(DATA_DIR, "leads.json")
FOTOS_DIR = os.path.join(DATA_DIR, "fotos")  # fotos das visitas de campo
PUBLIC_DIR = os.path.join(BASE_DIR, "public")

# ---------------------------------------------------------------------------
# ATLAS (prospeccao): banco SQLite com as 242 mil fazendas de Goias (SICAR +
# MapBiomas), compactado no repositorio (atlas.db.gz) e descompactado para
# data/atlas.db no primeiro boot. data/ fica fora do git — as edicoes da
# equipe (categorias, contatos, registros, territorios) vivem la e NAO sao
# sobrescritas em deploys (so descompacta se o arquivo ainda nao existe).
# ---------------------------------------------------------------------------
ATLAS_GZ = os.path.join(BASE_DIR, "atlas.db.gz")
ATLAS_DB = os.path.join(DATA_DIR, "atlas.db")
# Pivos centrais de irrigacao (ANA, dados abertos). Vem em arquivo proprio para
# nao depender do atlas.db.gz de 50 MB: a cada boot conferimos a versao.
PIVOS_GZ = os.path.join(BASE_DIR, "pivos.json.gz")
INICIO_ISO = datetime.now(timezone.utc).isoformat(timespec="seconds")
# marcos publicados — /api/status usa isto para conferir se o deploy chegou
RECURSOS = ["funil", "servicos", "curso", "tarefas", "atlas", "proprietarios",
            "potencial", "pivos"]
# A "versao" dos dados e a impressao digital do proprio atlas.db.gz: gerou um
# arquivo novo, o boot troca a base sozinho (preservando o que a equipe editou)
# — sem depender de ninguem lembrar de subir um numero de versao.
def _atlas_hash_gz():
    if not os.path.exists(ATLAS_GZ):
        return None
    h = hashlib.sha1()
    with open(ATLAS_GZ, "rb") as f:
        for bloco in iter(lambda: f.read(1024 * 1024), b""):
            h.update(bloco)
    return h.hexdigest()

# tabelas com o TRABALHO DA EQUIPE (migram para a base nova)
ATLAS_TABELAS_EQUIPE = ("pessoas", "fazenda_pessoas", "registros",
                        "territorios", "territorio_municipios")


def _atlas_versao_do(caminho):
    try:
        con = sqlite3.connect(caminho)
        r = con.execute("SELECT v FROM atlas_meta LIMIT 1").fetchone()
        con.close()
        return r[0] if r else None
    except Exception:
        return None


def _atlas_schema_equipe(con):
    """Tabelas que guardam o trabalho da equipe (nao vem no export do Atlas)."""
    con.executescript("""
        CREATE TABLE IF NOT EXISTS territorios (
            id INTEGER PRIMARY KEY AUTOINCREMENT, nome TEXT NOT NULL,
            criado_em TEXT DEFAULT (datetime('now','localtime')));
        CREATE TABLE IF NOT EXISTS territorio_municipios (
            territorio_id INTEGER NOT NULL, municipio_id INTEGER NOT NULL,
            PRIMARY KEY (territorio_id, municipio_id));
        CREATE TABLE IF NOT EXISTS registros (
            id INTEGER PRIMARY KEY AUTOINCREMENT, fazenda_id INTEGER NOT NULL,
            autor TEXT, texto TEXT NOT NULL,
            criado_em TEXT DEFAULT (datetime('now','localtime')));
    """)


def _atlas_migra_edicoes(antigo, novo):
    """Leva o trabalho da equipe do banco ANTIGO para a base NOVA.

    Duas regras que a versao anterior quebrava:
    1. NUNCA apagar o que ja vem na base nova (pessoas/vinculos do enriquecimento
       sao 22 mil linhas — o DELETE anterior jogava tudo fora).
    2. NUNCA confiar em id entre bases diferentes: fazenda e municipio sao
       religados pelo codigo do CAR / codigo do IBGE, que nao mudam. Linha cujo
       destino nao existe mais na base nova e descartada (e contada no aviso)."""
    con = sqlite3.connect(novo)
    _atlas_schema_equipe(con)   # as tabelas da equipe precisam existir aqui
    # sem esta chave o "INSERT OR IGNORE" nao ignora nada e duplica vinculos
    try:
        con.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_fp_unico "
                    "ON fazenda_pessoas(fazenda_id, pessoa_id)")
    except sqlite3.Error:
        pass
    con.execute("ATTACH ? AS velho", [antigo])
    tabs_velho = {r[0] for r in con.execute(
        "SELECT name FROM velho.sqlite_master WHERE type='table'").fetchall()}
    perdidas = 0

    def de_para(tabela, chave):
        """id do banco velho -> id do banco novo, casando pela chave estavel."""
        if tabela not in tabs_velho:
            return {}
        novos = {str(k).strip(): i for i, k in con.execute(
            "SELECT id, %s FROM %s WHERE %s IS NOT NULL" % (chave, tabela, chave)).fetchall()}
        mapa = {}
        for i, k in con.execute(
                "SELECT id, %s FROM velho.%s WHERE %s IS NOT NULL" % (chave, tabela, chave)).fetchall():
            j = novos.get(str(k).strip())
            if j is not None:
                mapa[i] = j
        return mapa

    faz = de_para("fazendas", "codigo_car")
    mun = de_para("municipios", "codigo_ibge")

    # --- pessoas: mantem as da base nova; traz as que so existem no banco velho
    if "pessoas" in tabs_velho:
        cols_n = [c[1] for c in con.execute("PRAGMA table_info(pessoas)").fetchall()]
        cols_v = [c[1] for c in con.execute("PRAGMA velho.table_info(pessoas)").fetchall()]
        cols = [c for c in cols_n if c in cols_v and c != "id"]
        lista = ", ".join(cols)
        prox = (con.execute("SELECT COALESCE(MAX(id),0) FROM pessoas").fetchone()[0] or 0) + 1
        pes = {}
        for linha in con.execute("SELECT id, %s FROM velho.pessoas" % lista).fetchall():
            vid, dados = linha[0], linha[1:]
            achou = con.execute(
                "SELECT id FROM pessoas WHERE (documento IS NOT NULL AND documento <> '' "
                "AND documento = ?) OR UPPER(nome) = UPPER(?)",
                [dados[cols.index("documento")] if "documento" in cols else None,
                 dados[cols.index("nome")] if "nome" in cols else ""]).fetchone()
            if achou:
                pes[vid] = achou[0]
                # completa o que faltava (telefone/e-mail que a equipe cadastrou)
                for campo in ("telefone", "email", "documento"):
                    if campo in cols:
                        con.execute("UPDATE pessoas SET %s = COALESCE(%s, ?) WHERE id = ?" % (campo, campo),
                                    [dados[cols.index(campo)], achou[0]])
            else:
                con.execute("INSERT INTO pessoas (id, %s) VALUES (?%s)" % (lista, ", ?" * len(cols)),
                            [prox] + list(dados))
                pes[vid] = prox
                prox += 1
    else:
        pes = {}

    # --- vinculos fazenda<->pessoa (religados pelo CAR)
    if "fazenda_pessoas" in tabs_velho:
        for fid, pid, rf, rc in con.execute(
                "SELECT fazenda_id, pessoa_id, relacao_fundiaria, relacao_comercial "
                "FROM velho.fazenda_pessoas").fetchall():
            f2, p2 = faz.get(fid), pes.get(pid)
            if f2 is None or p2 is None:
                perdidas += 1
                continue
            con.execute("INSERT OR IGNORE INTO fazenda_pessoas "
                        "(fazenda_id, pessoa_id, relacao_fundiaria, relacao_comercial) "
                        "VALUES (?,?,?,?)", [f2, p2, rf, rc])

    # --- registros da equipe (religados pelo CAR)
    if "registros" in tabs_velho:
        for fid, autor, texto, quando in con.execute(
                "SELECT fazenda_id, autor, texto, criado_em FROM velho.registros").fetchall():
            f2 = faz.get(fid)
            if f2 is None:
                perdidas += 1
                continue
            con.execute("INSERT INTO registros (fazenda_id, autor, texto, criado_em) "
                        "VALUES (?,?,?,?)", [f2, autor, texto, quando])

    # --- territorios (municipios religados pelo codigo do IBGE)
    if "territorios" in tabs_velho:
        for tid, nome, quando in con.execute(
                "SELECT id, nome, criado_em FROM velho.territorios").fetchall():
            cur = con.execute("INSERT INTO territorios (nome, criado_em) VALUES (?,?)", [nome, quando])
            novo_tid = cur.lastrowid
            for (mid,) in con.execute(
                    "SELECT municipio_id FROM velho.territorio_municipios WHERE territorio_id = ?",
                    [tid]).fetchall():
                m2 = mun.get(mid)
                if m2 is None:
                    perdidas += 1
                    continue
                con.execute("INSERT OR IGNORE INTO territorio_municipios "
                            "(territorio_id, municipio_id) VALUES (?,?)", [novo_tid, m2])

    # categorias marcadas pela equipe (casadas pelo codigo do CAR, que nao muda)
    if "fazendas" in tabs_velho:
        con.execute("""UPDATE fazendas SET categoria = (
                SELECT v.categoria FROM velho.fazendas v
                 WHERE v.codigo_car = fazendas.codigo_car)
            WHERE EXISTS (SELECT 1 FROM velho.fazendas v
                 WHERE v.codigo_car = fazendas.codigo_car
                   AND v.categoria IS NOT NULL AND v.categoria <> 'sem_categoria')""")
    con.commit()
    con.execute("DETACH velho")
    con.close()
    if perdidas:
        print("[atlas] AVISO: %d vínculo(s)/registro(s) não encontraram destino na base nova" % perdidas)


def atlas_boot():
    versao = _atlas_hash_gz()
    if versao and _atlas_versao_do(ATLAS_DB) != versao:
        os.makedirs(DATA_DIR, exist_ok=True)
        tmp = ATLAS_DB + ".novo"
        try:
            with gzip.open(ATLAS_GZ, "rb") as fin, open(tmp, "wb") as fout:
                shutil.copyfileobj(fin, fout)
            con = sqlite3.connect(tmp)
            con.executescript("CREATE TABLE IF NOT EXISTS atlas_meta (v TEXT);"
                              "DELETE FROM atlas_meta;")
            con.execute("INSERT INTO atlas_meta (v) VALUES (?)", [versao])
            con.commit()
            con.close()
            if os.path.exists(ATLAS_DB):
                # já existe base antiga: preserva o trabalho da equipe
                _atlas_migra_edicoes(ATLAS_DB, tmp)
                # cópia de segurança só se sobrar disco de verdade (o banco tem
                # ~90 MB; num servidor apertado a cópia é o que derrubaria tudo)
                try:
                    livre = shutil.disk_usage(DATA_DIR).free
                    if livre > os.path.getsize(ATLAS_DB) * 3:
                        shutil.copy2(ATLAS_DB, ATLAS_DB + ".anterior")
                except OSError:
                    pass
                print("[atlas] base atualizada (edições da equipe preservadas)")
            else:
                print("[atlas] banco instalado em %s" % ATLAS_DB)
            os.replace(tmp, ATLAS_DB)
            for sufixo in ("-wal", "-shm"):   # sobras do banco antigo
                try:
                    os.remove(ATLAS_DB + sufixo)
                except OSError:
                    pass
        except Exception:
            # descompressao pela metade (disco cheio, gz ruim): nao deixar lixo
            try:
                os.remove(tmp)
            except OSError:
                pass
            raise
    if os.path.exists(ATLAS_DB):
        con = sqlite3.connect(ATLAS_DB)
        # WAL: leituras nunca bloqueiam e escritas concorrem melhor (fica
        # gravado no arquivo — basta uma vez)
        con.execute("PRAGMA journal_mode=WAL")
        _atlas_schema_equipe(con)
        con.execute("CREATE INDEX IF NOT EXISTS idx_faz_latlng ON fazendas(latitude, longitude)")
        con.execute("CREATE INDEX IF NOT EXISTS idx_fp_pessoa ON fazenda_pessoas(pessoa_id)")
        con.execute("CREATE INDEX IF NOT EXISTS idx_pes_nome ON pessoas(nome)")
        # Sem estes, o SQLite monta um indice temporario A CADA consulta
        # ("AUTOMATIC PARTIAL COVERING INDEX") porque o export perdeu as chaves,
        # e a busca por cultura de cada fazenda varria a tabela inteira.
        con.execute("CREATE UNIQUE INDEX IF NOT EXISTS idx_faz_id ON fazendas(id)")
        con.execute("CREATE INDEX IF NOT EXISTS idx_fc_faz_cult "
                    "ON fazenda_culturas(fazenda_id, cultura_id)")
        con.execute("CREATE INDEX IF NOT EXISTS idx_fp_faz ON fazenda_pessoas(fazenda_id)")
        # MIGRACAO CRITICA: o export "CREATE TABLE AS SELECT" perdeu a chave
        # primaria de pessoas — sem ela, lastrowid NAO corresponde ao campo id
        # e um contato novo apontaria para OUTRA pessoa (ou sumiria). Reconstroi
        # a tabela com INTEGER PRIMARY KEY preservando os ids existentes.
        info = con.execute("PRAGMA table_info(pessoas)").fetchall()
        id_e_pk = any(c[1] == "id" and c[5] == 1 for c in info)
        if not id_e_pk:
            con.executescript("""
                BEGIN;
                CREATE TABLE pessoas_pk (
                    id INTEGER PRIMARY KEY, nome TEXT, tipo TEXT,
                    documento TEXT, telefone TEXT, email TEXT);
                INSERT INTO pessoas_pk (id, nome, tipo, documento, telefone, email)
                    SELECT id, nome, tipo, documento, telefone, email
                    FROM pessoas WHERE id IS NOT NULL;
                DROP TABLE pessoas;
                ALTER TABLE pessoas_pk RENAME TO pessoas;
                COMMIT;
            """)
            print("[atlas] tabela pessoas migrada para chave primaria de verdade")
        _pivos_boot(con)
        con.commit()
        con.close()


def _pivos_boot(con):
    """Instala/atualiza os pivos centrais no atlas.db a partir do pivos.json.gz.
    A fonte da verdade e o arquivo do repositorio, entao a troca da base do
    Atlas nao precisa preservar nada: se o arquivo mudar, a tabela e refeita."""
    if not os.path.exists(PIVOS_GZ):
        return
    try:
        h = hashlib.sha1()
        with open(PIVOS_GZ, "rb") as f:
            for bloco in iter(lambda: f.read(1024 * 512), b""):
                h.update(bloco)
        versao = h.hexdigest()
        con.execute("CREATE TABLE IF NOT EXISTS pivos_meta (v TEXT)")
        atual = con.execute("SELECT v FROM pivos_meta LIMIT 1").fetchone()
        tem_tab = con.execute("SELECT 1 FROM sqlite_master WHERE type='table' "
                              "AND name='pivos'").fetchone()
        if atual and atual[0] == versao and tem_tab:
            return
        with gzip.open(PIVOS_GZ, "rt", encoding="utf-8") as g:
            dados = json.load(g)
        con.execute("DROP TABLE IF EXISTS pivos")
        con.execute("CREATE TABLE pivos (id INTEGER PRIMARY KEY, lat REAL, lng REAL, "
                    "ha REAL, ano TEXT, ibge TEXT, municipio_id INT)")
        # de codigo do IBGE para o municipio do Atlas (a chave que os dois lados tem)
        mapa = {str(r[1]): r[0] for r in
                con.execute("SELECT id, codigo_ibge FROM municipios").fetchall() if r[1]}
        linhas = [(round(float(p[1]), 6), round(float(p[0]), 6), float(p[2] or 0),
                   str(p[4] or ""), str(p[3] or ""), mapa.get(str(p[3] or "")))
                  for p in dados.get("pivos", [])
                  if p and p[0] is not None and p[1] is not None]
        con.executemany("INSERT INTO pivos (lat, lng, ha, ano, ibge, municipio_id) "
                        "VALUES (?,?,?,?,?,?)", linhas)
        con.execute("CREATE INDEX IF NOT EXISTS idx_piv_latlng ON pivos(lat, lng)")
        con.execute("CREATE INDEX IF NOT EXISTS idx_piv_mun ON pivos(municipio_id)")
        con.execute("DELETE FROM pivos_meta")
        con.execute("INSERT INTO pivos_meta (v) VALUES (?)", [versao])
        con.commit()
        sem_mun = sum(1 for l in linhas if l[5] is None)
        print("[atlas] %d pivos centrais instalados (%d sem municipio)"
              % (len(linhas), sem_mun))
    except Exception as e:
        # sem pivos o CRM continua funcionando; melhor que derrubar o servidor
        print("[atlas] nao foi possivel instalar os pivos:", e)


def _sem_acento(txt):
    """MARIA JOSÉ -> MARIA JOSE (busca de família não pode depender de acento)."""
    if txt is None:
        return ""
    return "".join(c for c in unicodedata.normalize("NFD", str(txt))
                   if unicodedata.category(c) != "Mn").upper()


def atlas_con():
    con = sqlite3.connect(ATLAS_DB, timeout=15)
    con.row_factory = sqlite3.Row
    # busca por nome ignora acento e maiúscula (ROMUÁLDO acha ROMUALDO)
    con.create_function("semacento", 1, _sem_acento)
    return con


# planilhas em analise na importacao do Atlas (token -> linhas), em memoria
_atlas_imports = {}


def _atlas_dp(pts, tol):
    """Douglas-Peucker sobre uma linha aberta (portado do atlas-agro)."""
    if len(pts) < 3:
        return pts
    pilha, manter = [(0, len(pts) - 1)], {0, len(pts) - 1}
    while pilha:
        ini, fim = pilha.pop()
        x1, y1 = pts[ini]
        x2, y2 = pts[fim]
        dx, dy = x2 - x1, y2 - y1
        norma = (dx * dx + dy * dy) ** 0.5
        pior, idx = 0.0, -1
        for i in range(ini + 1, fim):
            x0, y0 = pts[i]
            if norma:
                d = abs(dy * x0 - dx * y0 + x2 * y1 - y2 * x1) / norma
            else:
                d = ((x0 - x1) ** 2 + (y0 - y1) ** 2) ** 0.5
            if d > pior:
                pior, idx = d, i
        if pior > tol and idx > 0:
            manter.add(idx)
            pilha.append((ini, idx))
            pilha.append((idx, fim))
    return [pts[i] for i in sorted(manter)]


def _atlas_decodifica(d):
    """Contorno em DELTA -> lista de aneis [[lon,lat],...].

    Formato: aneis separados por "|", pontos por ";", e cada ponto e a
    DIFERENCA (em 1e-5 grau) para o ponto anterior — 13x menor que guardar
    o geojson, sem perda visivel no mapa."""
    aneis = []
    for parte in str(d or "").split("|"):
        if not parte:
            continue
        x = y = 0
        anel = []
        for par in parte.split(";"):
            try:
                dx, dy = par.split(",")
                x += int(dx)
                y += int(dy)
            except ValueError:
                continue
            anel.append([x / 1e5, y / 1e5])
        if len(anel) >= 3:
            aneis.append(anel)
    return aneis


def _atlas_simplifica(anel, tol=0.0004):
    """Reduz pontos do contorno mantendo o formato (~40 m de tolerancia)."""
    if len(anel) < 30:
        return anel
    fechado = anel[0] == anel[-1]
    pts = anel[:-1] if fechado else anel
    meio = len(pts) // 2
    saida = _atlas_dp(pts[:meio + 1], tol) + _atlas_dp(pts[meio:], tol)[1:]
    if fechado and saida and saida[0] != saida[-1]:
        saida.append(saida[0])
    return saida if len(saida) >= 4 else anel


def _atlas_le_planilha(nome, dados):
    """(abas, cabecalho, linhas) de um CSV (stdlib) ou XLSX (se houver openpyxl)."""
    if nome.lower().endswith(".csv"):
        texto = dados.decode("utf-8-sig", "replace")
        amostra = texto[:8192]
        try:
            sep = csv.Sniffer().sniff(amostra, delimiters=",;\t").delimiter
        except Exception:
            sep = ";" if amostra.count(";") > amostra.count(",") else ","
        linhas = list(csv.reader(io.StringIO(texto), delimiter=sep))
        return ["(csv)"], (linhas[0] if linhas else []), linhas[1:100001]
    try:
        import openpyxl  # opcional; sem ele orientamos salvar como CSV
    except ImportError:
        raise ValueError("este servidor lê só .csv — no Excel use Arquivo → Salvar como → CSV e envie de novo")
    wb = openpyxl.load_workbook(io.BytesIO(dados), read_only=True, data_only=True)
    abas = wb.sheetnames
    ws = wb[abas[0]]
    it = ws.iter_rows(values_only=True)
    cab = [str(c) if c is not None else "" for c in next(it, [])]
    linhas = []
    for i, r in enumerate(it):
        if i >= 100000:
            break
        linhas.append(list(r))
    wb.close()
    return abas, cab, linhas


def _atlas_dentro(poligono, x, y):
    """Ponto dentro do poligono (algoritmo do raio)."""
    dentro = False
    n = len(poligono)
    for i in range(n):
        x1, y1 = poligono[i]
        x2, y2 = poligono[(i + 1) % n]
        if (y1 > y) != (y2 > y):
            xi = (x2 - x1) * (y - y1) / ((y2 - y1) or 1e-15) + x1
            if x < xi:
                dentro = not dentro
    return dentro


def _atlas_aneis_de(con, fazenda_id):
    """Anéis do contorno de uma fazenda, seja qual for o formato da base."""
    try:
        r = con.execute("SELECT d FROM fazenda_contorno_d WHERE fazenda_id = ?",
                        [fazenda_id]).fetchone()
        if r:
            return _atlas_decodifica(r["d"] if not isinstance(r, tuple) else r[0])
    except sqlite3.Error:
        pass
    try:
        r = con.execute("SELECT geojson FROM fazenda_contorno WHERE fazenda_id = ?",
                        [fazenda_id]).fetchone()
        if r:
            return json.loads(r["geojson"] if not isinstance(r, tuple) else r[0])
    except (sqlite3.Error, ValueError):
        pass
    return []


def _atlas_acha_fazenda(con, tem_contorno, car=None, lat=None, lon=None, nome=None, municipio=None):
    """A qual fazenda a linha da planilha se refere (CAR > coordenada > nome)."""
    if car:
        r = con.execute("SELECT id FROM fazendas WHERE codigo_car = ?", [str(car).strip()]).fetchone()
        if r:
            return r["id"], "código CAR"
    if lat is not None and lon is not None:
        d = 0.05  # ~5 km
        cands = con.execute(
            "SELECT id, latitude, longitude FROM fazendas "
            "WHERE latitude BETWEEN ? AND ? AND longitude BETWEEN ? AND ?",
            [lat - d, lat + d, lon - d, lon + d]).fetchall()
        melhor, dist = None, None
        for c in cands:
            if tem_contorno:
                # a base compacta guarda o contorno codificado (fazenda_contorno_d);
                # a antiga, em geojson. Usar a tabela errada quebrava a importação.
                aneis = _atlas_aneis_de(con, c["id"])
                for anel in aneis:
                    if _atlas_dentro(anel, lon, lat):
                        return c["id"], "coordenada dentro do contorno"
            dd = (c["latitude"] - lat) ** 2 + (c["longitude"] - lon) ** 2
            if dist is None or dd < dist:
                melhor, dist = c["id"], dd
        if melhor and dist is not None and dist ** 0.5 < 0.02:  # ate ~2 km
            return melhor, "coordenada próxima"
    if nome and municipio:
        r = con.execute(
            "SELECT f.id FROM fazendas f JOIN municipios m ON m.id = f.municipio_id "
            "WHERE UPPER(f.nome) = UPPER(?) AND UPPER(m.nome) = UPPER(?) LIMIT 1",
            [str(nome).strip(), str(municipio).strip()]).fetchone()
        if r:
            return r["id"], "nome + município"
    return None, None


def _atlas_importa(con, linhas, mapa, categoria, origem, tem_contorno):
    """Importa a planilha de contatos p/ o Atlas (portado do atlas-agro)."""
    def val(linha, campo):
        i = mapa.get(campo)
        if i is None or i == "" or int(i) >= len(linha):
            return None
        v = linha[int(i)]
        if v is None:
            return None
        v = str(v).strip()
        return v or None

    def num(linha, campo):
        v = val(linha, campo)
        if v is None:
            return None
        try:
            return float(v.replace(",", "."))
        except ValueError:
            return None

    criadas = ligadas = sem_fazenda = sem_nome = 0
    por_metodo = {}
    for n_linha, linha in enumerate(linhas):
        if n_linha and n_linha % 200 == 0:
            con.commit()  # solta a trava de escrita: o resto da equipe nao trava
        nome = val(linha, "nome")
        if not nome:
            sem_nome += 1
            continue
        doc = val(linha, "documento")
        tel = val(linha, "telefone")
        email = val(linha, "email")
        pessoa = None
        if doc:
            r = con.execute("SELECT id FROM pessoas WHERE documento = ?", [doc]).fetchone()
            if r:
                pessoa = r["id"]
        if pessoa is None:
            r = con.execute("SELECT id FROM pessoas WHERE UPPER(nome) = UPPER(?)", [nome]).fetchone()
            pessoa = r["id"] if r else None
        if pessoa is None:
            tipo = "juridica" if doc and len(re.sub(r"\D", "", doc)) == 14 else "fisica"
            cur = con.execute(
                "INSERT INTO pessoas (nome, tipo, documento, telefone, email) VALUES (?,?,?,?,?)",
                [nome, tipo, doc, tel, email])
            pessoa = cur.lastrowid
            criadas += 1
        else:
            con.execute(
                "UPDATE pessoas SET telefone = COALESCE(telefone, ?), "
                "email = COALESCE(email, ?), documento = COALESCE(documento, ?) WHERE id = ?",
                [tel, email, doc, pessoa])
        fid, metodo = _atlas_acha_fazenda(
            con, tem_contorno, car=val(linha, "codigo_car"), lat=num(linha, "latitude"),
            lon=num(linha, "longitude"), nome=val(linha, "fazenda"),
            municipio=val(linha, "municipio"))
        if fid:
            con.execute(
                "INSERT OR IGNORE INTO fazenda_pessoas "
                "(fazenda_id, pessoa_id, relacao_fundiaria, relacao_comercial) VALUES (?,?,?,?)",
                [fid, pessoa, val(linha, "relacao") or "proprietario",
                 categoria if categoria in ("cliente", "lead") else "nao_definida"])
            if categoria in ("cliente", "lead", "descartada"):
                con.execute("UPDATE fazendas SET categoria = ? WHERE id = ?", [categoria, fid])
            con.execute("INSERT INTO registros (fazenda_id, autor, texto) VALUES (?,?,?)",
                        [fid, origem, "Contato importado: %s%s" % (nome, (" · " + tel) if tel else "")])
            ligadas += 1
            por_metodo[metodo] = por_metodo.get(metodo, 0) + 1
        else:
            sem_fazenda += 1
    return {"ok": True, "linhas": len(linhas), "pessoas_novas": criadas,
            "ligadas_a_fazenda": ligadas, "sem_fazenda": sem_fazenda,
            "sem_nome": sem_nome, "como_ligou": por_metodo}

# Tipos de resultado de visita aceitos
RESULTADOS_VISITA = (
    "Interessado — avançar", "Em negociação", "Vai pensar", "Reagendar",
    "Sem interesse", "Fechou negócio", "Não encontrou o cliente", "Outro",
)

# Etapas do funil (ordem = ordem das colunas)
#   novo/triagem          -> fase do SDR (primeiro contato e qualificacao)
#   qualificado..ganho    -> funil de vendas (o "tipo" diz se e produtor ou
#                            prestador; cada tipo tem sua aba)
#   perdido               -> venda/lead que nao avancou (guardado p/ resgate)
STAGES = ["novo", "triagem", "qualificado", "decidindo", "negociacao", "proposta",
          "financiamento", "ganho", "desistiu", "perdido", "curioso"]

# Etapas do funil de vendas (ja qualificado). "decidindo" = cliente avaliando se
# vai adquirir o drone (antes de negociar); "financiamento" = proposta aceita,
# cliente aguardando a liberacao do recurso no banco.
SALES_STAGES = ["qualificado", "decidindo", "negociacao", "proposta",
                "financiamento", "ganho"]

# Painel de SERVIÇOS (pós-venda): depois de vender o drone (ganho), o cliente
# entra num funil separado de venda de serviços/pecas/manutencao.
SERVICO_STAGES = ["recebido_serv", "ofertado", "negociando_serv",
                  "proposta_serv", "vendido_serv", "recusado_serv"]

# Painel do CURSO (venda do curso de pilotagem): funil paralelo proprio, como
# Servicos. Entrada SO MANUAL — ninguem entra sozinho; o time marca na ficha.
CURSO_STAGES = ["interessado_curso", "ofertado_curso", "negociando_curso",
                "proposta_curso", "matriculado", "recusado_curso"]

# ---------------------------------------------------------------------------
# Etapas EDITAVEIS: o gestor renomeia qualquer coluna e cria/exclui etapas.
# As listas acima sao MUTADAS no lugar por recalcula_etapas() — todos os
# pontos do codigo que as referenciam enxergam a versao atual. Fixas: inicio
# (novo/triagem/qualificado) e fim (ganho/desistiu/perdido/curioso) de vendas.
# ---------------------------------------------------------------------------
STAGES_INICIO = ["novo", "triagem", "qualificado"]
STAGES_FIM = ["ganho", "desistiu", "perdido", "curioso"]
VENDAS_MEIO_PADRAO = ["decidindo", "negociacao", "proposta", "financiamento"]
SERVICO_PADRAO = list(SERVICO_STAGES)
CURSO_PADRAO = list(CURSO_STAGES)
MAX_ETAPAS_CUSTOM = 8

SERVICO_LABEL_PADRAO = {
    "recebido_serv": "🔧 Cliente com drone", "ofertado": "📞 Ofereci o serviço",
    "negociando_serv": "💬 Negociando", "proposta_serv": "📄 Proposta enviada",
    "vendido_serv": "🏆 Serviço vendido", "recusado_serv": "❌ Não quis",
}
CURSO_LABEL_PADRAO = {
    "interessado_curso": "🎓 Interessado", "ofertado_curso": "📞 Ofereci o curso",
    "negociando_curso": "💬 Negociando", "proposta_curso": "📄 Proposta enviada",
    "matriculado": "🏆 Matriculado", "recusado_curso": "❌ Não quis",
}


def _etapas_cfg():
    st = _db.get("settings", {}) or {}
    custom = st.get("etapas_custom") or {}
    rem = set(st.get("etapas_removidas") or [])
    return custom, rem


def _lista_funil(padrao, chave_custom):
    custom, rem = _etapas_cfg()
    lista = [s for s in padrao if s not in rem]
    lista += [e["key"] for e in (custom.get(chave_custom) or [])
              if isinstance(e, dict) and e.get("key")]
    return lista


def vendas_meio():
    """Etapas do MEIO do funil de vendas (entre qualificado e ganho) — as
    unicas de vendas que podem ser excluidas."""
    return _lista_funil(VENDAS_MEIO_PADRAO, "vendas")


def recalcula_etapas():
    """Reconstroi as listas de etapas a partir da configuracao salva."""
    meio = vendas_meio()
    STAGES[:] = STAGES_INICIO + meio + STAGES_FIM
    SALES_STAGES[:] = ["qualificado"] + meio + ["ganho"]
    ETAPAS_DE_VENDA[:] = meio + ["ganho"]
    SERVICO_STAGES[:] = _lista_funil(SERVICO_PADRAO, "servicos") or [SERVICO_PADRAO[0]]
    CURSO_STAGES[:] = _lista_funil(CURSO_PADRAO, "curso") or [CURSO_PADRAO[0]]


def rotulo_etapa(key):
    """Nome de exibicao da etapa: apelido do gestor > padrao > a propria chave."""
    if not key:
        return "—"
    st = _db.get("settings", {}) or {}
    r = (st.get("rotulos") or {}).get(key)
    if r:
        return r
    if key in STATUS_LABEL:
        return STATUS_LABEL[key]
    if key in SERVICO_LABEL_PADRAO:
        return SERVICO_LABEL_PADRAO[key]
    if key in CURSO_LABEL_PADRAO:
        return CURSO_LABEL_PADRAO[key]
    custom, _ = _etapas_cfg()
    for lista in custom.values():
        for e in (lista or []):
            if isinstance(e, dict) and e.get("key") == key:
                return e.get("label") or key
    return key


def etapas_publico():
    """Listas resolvidas (chave + nome + se e fixa) para o painel e o quadro.

    "renomeada" = o GESTOR personalizou o nome (ou a etapa foi criada por ele).
    O quadro so troca o rotulo quando renomeada — os nomes padrao da interface
    (com emojis e dicas) nao podem ser sobrescritos pelos textos internos."""
    meio = set(vendas_meio())
    st = _db.get("settings", {}) or {}
    rot = st.get("rotulos") or {}
    custom_keys = {e.get("key") for lista in (st.get("etapas_custom") or {}).values()
                   for e in (lista or []) if isinstance(e, dict)}

    def linhas(keys, todas_removiveis):
        return [{"key": k, "label": rotulo_etapa(k),
                 "renomeada": k in rot or k in custom_keys,
                 "fixa": (not todas_removiveis) and k not in meio} for k in keys]
    return {"vendas": linhas(STAGES, False),
            "servicos": linhas(SERVICO_STAGES, True),
            "curso": linhas(CURSO_STAGES, True)}

# Papeis da equipe (raias dos funis)
PAPEIS = ("sdr", "vendedor")

# Papeis de usuario (login):
#   admin    -> tudo + gerencia usuarios e niveis de acesso
#   gerente  -> tudo, exceto gerenciar usuarios
#   vendedor -> so leads do funil de Vendas que sao dele ou sem responsavel
#   sdr      -> so os leads dele (campo sdr)
PAPEIS_USUARIO = ("admin", "gerente", "vendedor", "sdr")

SESSAO_DIAS = 30

EDITABLE = {
    "nome", "telefone", "email", "regiao", "area_cultivada", "produto", "itens", "valor",
    "cargo", "decisor", "decisor_cargo", "formas_pagamento",
    "vendedor", "sdr", "responsavel", "tipo", "origem_canal", "campanha",
    "campanha_id", "utm_source", "utm_medium", "utm_campaign", "utm_content",
    "utm_term", "status", "observacoes", "lat", "lng", "recuperacao",
    "em_servicos", "status_servico", "valor_servico",
    "em_curso", "status_curso", "valor_curso",
}

# Canais aceitos para campanhas cadastradas
CANAIS = ("Meta", "Google", "WhatsApp", "TikTok", "Indicação", "Outro")

# Linha de produtos da empresa (lista fechada no formulario)
PRODUTOS = ("T25P", "T70P", "T55", "T100", "Peças e Serviços")

# Formas de pagamento aceitas (um lead pode combinar varias = pagamento misto)
PAGAMENTOS = (
    "À vista", "Financiamento", "Cartão BNDES", "Cartão de crédito",
    "Permuta / Troca", "Consórcio", "CPR", "Boleto / Parcelado", "Outro",
)
# Formas que aceitam entrada + parcelamento (as demais zeram esses campos)
PARCELAVEIS = frozenset({
    "Financiamento", "Cartão BNDES", "Cartão de crédito",
    "Consórcio", "CPR", "Boleto / Parcelado",
})

# Etapas que exigem telefone + e-mail preenchidos (nota fiscal / fechamento)
STAGES_EXIGEM_CONTATO = ("proposta", "financiamento", "ganho")

# Municipios oficiais (IBGE), carregados de public/cidades.json no boot.
# _CIDADES_CANON mapeia minusculo -> forma canonica "Nome - UF".
_CIDADES_CANON = {}
# Mesorregioes do IBGE (Goias): "Nome - GO" -> mesorregiao. De public/mesorregioes.json.
_MESO = {}
MESORREGIOES = ["Noroeste Goiano", "Norte Goiano", "Centro Goiano",
                "Leste Goiano", "Sul Goiano"]


def load_cidades():
    try:
        with open(os.path.join(PUBLIC_DIR, "cidades.json"), "r", encoding="utf-8") as f:
            for nome in json.load(f):
                _CIDADES_CANON[nome.lower()] = nome
        print("  %d cidades carregadas (IBGE)" % len(_CIDADES_CANON))
    except Exception as e:
        print("AVISO: nao carregou cidades.json (%s) — regiao fica sem validacao" % e)
    try:
        with open(os.path.join(PUBLIC_DIR, "mesorregioes.json"), "r", encoding="utf-8") as f:
            _MESO.update(json.load(f))
        print("  %d municipios com mesorregiao (GO)" % len(_MESO))
    except Exception as e:
        print("AVISO: nao carregou mesorregioes.json (%s)" % e)


def meso_da_regiao(regiao):
    """Mesorregiao (GO) da cidade do lead, ou None."""
    return _MESO.get(str(regiao or "").strip())


def canon_cidade(valor):
    """Retorna a forma canonica 'Nome - UF' ou None se nao reconhecida."""
    if not valor or not _CIDADES_CANON:
        return None
    return _CIDADES_CANON.get(str(valor).strip().lower())

MIME = {
    ".html": "text/html; charset=utf-8",
    ".js": "text/javascript; charset=utf-8",
    ".css": "text/css; charset=utf-8",
    ".json": "application/json; charset=utf-8",
    ".webmanifest": "application/manifest+json; charset=utf-8",
    ".svg": "image/svg+xml",
    ".ico": "image/x-icon",
    ".png": "image/png",
    ".webp": "image/webp",
    ".jpg": "image/jpeg",
    ".jpeg": "image/jpeg",
}

# ---------------------------------------------------------------------------
# Camada de dados (JSON em arquivo + lock para acesso concorrente)
# ---------------------------------------------------------------------------
_lock = threading.Lock()
# leads com envio de saudacao (Chatwoot) em andamento — trava de duplo-clique
_cw_em_voo = set()
_webhook_ultimo = None  # ultimo evento recebido do Chatwoot (desde o boot)

# Trava de forca-bruta no login: depois de muitas senhas erradas seguidas para
# um MESMO login, bloqueia novas tentativas por um tempo (chave = login, para
# valer mesmo que o atacante troque de IP).
LOGIN_MAX_FALHAS = 10
LOGIN_BLOQUEIO_S = int(os.environ.get("LOGIN_BLOQUEIO_S") or 600)  # 10 min
_login_falhas = {}  # login -> {"n": tentativas seguidas, "ate": bloqueado ate (epoch)}
# presenca "online": user_id -> ultima vez visto (ISO). So em memoria (efemero,
# nao vai pro disco); atualizado a cada requisicao autenticada.
_online = {}
ONLINE_LIMIAR_S = 100  # visto nos ultimos N segundos = online
# users: pessoas com login (a equipe: admin/gerente/vendedor/sdr)
# rr_sdr: indice do rodizio de SDRs | campaigns: campanhas | settings: config
# sessions: sessoes de login ativas (token -> user_id/validade)
_db = {"leads": [], "users": [], "rr_sdr": 0, "campaigns": [], "settings": {}, "sessions": {}}


def now_iso():
    return datetime.now(timezone.utc).isoformat()


BRT = timezone(timedelta(hours=-3))  # horario de Brasilia (sem horario de verao)


def dia_brt(iso):
    """Converte um timestamp ISO (UTC) para a data 'AAAA-MM-DD' em Brasilia."""
    if not iso:
        return None
    try:
        dt = datetime.fromisoformat(str(iso))
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return dt.astimezone(BRT).strftime("%Y-%m-%d")
    except (ValueError, TypeError):
        return None


def _parse_iso(iso):
    """ISO -> datetime aware (UTC). None se invalido."""
    if not iso:
        return None
    try:
        dt = datetime.fromisoformat(str(iso))
    except (ValueError, TypeError):
        return None
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt


def ts_offline(valor):
    """Horario de um registro feito OFFLINE (visita/nota): usa o instante do
    evento enviado pelo cliente, se for valido e plausivel (nao no futuro, nao
    mais de 30 dias atras). Caso contrario, usa agora. Assim uma visita feita as
    08h e sincronizada as 18h fica com 08h, sem confiar cegamente no relogio."""
    t = _parse_iso(valor)
    if not t:
        return now_iso()
    agora = datetime.now(timezone.utc)
    if t > agora + timedelta(minutes=1) or t < agora - timedelta(days=30):
        return now_iso()
    return t.isoformat()


HEAT_DIAS = 7        # janela do "termometro" (mesma regra do front)
HEAT_QUENTE = 3      # nº de atualizacoes recentes p/ virar "quente"


def heat_nivel(lead):
    """Termometro do lead: "" | "recente" (👍) | "quente" (🔥) — conta as
    entradas de historico dos ultimos HEAT_DIAS dias (igual ao front). A entrada
    automatica de criacao (tipo "novo") NAO conta: criar nao e engajamento."""
    hist = lead.get("historico") or []
    if not hist:
        return ""
    limite = datetime.now(timezone.utc) - timedelta(days=HEAT_DIAS)
    n = 0
    for h in hist:
        if h.get("tipo") == "novo":
            continue
        t = _parse_iso(h.get("data"))
        if t and t >= limite:
            n += 1
    if n >= HEAT_QUENTE:
        return "quente"
    if n >= 1:
        return "recente"
    return ""


def precisa_retorno(lead, cadencia_dias):
    """True quando o lead ativo passou do prazo sem contato e merece o alerta de
    retorno. Leads quentes sao cobrados na metade do prazo. Um lead que ja espera
    o registro da resposta (aguardando_resposta) NAO gera alerta de retorno."""
    if lead.get("status") in ("ganho", "perdido", "desistiu", "curioso"):
        return False
    if lead.get("aguardando_resposta") or lead.get("cliente_respondeu"):
        return False  # ja existe um aviso mais especifico no card
    t = _parse_iso(lead.get("updated_at") or lead.get("created_at"))
    if not t:
        return False
    idade_dias = (datetime.now(timezone.utc) - t).total_seconds() / 86400
    metade = max(1, (cadencia_dias + 1) // 2)
    # checa a idade (barato) antes do termometro (varre o historico): a maioria
    # dos leads cai fora da "banda do meio" e nem calcula o heat.
    if idade_dias < metade:
        return False                       # nem um lead quente dispara antes da metade
    if idade_dias >= cadencia_dias:
        return True                        # ate frio dispara; termometro nao muda nada
    return heat_nivel(lead) == "quente"    # banda do meio: so o quente (🔥) dispara


def tarefa_cobrando(lead):
    """True se ha tarefa ABERTA com prazo para hoje ou ja vencido (dia BRT)."""
    hoje = dia_brt(now_iso())
    for t in (lead.get("tarefas") or []):
        if not t.get("feita") and t.get("prazo") and str(t["prazo"])[:10] <= hoje:
            return True
    return False


def cadencia_dias_cfg():
    """Prazo (dias) do alerta de retorno, configuravel; padrao 2, limites 1..30."""
    try:
        v = int(_db.get("settings", {}).get("cadencia_dias") or 2)
    except (TypeError, ValueError):
        v = 2
    return max(1, min(30, v))


def resposta_horas_cfg():
    """Prazo (horas) p/ a resposta virar urgente; padrao 3, limites 1..168."""
    try:
        v = int(_db.get("settings", {}).get("resposta_horas") or 3)
    except (TypeError, ValueError):
        v = 3
    return max(1, min(168, v))


def new_id():
    return base64.urlsafe_b64encode(secrets.token_bytes(9)).decode().rstrip("=")


def load_db():
    global _db
    try:
        if os.path.exists(DB_FILE):
            with open(DB_FILE, "r", encoding="utf-8") as f:
                data = json.load(f)
            if not isinstance(data.get("leads"), list):
                data["leads"] = []
            if not isinstance(data.get("users"), list):
                data["users"] = []
            if not isinstance(data.get("rr_sdr"), int):
                data["rr_sdr"] = 0
            if not isinstance(data.get("campaigns"), list):
                data["campaigns"] = []
            if not isinstance(data.get("settings"), dict):
                data["settings"] = {}
            if not isinstance(data.get("sessions"), dict):
                data["sessions"] = {}
            # Migracao: a antiga "equipe" (members, sem login) vira usuarios com
            # senha pendente — o admin define a senha de cada um.
            if data.get("members") and not data["users"]:
                logins_usados = set()
                for m in data.pop("members"):
                    login = _slug_login(m.get("nome", ""))
                    while login in logins_usados:  # desambigua nomes parecidos
                        login = _slug_login(m.get("nome", "")) + secrets.token_hex(2)
                    logins_usados.add(login)
                    data["users"].append({
                        "id": m.get("id") or new_id(), "nome": m.get("nome", ""),
                        "login": login, "salt": "", "senha_hash": "",
                        "papel": m.get("papel", "sdr"), "ativo": m.get("ativo", True),
                    })
                print("  equipe antiga migrada para usuarios (defina as senhas no painel)")
            data.pop("members", None)
            # Migracao das etapas antigas:
            #   status "produtor" (recebido no funil) -> "qualificado"
            #   status "prestador" (fora do perfil)   -> qualificado + tipo prestador
            for l in data["leads"]:
                if l.get("status") == "produtor":
                    l["status"] = "qualificado"
                    l.setdefault("tipo", "") or l.__setitem__("tipo", l.get("tipo") or "produtor")
                elif l.get("status") == "prestador":
                    l["status"] = "qualificado"
                    l["tipo"] = "prestador"
                l.setdefault("qualificado_em", None)
                if l.get("tipo") and not l.get("qualificado_em"):
                    l["qualificado_em"] = l.get("updated_at") or l.get("created_at")
                # leads antigos que ja tem vendedor contam como atendidos
                l.setdefault("atendido_em", None)
                if l.get("vendedor") and not l.get("atendido_em"):
                    l["atendido_em"] = l.get("updated_at") or l.get("created_at")
                # congela a data de ganho/perda dos leads antigos na data atual
                # (updated_at) para que edicoes/notas futuras nao movam a vitoria
                # ou perda de dia no relatorio.
                l.setdefault("ganho_em", None)
                if l.get("status") == "ganho" and not l.get("ganho_em"):
                    l["ganho_em"] = l.get("updated_at") or l.get("created_at")
                l.setdefault("perdido_em", None)
                if l.get("status") == "perdido" and not l.get("perdido_em"):
                    l["perdido_em"] = l.get("updated_at") or l.get("created_at")
                l.setdefault("desistiu_em", None)
                if l.get("status") == "desistiu" and not l.get("desistiu_em"):
                    l["desistiu_em"] = l.get("updated_at") or l.get("created_at")
                if not isinstance(l.get("formas_pagamento"), list):
                    l["formas_pagamento"] = []
                if not isinstance(l.get("itens"), list):
                    # lead antigo tinha um produto unico -> vira 1 item do pedido
                    l["itens"] = itens_de_produto(l.get("produto"))
                # leads que ja existiam ANTES desta versao sao o lote de
                # recuperacao (clientes do passado); os novos nascem como atuais.
                l.setdefault("recuperacao", True)
                l.setdefault("em_servicos", False)
                l.setdefault("status_servico", "")
                l.setdefault("valor_servico", 0)
                l.setdefault("em_curso", False)
                l.setdefault("status_curso", "")
                l.setdefault("valor_curso", 0)
                if not isinstance(l.get("visitas"), list):
                    l["visitas"] = []
                if not isinstance(l.get("historico"), list):
                    l["historico"] = []
                l.setdefault("aguardando_resposta", None)
                l.setdefault("cliente_respondeu", None)
                if not isinstance(l.get("tarefas"), list):
                    l["tarefas"] = []
                if not isinstance(l.get("chatwoot_msgs_vistas"), list):
                    l["chatwoot_msgs_vistas"] = []
                # marcadores antigos eram ints (uma instancia so): viram o
                # formato novo por instancia ("123" = drones)
                l["chatwoot_msgs_vistas"] = [str(x) for x in l["chatwoot_msgs_vistas"]]
                l.setdefault("chatwoot_origem", "")
            _db = data
    except Exception as e:
        # Arquivo ilegivel (queda de energia, edicao manual): NUNCA sobrescrever.
        # Renomeia para .corrompido-<ts> e recomeca vazio; o original fica salvo.
        backup = "%s.corrompido-%d" % (DB_FILE, int(time.time()))
        try:
            os.replace(DB_FILE, backup)
            print("AVISO: banco ilegivel (%s). Copia guardada em: %s" % (e, backup))
        except OSError:
            print("AVISO: banco ilegivel e nao foi possivel criar backup:", e)
        _db = {"leads": [], "users": [], "rr_sdr": 0, "campaigns": [],
               "settings": {}, "sessions": {}}


def save_db():
    """Escrita atomica: grava em .tmp e renomeia."""
    try:
        os.makedirs(DATA_DIR, exist_ok=True)
        tmp = DB_FILE + ".tmp"
        with open(tmp, "w", encoding="utf-8") as f:
            json.dump(_db, f, ensure_ascii=False, indent=2)
        os.replace(tmp, DB_FILE)
    except Exception as e:
        print("Falha ao salvar o banco:", e)


def make_lead(partial=None):
    lead = {
        "id": new_id(),
        "source": "manual",
        "chatwoot_conversation_id": None,
        "chatwoot_contact_id": None,
        "nome": "",
        "telefone": "",
        "email": "",
        "regiao": "",
        "area_cultivada": "",
        "produto": "",       # espelho (nomes juntos) dos itens — busca/legado
        "itens": [],         # drones do pedido: lista de {produto, qtd}
        "valor": 0,
        "cargo": "",          # cargo de QUEM entrou em contato
        "decisor": "",        # quem decide/paga (vazio = o proprio contato)
        "decisor_cargo": "",  # cargo do decisor, quando for outra pessoa
        "formas_pagamento": [],  # lista de {tipo, valor} (misto = varias)
        "visitas": [],           # visitas de campo: {id, data, visitante, resultado, obs, foto}
        "historico": [],         # linha do tempo de atualizacoes: {data, autor, itens}
        "tipo": "",        # ""=nao classificado | "produtor" | "prestador"
        "sdr": "",         # SDR que recebeu/qualificou
        "vendedor": "",    # vendedor responsavel apos qualificar
        "responsavel": "", # dono atual do lead (SDR na triagem, vendedor depois)
        "origem_canal": "",
        "campanha": "",     # nome da campanha (texto exibido)
        "campanha_id": "",  # vinculo com uma campanha cadastrada
        "meta_ad_id": "",   # id do anuncio (Meta clique-pro-WhatsApp)
        "ctwa_clid": "",    # id de clique do anuncio (Meta)
        "utm_source": "",
        "utm_medium": "",
        "utm_campaign": "",
        "utm_content": "",
        "utm_term": "",
        "status": "novo",
        "recuperacao": False,  # True = cliente antigo em recuperacao (fica na aba
                               # "Recuperacao", fora do funil dos leads NOVOS)
        "em_servicos": False,  # True = tambem esta no painel de Servicos (pos-venda)
        "status_servico": "",  # etapa no funil de Servicos (SERVICO_STAGES)
        "valor_servico": 0,    # valor do negocio de servicos (separado do drone)
        "em_curso": False,     # True = tambem esta no painel do Curso (entrada manual)
        "status_curso": "",    # etapa no funil do Curso (CURSO_STAGES)
        "valor_curso": 0,      # valor do curso (separado do drone e dos servicos)
        "observacoes": "",
        "lat": None,   # localizacao exata da fazenda (ajustada no mapa);
        "lng": None,   # None = usar o centro da cidade (regiao) como aproximacao
        "last_message": "",
        "qualificado_em": None,  # quando o SDR classificou (entra no funil de vendas)
        "atendido_em": None,     # quando um vendedor assumiu (mede a agilidade)
        "ganho_em": None,        # data em que o negocio foi GANHO (fixa; p/ relatorio)
        "perdido_em": None,      # data em que o negocio foi PERDIDO (fixa; p/ relatorio)
        "desistiu_em": None,     # data em que o cliente DESISTIU da compra (fixa)
        "tarefas": [],     # tarefas do cliente: {id, texto, prazo, criada_por,
                           # criada_em, feita, feita_em, feita_por}
        "aguardando_resposta": None,  # ISO do contato por WhatsApp que ainda espera
                                      # o vendedor REGISTRAR o que o cliente respondeu
        "cliente_respondeu": None,    # ISO da ultima mensagem RECEBIDA do cliente que
                                      # ainda espera a equipe responder (aviso verde)
        "chatwoot_msgs_vistas": [],   # ids de mensagem ja processados do webhook
                                      # (reentregas nao registram de novo)
        "chatwoot_origem": "",        # em qual Chatwoot vive a conversa:
                                      # "" = venda de drones | "curso"
        "created_at": now_iso(),  # data/hora de ENTRADA do lead
        "updated_at": now_iso(),
    }
    if partial:
        lead.update(partial)
    return lead


def _num_pos(x, teto=None):
    """Converte para float >= 0 e finito; retorna 0.0 se invalido."""
    try:
        v = float(x) if x not in ("", None) else 0.0
    except (TypeError, ValueError):
        return 0.0
    if not math.isfinite(v) or v < 0:
        return 0.0
    if teto is not None and v > teto:
        v = teto
    return v


def salva_foto_visita(data_url, visita_id):
    """Recebe um data URL (base64) de imagem, valida e grava em fotos/<id>.jpg.
    Retorna o nome do arquivo ou None. Levanta ValueError se invalida/grande."""
    if not data_url or not isinstance(data_url, str):
        return None
    m = re.match(r"^data:image/(jpeg|jpg|png|webp);base64,(.+)$", data_url, re.DOTALL)
    if not m:
        raise ValueError("Formato de imagem inválido")
    try:
        raw = base64.b64decode(m.group(2), validate=True)
    except Exception:
        raise ValueError("Imagem corrompida")
    if len(raw) > 8 * 1024 * 1024:
        raise ValueError("Foto muito grande (máx. 8 MB)")
    # confere a assinatura do arquivo (nao confia so na extensao)
    ok = raw[:3] == b"\xff\xd8\xff" or raw[:8] == b"\x89PNG\r\n\x1a\n" or raw[:4] == b"RIFF"
    if not ok:
        raise ValueError("Arquivo não é uma imagem válida")
    ext = "png" if raw[:8] == b"\x89PNG\r\n\x1a\n" else ("webp" if raw[:4] == b"RIFF" else "jpg")
    nome = "%s.%s" % (visita_id, ext)
    os.makedirs(FOTOS_DIR, exist_ok=True)
    with open(os.path.join(FOTOS_DIR, nome), "wb") as f:
        f.write(raw)
    return nome


def sanitiza_pagamentos(value):
    """Aceita uma lista de {tipo, valor, entrada, parcelas}; descarta lixo/NaN,
    remove tipos repetidos (pagamento misto = formas distintas) e zera
    entrada/parcelas nas formas que nao sao parcelaveis."""
    if not isinstance(value, list):
        return []
    out = []
    vistos = set()
    for item in value:
        if not isinstance(item, dict):
            continue
        tipo = str(item.get("tipo") or "").strip()
        if tipo not in PAGAMENTOS or tipo in vistos:
            continue
        vistos.add(tipo)
        parcelavel = tipo in PARCELAVEIS
        out.append({
            "tipo": tipo,
            "valor": round(_num_pos(item.get("valor")), 2),
            "entrada": round(_num_pos(item.get("entrada")), 2) if parcelavel else 0.0,
            "parcelas": int(_num_pos(item.get("parcelas"), teto=360)) if parcelavel else 0,
        })
    return out


def sanitiza_itens(value):
    """Aceita uma lista de {produto, qtd} (os drones do pedido); mantem so
    produtos validos, SOMA quantidades de um produto repetido e limita 1..99."""
    if not isinstance(value, list):
        return []
    somas, ordem = {}, []
    for item in value:
        if not isinstance(item, dict):
            continue
        prod = str(item.get("produto") or "").strip()
        if prod not in PRODUTOS:
            continue
        try:
            qf = float(item.get("qtd") or 1)
            q = int(qf) if math.isfinite(qf) else 1  # inf/NaN (via API) não podem estourar
        except (TypeError, ValueError):
            q = 1
        q = max(1, min(99, q))
        if prod not in somas:
            somas[prod] = 0
            ordem.append(prod)
        somas[prod] = min(99, somas[prod] + q)
    return [{"produto": p, "qtd": somas[p]} for p in ordem]


def resumo_produtos(itens):
    """Nomes dos produtos do pedido, juntos — usado na busca e como espelho do
    campo legado `produto`. Ex.: [T25P x2, T70P] -> 'T25P, T70P'."""
    return ", ".join(it["produto"] for it in (itens or []) if it.get("produto"))


def itens_de_produto(produto):
    """Um produto unico (webhook/importacao/legado) vira uma lista de 1 item."""
    p = str(produto or "").strip()
    return [{"produto": p, "qtd": 1}] if p in PRODUTOS else []


# ---------------------------------------------------------------------------
# Historico (linha do tempo de atualizacoes do lead)
# ---------------------------------------------------------------------------
STATUS_LABEL = {
    "novo": "Novo lead", "triagem": "Em triagem", "qualificado": "Qualificado",
    "decidindo": "Decidindo", "negociacao": "Em negociação", "proposta": "Proposta enviada",
    "financiamento": "Aguardando financiamento",
    "ganho": "Fechado (ganho)", "desistiu": "Desistiu da compra", "perdido": "Perdido p/ concorrente",
    "curioso": "Só curioso (sem perspectiva)",
}
HIST_LABEL = {
    "status": "Etapa", "vendedor": "Vendedor", "sdr": "SDR", "tipo": "Classificação",
    "valor": "Valor", "produto": "Drone", "regiao": "Cidade", "nome": "Nome",
    "telefone": "Telefone", "email": "E-mail", "area_cultivada": "Área cultivada",
    "cargo": "Cargo do contato", "decisor": "Decisor", "decisor_cargo": "Cargo do decisor",
    "campanha": "Campanha", "observacoes": "Observações", "formas_pagamento": "Forma de pagamento",
    "origem_canal": "Canal", "lat": "Localização", "lng": "Localização",
}


def registra_hist(lead, autor, itens, papel=None, tipo=None, op_id=None, quando=None):
    """Anexa uma entrada na linha do tempo do lead (mantem as ultimas 300).

    papel = cargo de quem fez (admin/gerente/vendedor/sdr) para aparecer no
    painel; tipo = 'nota' quando for uma atualizacao escrita a mao; op_id =
    id do registro offline (para nao duplicar em reenvios); quando = horario do
    evento (registros offline) — padrao agora."""
    itens = [i for i in itens if i]
    if not itens:
        return
    entrada = {"data": quando or now_iso(), "autor": autor or "Sistema", "itens": itens}
    if papel:
        entrada["papel"] = papel
    if tipo:
        entrada["tipo"] = tipo
    if op_id:
        entrada["op_id"] = op_id
    lead.setdefault("historico", []).append(entrada)
    if len(lead["historico"]) > 300:
        lead["historico"] = lead["historico"][-300:]


def descreve_mudancas(antes, depois, campos):
    """Gera frases legiveis do que mudou (so os campos informados)."""
    itens = []
    vistos = set()
    for k in campos:
        if k in ("lat", "lng"):
            k = "localizacao"
        if k in vistos:
            continue
        vistos.add(k)
        if k == "localizacao":
            if antes.get("lat") != depois.get("lat") or antes.get("lng") != depois.get("lng"):
                itens.append("📍 Localização da fazenda atualizada")
            continue
        a, d = antes.get(k), depois.get(k)
        if a == d:
            continue
        if k == "status":
            itens.append("➡️ Etapa: %s → %s" % (rotulo_etapa(a), rotulo_etapa(d)))
        elif k == "tipo":
            itens.append("✅ Classificado como %s" % (d or "—"))
        elif k == "valor":
            itens.append("💰 Valor: R$ %s" % ("{:,.0f}".format(float(d or 0)).replace(",", ".")))
        elif k == "vendedor":
            itens.append("👤 Vendedor: %s" % (d or "(removido)"))
        elif k == "sdr":
            itens.append("📞 SDR: %s" % (d or "(removido)"))
        elif k == "itens":
            itens.append("📦 Pedido (drones) atualizado")
        elif k == "status_servico":
            itens.append("🔧 Serviços — etapa: %s" % rotulo_etapa(d))
        elif k == "valor_servico":
            itens.append("🔧 Serviços — valor: R$ %s" % ("{:,.0f}".format(float(d or 0)).replace(",", ".")))
        elif k == "em_curso":
            itens.append("🎓 " + ("Entrou no painel do Curso" if d else "Saiu do painel do Curso"))
        elif k == "status_curso":
            itens.append("🎓 Curso — etapa: %s" % rotulo_etapa(d))
        elif k == "valor_curso":
            itens.append("🎓 Curso — valor: R$ %s" % ("{:,.0f}".format(float(d or 0)).replace(",", ".")))
        elif k == "formas_pagamento":
            itens.append("💳 Forma de pagamento atualizada")
        elif k == "observacoes":
            itens.append("📝 Observações atualizadas")
        elif k in HIST_LABEL:
            itens.append("✏️ %s: %s" % (HIST_LABEL[k], str(d)[:60] if d else "(vazio)"))
    return itens


def apply_updates(lead, updates):
    """Aplica edicoes manuais. Levanta ValueError com mensagem amigavel quando
    uma regra de negocio e violada (as rotas devolvem 400 com essa mensagem)."""
    status_antes = lead.get("status")
    for key, value in updates.items():
        if key not in EDITABLE:
            continue
        if key == "status" and value not in STAGES:
            continue
        if key == "tipo" and value not in ("", "produtor", "prestador", "pecuarista", "curso"):
            continue
        if key in ("telefone", "email") and not str(value or "").strip() and str(lead.get(key) or "").strip():
            campo = "Telefone" if key == "telefone" else "E-mail"
            raise ValueError("%s é obrigatório e não pode ficar vazio" % campo)
        if key == "telefone" and str(value or "").strip() and len(norm_phone(value)) < 8:
            raise ValueError("Telefone inválido — informe DDD e número")
        if key == "produto":
            # produto unico (import/webhook/legado): vira 1 item; itens e a fonte
            if value and value not in PRODUTOS:
                raise ValueError("Produto inválido — escolha um da lista")
            lead["itens"] = itens_de_produto(value)
            lead["produto"] = value if value in PRODUTOS else ""
            continue
        if key == "itens":
            # os drones do pedido; `produto` vira o espelho (nomes juntos) p/ busca
            lead["itens"] = sanitiza_itens(value)
            lead["produto"] = resumo_produtos(lead["itens"])
            continue
        if key == "formas_pagamento":
            lead["formas_pagamento"] = sanitiza_pagamentos(value)
            continue
        if key == "recuperacao":
            lead["recuperacao"] = bool(value)
            continue
        if key == "em_servicos":
            lead["em_servicos"] = bool(value)
            if lead["em_servicos"] and not lead.get("status_servico"):
                lead["status_servico"] = SERVICO_STAGES[0]  # entrou -> primeira etapa
            continue
        if key == "status_servico":
            if value in SERVICO_STAGES:
                lead["status_servico"] = value
                lead["em_servicos"] = True  # ter etapa de servico = estar no painel
            elif value == "":
                lead["status_servico"] = ""
            continue
        if key == "valor_servico":
            try:
                vs = float(value) if value not in ("", None) else 0.0
            except (TypeError, ValueError):
                vs = 0.0
            lead["valor_servico"] = vs if math.isfinite(vs) and vs >= 0 else 0.0
            continue
        if key == "em_curso":
            lead["em_curso"] = bool(value)
            if lead["em_curso"] and not lead.get("status_curso"):
                lead["status_curso"] = CURSO_STAGES[0]  # entrou -> primeira etapa
            continue
        if key == "status_curso":
            if value in CURSO_STAGES:
                lead["status_curso"] = value
                lead["em_curso"] = True  # ter etapa de curso = estar no painel
            elif value == "":
                lead["status_curso"] = ""
            continue
        if key == "valor_curso":
            try:
                vc = float(value) if value not in ("", None) else 0.0
            except (TypeError, ValueError):
                vc = 0.0
            lead["valor_curso"] = vc if math.isfinite(vc) and vc >= 0 else 0.0
            continue
        if key == "regiao" and value and _CIDADES_CANON:
            canon = canon_cidade(value)
            if not canon:
                raise ValueError("Cidade não reconhecida — escolha uma da lista (ex.: Rio Verde - GO)")
            value = canon  # padroniza a grafia
        if key == "valor":
            try:
                v = float(value) if value not in ("", None) else 0.0
            except (TypeError, ValueError):
                v = 0.0
            # NaN/Infinity passam no float() mas quebrariam o JSON do banco
            lead["valor"] = v if math.isfinite(v) else 0.0
            continue
        if key in ("lat", "lng"):
            if value in ("", None):
                lead[key] = None  # volta a usar o centro da cidade
                continue
            try:
                v = float(value)
            except (TypeError, ValueError):
                raise ValueError("Coordenada inválida")
            limite = 90 if key == "lat" else 180
            if not math.isfinite(v) or abs(v) > limite:
                raise ValueError("Coordenada inválida")
            lead[key] = round(v, 6)
            continue
        lead[key] = value

    # "Curioso" e um lead de nivel SDR (sem perspectiva de compra): sai do funil
    # de vendas, entao zera o tipo/qualificacao por QUALQUER caminho (arraste,
    # seletor de etapa do modal ou PATCH) — senao ficaria contado como produtor.
    if lead.get("status") == "curioso":
        lead["tipo"] = ""
        lead["qualificado_em"] = None
    # Consistencia: entrar no funil de vendas sem tipo assume "produtor";
    # ao ganhar o tipo pela primeira vez, marca a data de qualificacao.
    if lead.get("status") in SALES_STAGES and not lead.get("tipo"):
        lead["tipo"] = "produtor"
    if lead.get("tipo") and not lead.get("qualificado_em"):
        lead["qualificado_em"] = now_iso()
    # CLASSIFICADO como Curso: o funil de venda dele e o do curso — entra
    # sozinho no painel 🎓 (o invariante la embaixo poe a primeira etapa).
    # So no ATO da classificacao: tirar do painel depois continua manual.
    if updates.get("tipo") == "curso" and not lead.get("em_curso"):
        lead["em_curso"] = True
    # Primeiro momento em que um vendedor assume o lead = "atendimento".
    if lead.get("vendedor") and not lead.get("atendido_em"):
        lead["atendido_em"] = now_iso()
    # Carimba a data de ganho/perda no momento da TRANSICAO (nao em cada save),
    # para o relatorio nao depender de updated_at (que muda com nota/visita/edicao).
    if lead.get("status") == "ganho" and status_antes != "ganho":
        lead["ganho_em"] = now_iso()
        # vendeu o drone -> entra AUTOMATICAMENTE no painel de Servicos (pos-venda),
        # sem sair do funil de drones (o ganho continua contando).
        if not lead.get("em_servicos"):
            lead["em_servicos"] = True
            lead["status_servico"] = SERVICO_STAGES[0]
    if lead.get("status") == "perdido" and status_antes != "perdido":
        lead["perdido_em"] = now_iso()
    if lead.get("status") == "desistiu" and status_antes != "desistiu":
        lead["desistiu_em"] = now_iso()
    # Lead encerrado nao deve mais cobrar "registre a resposta" nem "responda".
    if lead.get("status") in ("ganho", "perdido", "desistiu", "curioso"):
        lead["aguardando_resposta"] = None
        lead["cliente_respondeu"] = None

    # Nota fiscal exige contato completo: barra a MUDANCA para essas etapas
    if updates.get("status") in STAGES_EXIGEM_CONTATO and (
            not str(lead.get("telefone") or "").strip() or not str(lead.get("email") or "").strip()):
        # lista gerada das proprias etapas: nao desatualiza ao criar uma nova
        etapas = "/".join('"%s"' % rotulo_etapa(s) for s in STAGES_EXIGEM_CONTATO)
        raise ValueError("Para mover para %s, preencha telefone e e-mail do lead (nota fiscal)" % etapas)

    # Ao definir um vendedor, a posse do lead passa para ele (a menos que o
    # responsavel tenha sido informado explicitamente na mesma atualizacao).
    if updates.get("vendedor") and "responsavel" not in updates:
        lead["responsavel"] = updates["vendedor"]

    # Vinculo manual com campanha cadastrada: valida e espelha o nome no lead.
    if "campanha_id" in updates:
        camp = next((c for c in _db.get("campaigns", []) if c["id"] == updates["campanha_id"]), None)
        if camp:
            lead["campanha_id"] = camp["id"]
            lead["campanha"] = camp["nome"]
            if not lead.get("origem_canal"):
                lead["origem_canal"] = camp.get("canal", "")
        else:
            lead["campanha_id"] = ""

    # Invariante do painel de Serviços: dentro do painel SEMPRE há uma etapa
    # válida (senão o card ficaria contado mas invisível); fora do painel, etapa
    # vazia. em_servicos manda; a etapa se ajusta.
    if lead.get("em_servicos"):
        if lead.get("status_servico") not in SERVICO_STAGES:
            lead["status_servico"] = SERVICO_STAGES[0]
    else:
        lead["status_servico"] = ""
    # Mesmo invariante para o painel do Curso (entrada e saida so manuais).
    if lead.get("em_curso"):
        if lead.get("status_curso") not in CURSO_STAGES:
            lead["status_curso"] = CURSO_STAGES[0]
    else:
        lead["status_curso"] = ""

    lead["updated_at"] = now_iso()
    return lead


def norm_phone(v):
    return re.sub(r"[^0-9]", "", str(v or ""))


def _canon_br(d):
    """Forma canonica para comparacao de numero brasileiro: remove o DDI 55,
    o 0 de tronco e o 9º dígito do celular (o WhatsApp ora inclui, ora omite)."""
    if d.startswith("55") and len(d) in (12, 13):
        d = d[2:]
    if d.startswith("0") and len(d) in (11, 12):  # 0 + DDD + numero
        d = d[1:]
    if len(d) == 11 and d[2] == "9":  # DDD + 9 + 8 digitos
        d = d[:2] + d[3:]
    return d


def same_phone(a, b):
    """Mesmo numero ignorando formatacao, +55/0/DDD e o 9º dígito do celular.
    Comparacao por igualdade canonica — sem casamento por sufixo, que mesclava
    numeros internacionais parecidos com numeros brasileiros."""
    da, db = norm_phone(a), norm_phone(b)
    if not da or not db:
        return False
    return da == db or _canon_br(da) == _canon_br(db)


def find_duplicado(telefone, email, exclude_id=None):
    """Procura outro lead com o mesmo telefone (ou e-mail). Retorna (lead, campo)."""
    email_n = str(email or "").strip().lower()
    for l in _db["leads"]:
        if exclude_id and l["id"] == exclude_id:
            continue
        if telefone and same_phone(l.get("telefone"), telefone):
            return l, "telefone"
        if email_n and str(l.get("email") or "").strip().lower() == email_n:
            return l, "e-mail"
    return None, None


# ---------------------------------------------------------------------------
# Importacao em massa (CSV)
# ---------------------------------------------------------------------------
def _slug_header(h):
    """Normaliza cabecalho de coluna: 'Região do Produtor' -> 'regiaodoprodutor'."""
    s = unicodedata.normalize("NFD", str(h or "")).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9]", "", s.lower())


# nomes de coluna aceitos -> campo do lead
IMPORT_COLS = {
    "nome": "nome", "cliente": "nome", "nomedoprodutor": "nome",
    "tipo": "tipo",
    "telefone": "telefone", "celular": "telefone", "whatsapp": "telefone",
    "fone": "telefone", "telefonewhatsapp": "telefone",
    "email": "email", "emailnf": "email",
    "regiao": "regiao", "cidade": "regiao", "municipio": "regiao", "regiaodoprodutor": "regiao",
    "area": "area_cultivada", "areacultivada": "area_cultivada",
    "produto": "produto", "produtodeinteresse": "produto",
    "cargo": "cargo", "cargodocontato": "cargo", "funcao": "cargo",
    "decisor": "decisor", "quemdecide": "decisor", "quempaga": "decisor",
    "pagamento": "pagamento", "formapagamento": "pagamento", "formadepagamento": "pagamento",
    "valor": "valor", "valorestimado": "valor", "valorestimadors": "valor",
    "sdr": "sdr", "vendedor": "vendedor", "vendedorresponsavel": "vendedor",
    "canal": "origem_canal", "origemcanal": "origem_canal", "origem": "origem_canal",
    "campanha": "campanha",
    "observacoes": "observacoes", "observacao": "observacoes", "obs": "observacoes",
    "status": "status", "etapa": "status",
}


def parse_hectares(s):
    """Extrai o numero de hectares de um texto livre ('3600', '500 ha',
    '1.200 hectares', '1.200,5'). Retorna float ou None."""
    if not s:
        return None
    m = re.search(r"[\d.,]+", str(s))
    if not m:
        return None
    num = m.group(0).strip(".,")
    if "." in num and "," in num:            # 1.200,5 -> 1200.5
        num = num.replace(".", "").replace(",", ".")
    elif "," in num:                          # 1200,5 -> 1200.5
        num = num.replace(",", ".")
    elif num.count(".") == 1 and len(num.split(".")[1]) == 3:
        num = num.replace(".", "")            # 1.200 -> 1200 (separador de milhar)
    try:
        return float(num)
    except ValueError:
        return None


def _parse_valor_br(texto):
    """Aceita '250000', '250000.50' e o formato brasileiro '250.000,50'."""
    t = str(texto or "").strip().replace("R$", "").strip()
    if not t:
        return 0.0
    if "," in t:
        t = t.replace(".", "").replace(",", ".")
    try:
        v = float(t)
    except ValueError:
        return 0.0
    return v if math.isfinite(v) else 0.0


def importar_csv(texto):
    """Importa leads de um CSV. Retorna (criados, rejeitados). Chamar SEM o lock."""
    primeira = texto.splitlines()[0] if texto.splitlines() else ""
    delim = ";" if primeira.count(";") >= primeira.count(",") else ","
    reader = csv.DictReader(io.StringIO(texto), delimiter=delim)
    if not reader.fieldnames:
        raise ValueError("Cabeçalho não encontrado no arquivo")
    campos = {orig: IMPORT_COLS.get(_slug_header(orig)) for orig in reader.fieldnames}
    if "telefone" not in campos.values():
        raise ValueError("O CSV precisa ter uma coluna de telefone (e de preferência nome e email)")

    criados = 0
    rejeitados = []
    with _lock:
        for i, row in enumerate(reader, start=2):  # linha 1 = cabecalho
            if criados >= 2000:
                rejeitados.append({"linha": i, "motivo": "limite de 2000 leads por importação — divida o arquivo"})
                break
            dados = {}
            for orig, alvo in campos.items():
                if alvo and row.get(orig) is not None:
                    dados[alvo] = str(row.get(orig) or "").strip()
            if not any(dados.values()):
                continue  # linha vazia
            nome = dados.get("nome", "")
            tel = dados.get("telefone", "")
            email = dados.get("email", "")
            if not tel or not email:
                rejeitados.append({"linha": i, "motivo": "%s: sem telefone e/ou e-mail (obrigatórios)" % (nome or "(sem nome)")})
                continue
            if len(norm_phone(tel)) < 8:
                rejeitados.append({"linha": i, "motivo": "%s: telefone inválido (%s)" % (nome or "(sem nome)", tel)})
                continue
            dup, campo_dup = find_duplicado(tel, email)
            if dup:
                rejeitados.append({"linha": i, "motivo": "%s: mesmo %s de \"%s\" (duplicado)" % (
                    nome or tel, campo_dup, dup.get("nome") or "lead existente")})
                continue

            lead = make_lead({"source": "importacao"})
            lead["nome"] = nome
            lead["telefone"] = tel
            lead["email"] = email
            # tolerante como o webhook: padroniza quando reconhece, aceita quando nao
            reg = dados.get("regiao", "")
            lead["regiao"] = canon_cidade(reg) or reg
            prod = dados.get("produto", "")
            for p in PRODUTOS:
                if prod and prod.lower() == p.lower():
                    prod = p
                    break
            lead["produto"] = prod if prod in PRODUTOS else ""
            lead["itens"] = itens_de_produto(prod)
            lead["valor"] = _parse_valor_br(dados.get("valor"))
            lead["area_cultivada"] = dados.get("area_cultivada", "")
            lead["cargo"] = dados.get("cargo", "")
            lead["decisor"] = dados.get("decisor", "")
            # "Financiamento + Permuta" -> duas formas (valores ficam zerados)
            pg = dados.get("pagamento", "")
            if pg:
                # casa ignorando espacos: "Permuta/Troca" == "Permuta / Troca"
                nomes = {re.sub(r"\s+", "", p.lower()): p for p in PAGAMENTOS}
                # separadores: + ; , (NUNCA "/", pois nomes tem "Permuta / Troca")
                achadas = [nomes[re.sub(r"\s+", "", t.lower())]
                           for t in re.split(r"[+;,]", pg) if re.sub(r"\s+", "", t.lower()) in nomes]
                lead["formas_pagamento"] = sanitiza_pagamentos(
                    [{"tipo": t} for t in achadas])
            lead["sdr"] = dados.get("sdr", "")
            lead["vendedor"] = dados.get("vendedor", "")
            lead["responsavel"] = lead["vendedor"] or lead["sdr"]
            if lead["vendedor"]:
                lead["atendido_em"] = now_iso()
            lead["origem_canal"] = dados.get("origem_canal", "")
            lead["campanha"] = dados.get("campanha", "")
            lead["observacoes"] = dados.get("observacoes", "")
            st = dados.get("status", "").lower()
            lead["status"] = st if st in STAGES else "novo"
            tp = dados.get("tipo", "").lower()
            if tp in ("produtor", "prestador", "pecuarista", "curso"):
                lead["tipo"] = tp
                if tp == "curso":  # classificado como curso ja entra no painel
                    lead["em_curso"] = True
                    lead["status_curso"] = CURSO_STAGES[0]
            elif lead["status"] in SALES_STAGES:
                lead["tipo"] = "produtor"
            if lead.get("tipo"):
                lead["qualificado_em"] = now_iso()
            if lead["status"] == "ganho":
                lead["ganho_em"] = now_iso()
            elif lead["status"] == "perdido":
                lead["perdido_em"] = now_iso()
            elif lead["status"] == "desistiu":
                lead["desistiu_em"] = now_iso()
            _db["leads"].append(lead)
            criados += 1
        if criados:
            save_db()
    return criados, rejeitados


# ---------------------------------------------------------------------------
# Usuarios, senhas e sessoes
# ---------------------------------------------------------------------------
def _slug_login(nome):
    s = unicodedata.normalize("NFD", str(nome or "")).encode("ascii", "ignore").decode()
    return re.sub(r"[^a-z0-9.]", "", s.lower().replace(" ", ".")) or "usuario"


def hash_senha(senha, salt=None):
    salt = salt or secrets.token_hex(16)
    h = hashlib.pbkdf2_hmac("sha256", str(senha).encode(), bytes.fromhex(salt), 120000).hex()
    return salt, h


def verifica_senha(user, senha):
    if not user.get("senha_hash") or not user.get("salt"):
        return False  # senha ainda nao definida pelo admin
    _, h = hash_senha(senha, user["salt"])
    return secrets.compare_digest(h, user["senha_hash"])


def ensure_admin():
    """Garante que exista ao menos um administrador para dar o primeiro acesso.
    A senha padrao fica marcada (senha_padrao=True) para o painel avisar que
    precisa ser trocada; a flag some quando o admin define uma senha nova."""
    if any(u.get("papel") == "admin" for u in _db["users"]):
        return None
    senha = "novaera123"
    salt, h = hash_senha(senha)
    _db["users"].append({
        "id": new_id(), "nome": "Administrador", "login": "admin",
        "salt": salt, "senha_hash": h, "papel": "admin", "ativo": True,
        "senha_padrao": True,
    })
    save_db()
    return senha


def ensure_webhook_token():
    """Resolve o token do webhook: env var ou um aleatorio persistido."""
    global WEBHOOK_TOKEN
    if WEBHOOK_TOKEN:
        return WEBHOOK_TOKEN
    tok = _db.get("settings", {}).get("webhook_token")
    if not tok:
        tok = secrets.token_urlsafe(24)
        _db.setdefault("settings", {})["webhook_token"] = tok
        save_db()
    WEBHOOK_TOKEN = tok
    return tok


def cria_sessao(user_id):
    token = secrets.token_urlsafe(32)
    _db["sessions"][token] = {"user_id": user_id, "exp": time.time() + SESSAO_DIAS * 86400}
    # faxina de sessoes vencidas
    agora = time.time()
    _db["sessions"] = {t: s for t, s in _db["sessions"].items() if s.get("exp", 0) > agora}
    save_db()
    return token


def usuario_da_sessao(token):
    if not token:
        return None
    s = _db.get("sessions", {}).get(token)
    if not s or s.get("exp", 0) < time.time():
        return None
    u = next((x for x in _db["users"] if x["id"] == s["user_id"]), None)
    if not u or not u.get("ativo", True):
        return None
    return u


def nome_em_uso(nome, exclude_id=None):
    alvo = str(nome or "").strip().lower()
    return any(str(u.get("nome") or "").strip().lower() == alvo and u["id"] != exclude_id
               for u in _db.get("users", []))


def renomeia_dono_leads(antigo, novo):
    """A posse do lead e gravada pelo NOME; ao renomear, atualiza todos os leads
    para o dono nao virar orfao (perder visibilidade e sair do rodizio)."""
    if not antigo or antigo == novo:
        return
    for l in _db["leads"]:
        for campo in ("sdr", "vendedor", "responsavel"):
            if l.get(campo) == antigo:
                l[campo] = novo
        for v in l.get("visitas", []):
            if v.get("visitante") == antigo:
                v["visitante"] = novo


def pode_recuperacao(user):
    """Quem pode ver/trabalhar o painel de Recuperação: admin/gerente sempre;
    SDR/vendedor só quando liberado individualmente (acesso_recuperacao)."""
    return user.get("papel") in ("admin", "gerente") or bool(user.get("acesso_recuperacao"))


def user_publico(u):
    return {"id": u["id"], "nome": u["nome"], "login": u["login"],
            "papel": u["papel"], "ativo": u.get("ativo", True),
            "senha_definida": bool(u.get("senha_hash")),
            "senha_padrao": bool(u.get("senha_padrao")),
            "acesso_recuperacao": bool(u.get("acesso_recuperacao")),
            "pode_recuperacao": pode_recuperacao(u),
            "recebe_leads": u.get("recebe_leads", True),
            "pode_mover": u.get("pode_mover", True)}


POTENCIAL_PADRAO = []


def potencial_regras():
    """Regras de potencial comercial: 'a cada X ha da cultura Y vendo 1 Z de R$ W'.
    Ficam nas settings do CRM (e nao no atlas.db) para sobreviverem a troca da base."""
    rs = _db.get("settings", {}).get("potencial_regras")
    if not isinstance(rs, list):
        return []
    limpas = []
    for r in rs:
        if not isinstance(r, dict):
            continue
        try:
            ha = float(r.get("ha_por_unidade") or 0)
            valor = float(r.get("valor_unidade") or 0)
        except (TypeError, ValueError):
            continue
        cid = r.get("cultura_id")
        if ha <= 0 or not str(cid).isdigit():
            continue
        limpas.append({"id": str(r.get("id") or "")[:40],
                       "cultura_id": int(cid),
                       "cultura": str(r.get("cultura") or "")[:60],
                       "ha_por_unidade": ha,
                       "produto": str(r.get("produto") or "Produto")[:60],
                       "valor_unidade": valor,
                       "ativa": r.get("ativa") is not False})
    return limpas


def settings_publico():
    """Config exposta ao painel — sem os segredos (tokens do webhook e do Chatwoot)."""
    st = _db.get("settings", {})
    pub = {k: v for k, v in st.items()
           if k not in ("webhook_token", "chatwoot_token", "curso_chatwoot_token")}
    pub["chatwoot_token_definido"] = bool(st.get("chatwoot_token"))
    pub["curso_chatwoot_token_definido"] = bool(st.get("curso_chatwoot_token"))
    return pub


def chatwoot_base_url(valor):
    """Extrai a RAIZ do endereço do Chatwoot. Aceita o endereço colado direto do
    navegador (ex.: https://chat.x.com.br/app/accounts/2/dashboard) e devolve so
    https://chat.x.com.br — senao a API seria chamada num caminho errado (404)."""
    u = str(valor or "").strip()
    if not u:
        return ""
    if not u.startswith(("http://", "https://")):
        u = "https://" + u
    m = re.match(r"^(https?://[^/\s]+)", u)
    return m.group(1) if m else ""


def chatwoot_cfg(origem=""):
    """Config do Chatwoot. origem="" = instancia da VENDA DE DRONES (padrao);
    origem="curso" = instancia separada do CURSO (cada lead conversa na sua)."""
    st = _db.get("settings", {})
    p = "curso_chatwoot_" if origem == "curso" else "chatwoot_"
    return (chatwoot_base_url(st.get(p + "url")),
            str(st.get(p + "account_id") or "").strip(),
            str(st.get(p + "token") or ""),
            str(st.get(p + "saudacao") or ""))


def origem_do_lead(lead):
    """Em qual Chatwoot vive a conversa deste lead ("" = drones, "curso")."""
    return "curso" if (lead or {}).get("chatwoot_origem") == "curso" else ""


def inbox_cfg_key(origem):
    return "curso_chatwoot_inbox_id" if origem == "curso" else "chatwoot_inbox_id"


def conv_id_valido(conv):
    """conversation_id só é usado em URL se for numérico (evita injeção de path
    numa chamada autenticada com o token). Retorna a string de dígitos ou None."""
    s = str(conv).strip() if conv is not None else ""
    return s if s.isdigit() else None


def chatwoot_conversa_url(base, acc, conv):
    conv = conv_id_valido(conv)
    if base and acc and conv:
        return "%s/app/accounts/%s/conversations/%s" % (base, acc, conv)
    return None


class _SemRedirect(urllib.request.HTTPRedirectHandler):
    """POST autenticado nao pode seguir redirect (viraria GET sem corpo e o
    'sucesso' seria mentira). 3xx vira HTTPError, tratado como recusa."""
    def redirect_request(self, *a, **k):
        return None


_ABRIDOR_SEM_REDIRECT = urllib.request.build_opener(_SemRedirect)


def chatwoot_headers(token, json_body=False):
    """Cabecalhos de autenticacao do Chatwoot. O token vai em DOIS nomes:
    'api_access_token' (o oficial) e 'Api-Access-Token' (com hifens) — nginx/
    openresty, por padrao, DESCARTA cabecalhos com underscore no nome, e o
    Chatwoot atras dele nunca via o token (401 mesmo com o token certo). O
    Rails normaliza hifens p/ underscore, entao os dois chegam iguais."""
    h = {"api_access_token": token, "Api-Access-Token": token}
    if json_body:
        h["Content-Type"] = "application/json"
    return h


def chatwoot_envia_mensagem(base, acc, conv, token, content):
    """Envia uma mensagem OUTGOING numa conversa do Chatwoot pela API. Retorna
    True se aceitou (2xx). Levanta excecao em erro de rede/HTTP (o chamador trata)."""
    conv = conv_id_valido(conv)
    if not conv:
        return False
    url = "%s/api/v1/accounts/%s/conversations/%s/messages" % (base, acc, conv)
    dados = json.dumps({"content": content, "message_type": "outgoing",
                        "private": False}).encode("utf-8")
    req = urllib.request.Request(url, data=dados, method="POST",
                                 headers=chatwoot_headers(token, json_body=True))
    with _ABRIDOR_SEM_REDIRECT.open(req, timeout=12) as resp:
        return 200 <= getattr(resp, "status", 200) < 300


ANEXO_MAX_BYTES = 10 * 1024 * 1024  # 10 MB por anexo (audio/imagem/video)
ANEXO_MIMES = {
    "image/jpeg", "image/png", "image/webp", "image/gif",
    "audio/webm", "audio/ogg", "audio/mpeg", "audio/mp4", "audio/aac", "audio/wav", "audio/x-m4a",
    "video/mp4", "video/webm", "video/quicktime", "video/3gpp",
}


def chatwoot_envia_anexo(base, acc, conv, token, content, nome, mime, dados):
    """Envia uma mensagem OUTGOING com ANEXO (multipart) na conversa do
    Chatwoot. content e opcional (legenda). Levanta excecao em erro."""
    conv = conv_id_valido(conv)
    if not conv:
        return False
    fronteira = "----cwanexo" + secrets.token_hex(12)
    partes = []
    campos = {"message_type": "outgoing", "private": "false"}
    if content:
        campos["content"] = content
    for k, v in campos.items():
        partes.append(("--%s\r\nContent-Disposition: form-data; name=\"%s\"\r\n\r\n%s\r\n"
                       % (fronteira, k, v)).encode("utf-8"))
    nome_seguro = re.sub(r"[^A-Za-z0-9._-]", "_", str(nome or "arquivo"))[:80] or "arquivo"
    partes.append(("--%s\r\nContent-Disposition: form-data; name=\"attachments[]\"; "
                   "filename=\"%s\"\r\nContent-Type: %s\r\n\r\n"
                   % (fronteira, nome_seguro, mime)).encode("utf-8"))
    partes.append(dados)
    partes.append(("\r\n--%s--\r\n" % fronteira).encode("utf-8"))
    corpo = b"".join(partes)
    url = "%s/api/v1/accounts/%s/conversations/%s/messages" % (base, acc, conv)
    cab = chatwoot_headers(token)
    cab["Content-Type"] = "multipart/form-data; boundary=%s" % fronteira
    req = urllib.request.Request(url, data=corpo, method="POST", headers=cab)
    with _ABRIDOR_SEM_REDIRECT.open(req, timeout=60) as resp:  # upload pode demorar
        return 200 <= getattr(resp, "status", 200) < 300


def _cw_data_msg(ts):
    """created_at do Chatwoot vem como unix (int) ou ISO — normaliza p/ ISO."""
    if ts in (None, ""):
        return None
    try:
        return datetime.fromtimestamp(int(ts), tz=timezone.utc).isoformat()
    except (TypeError, ValueError, OSError):
        return str(ts)


def _cw_req(base, caminho, token, corpo=None):
    """GET/POST na API do Chatwoot; devolve o JSON da resposta (dict/list).
    corpo=None -> GET. Levanta excecao em erro (o chamador traduz)."""
    url = "%s%s" % (base, caminho)
    dados = json.dumps(corpo).encode("utf-8") if corpo is not None else None
    req = urllib.request.Request(url, data=dados, method="POST" if corpo is not None else "GET",
                                 headers=chatwoot_headers(token, json_body=corpo is not None))
    with _ABRIDOR_SEM_REDIRECT.open(req, timeout=12) as resp:
        try:
            return json.loads(resp.read().decode("utf-8"))
        except Exception:
            return {}


def _cw_payload(d):
    """As respostas do Chatwoot vem em formatos variados ({payload: ...},
    {payload: {contact: ...}}, ou direto). Desembrulha ate o miolo."""
    if isinstance(d, dict) and "payload" in d:
        d = d["payload"]
    if isinstance(d, dict) and "contact" in d and isinstance(d["contact"], dict):
        d = d["contact"]
    return d


def _cw_conv_num(c):
    """Numero da conversa (o que aparece na URL do painel): display_id
    quando existe, senao id."""
    if not isinstance(c, dict):
        return None
    return conv_id_valido(c.get("display_id") if c.get("display_id") is not None else c.get("id"))


class ChatwootErro(Exception):
    """Erro amigavel (em PT) para mostrar ao usuario no toast."""


def chatwoot_conectar(base, acc, token, inbox_id, nome, telefone):
    """Acha (ou cria) o CONTATO pelo telefone e devolve o numero de uma conversa
    dele — reusa a conversa antiga se existir; senao cria uma nova na caixa de
    entrada configurada. Retorna (conversa_num, contato_id)."""
    dig = norm_phone(telefone).lstrip("0")  # 0 de tronco fora (011... -> 11...)
    if len(dig) < 8:
        raise ChatwootErro("Este lead não tem telefone válido para conectar no Chatwoot")
    if len(dig) <= 11:
        dig = "55" + dig
    e164 = "+" + dig

    def _busca_contato():
        # busca pelos ULTIMOS 8 digitos (estavel com/sem o 9º dígito) e valida
        # cada candidato com same_phone (igualdade canonica: DDI/0/9º dígito) —
        # sufixo solto casava contato de OUTRO DDD ou telefone curto/lixo.
        try:
            res = _cw_req(base, "/api/v1/accounts/%s/contacts/search?q=%s" % (acc, dig[-8:]), token)
        except Exception:
            return None
        achados = _cw_payload(res)
        if not isinstance(achados, list):
            return None
        for c in achados:
            tel_c = norm_phone((c or {}).get("phone_number"))
            if tel_c and len(tel_c) >= 8 and same_phone(tel_c, dig):
                return c.get("id")
        return None

    contato_id = _busca_contato()
    if not contato_id:
        try:
            res = _cw_req(base, "/api/v1/accounts/%s/contacts" % acc, token,
                          {"name": nome or e164, "phone_number": e164})
            c = _cw_payload(res)
            contato_id = c.get("id") if isinstance(c, dict) else None
        except urllib.error.HTTPError as e:
            if e.code == 422:  # telefone ja cadastrado la — busca de novo
                contato_id = _busca_contato()
            else:
                raise ChatwootErro("O Chatwoot recusou o cadastro do contato (HTTP %d)" % e.code)
    if not contato_id:
        raise ChatwootErro("Não consegui achar nem cadastrar o contato no Chatwoot")

    # conversa antiga do contato? (recuperacao: reaproveita o historico).
    # Falha na CONSULTA e erro (tentar de novo) — tratar como "sem conversa"
    # criaria uma conversa duplicada e o historico antigo se perderia.
    try:
        res = _cw_req(base, "/api/v1/accounts/%s/contacts/%s/conversations" % (acc, contato_id), token)
    except Exception:
        raise ChatwootErro("Não consegui consultar as conversas do cliente no Chatwoot — "
                           "tente de novo em alguns segundos")
    convs = _cw_payload(res)
    if isinstance(convs, list):
        candidatas = [c for c in convs if _cw_conv_num(c)]
        escolhida = None
        # prefere a conversa da CAIXA configurada (canal oficial de WhatsApp);
        # sem preferencia aplicavel, fica a primeira (mais recente no Chatwoot)
        if inbox_id:
            escolhida = next((c for c in candidatas
                              if str(c.get("inbox_id") or "") == str(inbox_id)), None)
        if not escolhida and candidatas:
            escolhida = candidatas[0]
        if escolhida:
            return int(_cw_conv_num(escolhida)), contato_id

    # nenhuma conversa: cria uma nova na caixa de entrada configurada
    if not inbox_id:
        raise ChatwootErro("Este cliente não tem conversa no Chatwoot ainda — preencha o "
                           "Nº da caixa de entrada em Gerenciar → Campanhas (o Testar conexão "
                           "mostra os números) para o CRM poder criar a conversa")
    try:
        res = _cw_req(base, "/api/v1/accounts/%s/contacts/%s/contact_inboxes" % (acc, contato_id),
                      token, {"inbox_id": int(inbox_id)})
        src = _cw_payload(res)
        source_id = src.get("source_id") if isinstance(src, dict) else None
        corpo = {"inbox_id": int(inbox_id), "contact_id": contato_id, "status": "open"}
        if source_id:
            corpo["source_id"] = source_id
        res = _cw_req(base, "/api/v1/accounts/%s/conversations" % acc, token, corpo)
        num = _cw_conv_num(_cw_payload(res))
        if not num:
            raise ChatwootErro("O Chatwoot criou a conversa mas não devolveu o número dela — "
                               "recarregue e tente de novo")
        # int: o webhook compara conversation_id como numero — string quebraria
        # o vinculo quando o cliente responder
        return int(num), contato_id
    except ChatwootErro:
        raise
    except urllib.error.HTTPError as e:
        raise ChatwootErro("O Chatwoot recusou a criação da conversa (HTTP %d) — confira o "
                           "Nº da caixa de entrada" % e.code)
    except Exception as e:
        raise ChatwootErro("Não consegui criar a conversa no Chatwoot (%s)" % type(e).__name__)


# Conexao EM LOTE dos leads da recuperacao (roda numa thread; o painel
# acompanha o progresso). IMPORTANTE: o lote SO VINCULA as conversas — nao
# envia nenhuma mensagem (disparo em massa derruba numero de WhatsApp).
_conectar_lote = {"rodando": False, "total": 0, "feitos": 0, "conectados": 0,
                  "ja_tinham": 0, "falhas": 0, "ultimo_erro": None, "terminado_em": None}

# Importacao dos leads ANTIGOS do Chatwoot do CURSO (lote de recuperacao):
# varre todas as conversas existentes la e cria cada cliente como lead 🎓 de
# recuperacao, ja conectado a conversa antiga. NUNCA envia mensagem.
_importar_curso = {"rodando": False, "paginas": 0, "vistos": 0, "criados": 0,
                   "ja_no_crm": 0, "sem_contato": 0, "falhas": 0,
                   "ultimo_erro": None, "terminado_em": None}


def _cw_lista_conversas(base, acc, token, pagina):
    """Uma pagina da lista de conversas (todas, inclusive resolvidas)."""
    url = "%s/api/v1/accounts/%s/conversations?status=all&page=%d" % (base, acc, pagina)
    req = urllib.request.Request(url, headers=chatwoot_headers(token))
    with _ABRIDOR_SEM_REDIRECT.open(req, timeout=20) as resp:
        dados = json.loads(resp.read().decode("utf-8", "replace"))
    if isinstance(dados, dict):
        d = dados.get("data")
        if isinstance(d, dict) and isinstance(d.get("payload"), list):
            return d["payload"]
        if isinstance(dados.get("payload"), list):
            return dados["payload"]
    return []


def _roda_importar_curso(base, acc, token):
    st = _importar_curso
    try:
        criados_sem_save = 0
        esgotou_teto = True  # vira False quando as paginas ACABAM de verdade
        for pagina in range(1, 401):  # teto de seguranca (~10.000 conversas)
            if not st["rodando"]:
                esgotou_teto = False
                break
            try:
                convs = _cw_lista_conversas(base, acc, token, pagina)
            except Exception as e:
                st["falhas"] += 1
                st["ultimo_erro"] = "Falha ao listar as conversas (%s)" % type(e).__name__
                esgotou_teto = False
                break
            if not convs:
                esgotou_teto = False
                break  # acabaram as paginas
            st["paginas"] = pagina
            for c in convs:
                if not st["rodando"] or not isinstance(c, dict):
                    continue
                st["vistos"] += 1
                cid = c.get("display_id") if c.get("display_id") is not None else c.get("id")
                s = str(cid).strip() if cid is not None else ""
                cid = int(s) if s.isdigit() else None
                sender = (c.get("meta") or {}).get("sender") or {}
                nome = str(sender.get("name") or sender.get("pushname") or "").strip()
                tel = str(sender.get("phone_number") or "").strip()
                email = str(sender.get("email") or "").strip()
                contato_id = sender.get("id")
                if not tel and not email:
                    st["sem_contato"] += 1  # sem como identificar/contatar
                    continue
                email_n = email.lower()
                with _lock:
                    ja = next((l for l in _db["leads"] if (
                        (cid is not None and l.get("chatwoot_conversation_id") == cid
                         and origem_do_lead(l) == "curso")
                        or (tel and same_phone(l.get("telefone"), tel))
                        or (email_n and str(l.get("email") or "").strip().lower() == email_n))), None)
                    if ja:
                        st["ja_no_crm"] += 1  # ja existe (conversa/telefone/e-mail)
                        continue
                    lead = make_lead({"source": "chatwoot", "nome": nome,
                                      "telefone": tel, "email": email})
                    lead["recuperacao"] = True
                    lead["tipo"] = "curso"
                    lead["em_curso"] = True
                    lead["status_curso"] = CURSO_STAGES[0]
                    lead["chatwoot_origem"] = "curso"
                    lead["qualificado_em"] = now_iso()
                    if cid is not None:
                        lead["chatwoot_conversation_id"] = cid
                    if contato_id is not None:
                        lead["chatwoot_contact_id"] = contato_id
                    registra_hist(lead, "Sistema",
                                  ["🎓 Lead antigo importado do Chatwoot do curso (lote de recuperação)"])
                    _db["leads"].append(lead)
                    st["criados"] += 1
                    criados_sem_save += 1
                    if criados_sem_save >= 10:
                        save_db()
                        criados_sem_save = 0
            time.sleep(0.15)  # nao martelar a API do Chatwoot
        if esgotou_teto:
            # parou no teto com paginas sobrando: NAO pode parecer completo
            st["falhas"] += 1
            st["ultimo_erro"] = ("Parei no limite de segurança (%d conversas verificadas) — "
                                 "o Chatwoot ainda tinha mais conversas além dessas" % st["vistos"])
    finally:
        with _lock:
            save_db()
        st["rodando"] = False
        st["terminado_em"] = now_iso()


def _roda_conectar_lote(base, acc, token, inbox):
    st = _conectar_lote
    erros_seguidos = 0
    try:
        with _lock:
            ids = [l["id"] for l in _db["leads"]
                   if l.get("recuperacao")
                   # lote dos DRONES: lead do curso conversa na OUTRA instancia
                   and l.get("tipo") != "curso" and origem_do_lead(l) != "curso"
                   and not conv_id_valido(l.get("chatwoot_conversation_id"))
                   and len(norm_phone(l.get("telefone"))) >= 8]
        st["total"] = len(ids)
        alterados = 0
        for lid in ids:
            if not st["rodando"]:
                break  # alguem pediu para parar
            with _lock:
                lead = next((l for l in _db["leads"] if l["id"] == lid), None)
                if not lead or conv_id_valido(lead.get("chatwoot_conversation_id")):
                    st["feitos"] += 1
                    st["ja_tinham"] += 1
                    continue
                if lid in _cw_em_voo:  # clique manual em andamento: pula
                    st["feitos"] += 1
                    continue
                _cw_em_voo.add(lid)
                nome = str(lead.get("nome") or "")
                tel = lead.get("telefone") or ""
            try:
                conv, contato = chatwoot_conectar(base, acc, token, inbox, nome, tel)
                with _lock:
                    l2 = next((l for l in _db["leads"] if l["id"] == lid), None)
                    if l2 and not conv_id_valido(l2.get("chatwoot_conversation_id")):
                        l2["chatwoot_conversation_id"] = conv
                        if contato and not l2.get("chatwoot_contact_id"):
                            l2["chatwoot_contact_id"] = contato
                        registra_hist(l2, "Sistema", ["🔗 Conversa conectada no Chatwoot (lote da recuperação)"])
                        l2["updated_at"] = now_iso()
                        alterados += 1
                        if alterados % 10 == 0:
                            save_db()  # salva aos poucos (nao 320 gravacoes)
                st["conectados"] += 1
                erros_seguidos = 0
            except ChatwootErro as e:
                st["falhas"] += 1
                st["ultimo_erro"] = str(e)
                erros_seguidos += 1
            except Exception as e:
                st["falhas"] += 1
                st["ultimo_erro"] = "Erro inesperado (%s)" % type(e).__name__
                erros_seguidos += 1
            finally:
                with _lock:
                    _cw_em_voo.discard(lid)
                st["feitos"] += 1
            if erros_seguidos >= 8:
                st["ultimo_erro"] = (st["ultimo_erro"] or "falhas") + \
                    " — lote interrompido: muitas falhas seguidas (confira a configuração e rode de novo)"
                break
            time.sleep(0.2)  # ritmo suave: nao metralhar o Chatwoot
    finally:
        with _lock:
            save_db()
        st["rodando"] = False
        st["terminado_em"] = now_iso()


# ---------------------------------------------------------------------------
# Permissoes por papel
# ---------------------------------------------------------------------------
VENDAS_STATUSES = SALES_STAGES  # etapas em que o lead esta no funil de vendas


def no_funil_vendas(lead):
    return lead.get("status") in VENDAS_STATUSES or (
        lead.get("status") in ("perdido", "desistiu") and bool(lead.get("tipo")))


def em_fase_sdr(lead):
    """Lead ainda na triagem (o territorio do SDR)."""
    return lead.get("status") in ("novo", "triagem", "curioso") or (
        lead.get("status") == "perdido" and not lead.get("tipo"))


# Etapas que sao TRABALHO DE VENDEDOR — um SDR nao pode mover um lead para ca
# (o SDR vai ate a qualificacao; dali em diante e venda).
ETAPAS_DE_VENDA = ["decidindo", "negociacao", "proposta", "financiamento", "ganho"]


def pode_ver_lead(user, lead):
    p = user.get("papel")
    if p in ("admin", "gerente"):
        return True
    if p == "sdr":
        return lead.get("sdr") == user.get("nome")
    if p == "vendedor":
        # vendedor faz TAMBEM o servico do SDR: ve toda a fase de triagem
        if em_fase_sdr(lead):
            return True
        dono = str(lead.get("vendedor") or "").strip()
        return no_funil_vendas(lead) and dono in ("", user.get("nome"))
    return False


def active_members(papel=None):
    """Equipe = usuarios ativos com papel de raia (sdr/vendedor)."""
    out = [u for u in _db.get("users", []) if u.get("ativo", True) and u.get("papel") in PAPEIS]
    if papel:
        out = [u for u in out if u.get("papel") == papel]
    return out


def next_sdr():
    """Escolhe o proximo SDR no rodizio (round-robin). Retorna nome ou ''.

    So entram SDRs marcados como "recebe leads" (o administrador escolhe no
    painel de equipe). Sem nenhum marcado, o lead novo fica sem responsavel —
    visivel na coluna "Sem responsavel" da triagem.

    O rodizio guarda o NOME do ultimo contemplado, nao um indice: como a lista
    agora muda de tamanho toda hora (ligar/desligar SDR), um indice posicional
    pularia ou repetiria alguem a cada mudanca — vies que se acumula."""
    sdrs = [u.get("nome", "") for u in active_members("sdr") if u.get("recebe_leads", True)]
    sdrs = [n for n in sdrs if n]
    if not sdrs:
        return ""
    ultimo = _db.get("rr_sdr_ultimo") or ""
    if ultimo in sdrs:
        escolhido = sdrs[(sdrs.index(ultimo) + 1) % len(sdrs)]
    else:
        escolhido = sdrs[0]
    _db["rr_sdr_ultimo"] = escolhido
    return escolhido


# ---------------------------------------------------------------------------
# Integracao com o Chatwoot
# ---------------------------------------------------------------------------
def parse_utms(referer):
    out = {}
    if not referer or not isinstance(referer, str):
        return out
    try:
        qs = parse_qs(urlparse(referer).query)
        for f in ("utm_source", "utm_medium", "utm_campaign", "utm_content", "utm_term"):
            if qs.get(f):
                out[f] = qs[f][0]
    except Exception:
        pass
    return out


SOURCE_CANAL = {
    "facebook": "Meta", "instagram": "Meta", "fb": "Meta", "ig": "Meta",
    "meta": "Meta", "google": "Google", "adwords": "Google", "youtube": "Google",
    "whatsapp": "WhatsApp", "wa": "WhatsApp", "tiktok": "TikTok",
}


def guess_canal(utms, referer):
    """Deduz o canal comparando tokens exatos (nunca substring solta — 'ig'
    dentro de 'campaign' ou 'digital' nao pode virar Meta)."""
    src = (utms.get("utm_source") or "").strip().lower()
    if src in SOURCE_CANAL:
        return SOURCE_CANAL[src]

    host, query = "", ""
    try:
        parsed = urlparse(referer or "")
        host = (parsed.netloc or "").lower()
        query = (parsed.query or "").lower()
    except Exception:
        pass
    if host:
        if re.search(r"(^|\.)(facebook\.com|instagram\.com|fb\.me|fb\.com|meta\.com)$", host):
            return "Meta"
        if re.search(r"(^|\.)(google\.[a-z.]+|youtube\.com)$", host):
            return "Google"
        if re.search(r"(^|\.)(wa\.me|whatsapp\.com)$", host):
            return "WhatsApp"
        if re.search(r"(^|\.)tiktok\.com$", host):
            return "TikTok"
    # ids de clique presentes na URL denunciam a origem
    if "fbclid=" in query:
        return "Meta"
    if "gclid=" in query:
        return "Google"
    return utms.get("utm_source", "")


def find_referral(payload):
    """
    Procura os dados de anuncio "clique-pro-WhatsApp" (Meta) no evento.

    Quando alguem clica num anuncio do Facebook/Instagram que abre o WhatsApp,
    a primeira mensagem chega com um bloco "referral" (id do anuncio, titulo,
    ctwa_clid...). O lugar exato varia conforme a versao do Chatwoot, entao
    fazemos uma busca em largura (limitada) por um dicionario com essa cara.
    """
    def looks_like_referral(d):
        if not isinstance(d, dict):
            return False
        if d.get("ctwa_clid"):
            return True
        if d.get("source_id") and d.get("source_type"):
            return True
        if d.get("source_url") and (d.get("headline") or d.get("body")):
            return True
        return False

    queue = [payload]
    seen = 0
    while queue and seen < 200:
        node = queue.pop(0)
        seen += 1
        if looks_like_referral(node):
            return {
                "source_id": str(node.get("source_id") or ""),
                "source_type": str(node.get("source_type") or ""),
                "headline": str(node.get("headline") or ""),
                "body": str(node.get("body") or ""),
                "source_url": str(node.get("source_url") or ""),
                "ctwa_clid": str(node.get("ctwa_clid") or ""),
            }
        if isinstance(node, dict):
            queue.extend(node.values())
        elif isinstance(node, list):
            queue.extend(node)
    return None


def match_campaign(utms, message_text):
    """
    Tenta vincular o lead a uma campanha cadastrada. Prioridade:
      1. utm_campaign igual ao codigo, ao utm_campaign ou ao nome da campanha
      2. "#CODIGO" presente no texto da mensagem (link de WhatsApp do anuncio)
      3. palavra-chave da campanha presente na mensagem
    """
    campaigns = [c for c in _db.get("campaigns", []) if c.get("ativo", True)]
    if not campaigns:
        return None

    utm_c = (utms.get("utm_campaign") or "").strip().lower()
    if utm_c:
        for c in campaigns:
            candidates = {c.get("codigo", ""), c.get("utm_campaign", ""), c.get("nome", "")}
            if utm_c in {x.strip().lower() for x in candidates if x}:
                return c

    text = (message_text or "").lower()
    if text:
        for c in campaigns:
            code = (c.get("codigo") or "").strip().lower()
            # fronteira apos o codigo: evita "#SOJA" casar com "#SOJA25"
            if code and re.search("#" + re.escape(code) + r"(?![a-z0-9])", text):
                return c
        # keywords mais longas primeiro ("soja premium" vence "soja"), e com
        # fronteira de palavra ("uva" nao pode casar "chuva", "milho"/"milhoes")
        by_len = sorted(campaigns, key=lambda c: len(c.get("keyword") or ""), reverse=True)
        for c in by_len:
            kw = (c.get("keyword") or "").strip().lower()
            if kw and re.search(r"\b" + re.escape(kw) + r"\b", text):
                return c
    return None


def gen_codigo(nome):
    """Gera um codigo curto e unico a partir do nome (ex.: 'Soja Safra 2025' -> SOJA25)."""
    letters = re.sub(r"[^A-Za-z]", "", nome or "").upper()[:4] or "CAMP"
    digits = re.sub(r"[^0-9]", "", nome or "")[-2:]
    base = letters + digits
    existing = {(c.get("codigo") or "").upper() for c in _db.get("campaigns", [])}
    codigo, n = base, 1
    while codigo.upper() in existing:
        n += 1
        codigo = "%s%d" % (base, n)
    return codigo


# Eventos do Chatwoot que realmente representam conversa/mensagem de lead.
# contact_updated & cia. tem "id" de CONTATO no topo — tratar como conversa
# criaria leads fantasma com id trocado (colisao garantida entre sequencias).
CONVERSATION_EVENTS = ("conversation_created", "conversation_updated", "conversation_status_changed")
MESSAGE_EVENTS = ("message_created", "message_updated")


def _is_incoming(msg):
    """So mensagem do LEAD conta (nao a resposta do atendente nem nota privada)."""
    if not isinstance(msg, dict):
        return False
    if msg.get("private"):
        return False
    mt = msg.get("message_type")
    return mt in ("incoming", 0, None)  # None: payload minimo/sintetico


def _msg_id_int(payload):
    """Id numerico da mensagem do webhook (0 quando nao veio/nao e numero)."""
    mid = payload.get("id")
    if isinstance(mid, int):
        return mid
    s = str(mid or "").strip()
    return int(s) if s.isdigit() else 0


def _conv_ids_do_evento(payload):
    """Numero da conversa num evento de MENSAGEM.

    O Chatwoot manda o display_id (o numero que aparece na tela e nas URLs —
    e o que o CRM guarda) e um id global da instalacao. Quando o display_id
    veio, ele e o UNICO candidato: misturar os dois casaria a conversa com o
    lead errado (o display_id de um lead pode coincidir com o id global de
    OUTRA conversa). O id global so serve de reserva quando o payload nao
    trouxe display_id."""
    conv = payload.get("conversation") or {}

    def _n(v):
        s = str(v).strip() if v is not None else ""
        return int(s) if s.isdigit() else None

    disp = _n(conv.get("display_id"))
    if disp is not None:
        return disp, {disp}
    for v in (conv.get("id"), payload.get("conversation_id")):
        n = _n(v)
        if n is not None:
            return n, {n}
    return None, set()


def _resumo_msg(payload):
    """Frase curta descrevendo a mensagem (texto ate 200 letras + anexos)."""
    texto = str(payload.get("content") or "").strip()
    if len(texto) > 200:
        texto = texto[:200] + "…"
    rotulos = []
    for a in (payload.get("attachments") or []):
        ft = str((a or {}).get("file_type") or "") if isinstance(a, dict) else ""
        rotulos.append({"image": "🖼 imagem", "audio": "🎤 áudio",
                        "video": "🎬 vídeo"}.get(ft, "📎 arquivo"))
    partes = " + ".join(rotulos)
    if texto and partes:
        return '%s "%s"' % (partes, texto)
    if partes:
        return partes
    return '"%s"' % texto if texto else ""


def _marca_msg(origem, msg_id):
    """Marcador de dedup POR INSTANCIA: ids de mensagem sao sequenciais por
    instalacao do Chatwoot — o id 500 dos drones e o 500 do curso sao mensagens
    diferentes e nao podem se anular."""
    return ("curso:%d" % msg_id) if origem == "curso" else str(msg_id)


def _registra_msg_cliente(lead, payload, origem=""):
    """Mensagem RECEBIDA do cliente: grava o que ele disse no historico, baixa
    o alerta "registre a resposta" (ja esta registrada) e acende o aviso verde
    "cliente respondeu — responda!". Chamar SEGURANDO o _lock."""
    # dedup por CONJUNTO de marcadores ja vistos (nao por "maior id"): webhooks
    # chegam fora de ordem e mensagem atrasada legitima nao pode ser descartada
    msg_id = _msg_id_int(payload)
    marca = _marca_msg(origem, msg_id) if msg_id else ""
    vistos = lead.get("chatwoot_msgs_vistas")
    if not isinstance(vistos, list):
        vistos = []
    if marca and marca in vistos:
        return False  # reentrega da mesma mensagem (o Chatwoot repete envios)
    if marca:
        vistos.append(marca)
        lead["chatwoot_msgs_vistas"] = vistos[-80:]
    # Cliente dado como perdido/desistiu/curioso que volta a escrever NA MESMA
    # conversa REABRE na triagem (mesma regra da conversa nova) — a equipe
    # precisa enxergar a volta. Ganho continua ganho (vira pos-venda).
    if lead.get("status") in ("perdido", "desistiu", "curioso"):
        lead["status"] = "novo"
        lead["tipo"] = ""
        registra_hist(lead, "Chatwoot", ["🔄 Cliente voltou a escrever — reaberto na triagem"])
    corpo = _resumo_msg(payload)
    if corpo:
        item = "📥 Cliente: " + corpo
        # mensagens em sequencia (< 30 min) entram na MESMA entrada do
        # historico — cliente que manda 6 frases nao vira 6 entradas
        ult = lead["historico"][-1] if lead.get("historico") else None
        junta = False
        if ult and ult.get("tipo") == "resposta" and len(ult.get("itens") or []) < 30:
            try:
                t = datetime.fromisoformat(str(ult.get("data", "")).replace("Z", "+00:00"))
                junta = (datetime.now(timezone.utc) - t) < timedelta(minutes=30)
            except Exception:
                junta = False
        if junta:
            ult["itens"].append(item)
        else:
            registra_hist(lead, "Cliente", [item], tipo="resposta")
    # lead encerrado nao ganha cobranca (mesma regra dos outros alertas)
    if lead.get("status") not in ("ganho", "perdido", "desistiu", "curioso"):
        lead["cliente_respondeu"] = now_iso()
        lead["aguardando_resposta"] = None
    lead["updated_at"] = now_iso()
    return True


def _dt_msg(payload):
    """Horario de criacao da mensagem do webhook como datetime UTC.

    O Chatwoot manda created_at ora como unix (numero), ora como texto
    ("2026-07-29T13:05:57Z" / "2026-07-29 13:05:57 UTC"). None se ilegivel."""
    v = payload.get("created_at")
    if v is None:
        return None
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        try:
            return datetime.fromtimestamp(v, tz=timezone.utc)
        except Exception:
            return None
    s = str(v).strip().replace(" UTC", "+00:00").replace("Z", "+00:00")
    if "T" not in s:
        s = s.replace(" ", "T", 1)
    try:
        dt = datetime.fromisoformat(s)
        return dt if dt.tzinfo else dt.replace(tzinfo=timezone.utc)
    except Exception:
        return None


def _msg_atendente(payload, origem=""):
    """Mensagem ENVIADA pelo atendente (respondida direto no painel do Chatwoot,
    ou o eco da que o CRM acabou de mandar): apaga o aviso "cliente respondeu"
    e reinicia o relogio da resposta pendente. NUNCA cria nem mescla lead —
    o sender de uma mensagem outgoing e o ATENDENTE, nao o cliente."""
    if payload.get("event") != "message_created" or payload.get("private"):
        return {"ok": True, "ignored": "mensagem sem efeito (nota/edicao)"}
    if payload.get("message_type") not in ("outgoing", 1):
        return {"ok": True, "ignored": "atividade/template"}
    # So atendente HUMANO conta (sender tipo "user"). Robo, regra de automacao
    # e campanha saem sem sender (ou com tipo proprio) — resposta automatica
    # nao pode apagar o aviso "cliente respondeu" da equipe.
    if str((payload.get("sender") or {}).get("type") or "").lower() != "user":
        return {"ok": True, "ignored": "mensagem automatica (sem atendente)"}
    _, cands = _conv_ids_do_evento(payload)
    if not cands:
        return {"ok": True, "ignored": "sem numero de conversa"}
    with _lock:
        lead = next((l for l in _db["leads"]
                     if l.get("chatwoot_conversation_id") in cands
                     and origem_do_lead(l) == origem), None)
        if not lead:
            return {"ok": True, "ignored": "conversa sem lead"}
        msg_id = _msg_id_int(payload)
        marca = _marca_msg(origem, msg_id) if msg_id else ""
        vistos = lead.get("chatwoot_msgs_vistas")
        if not isinstance(vistos, list):
            vistos = []
        if marca and marca in vistos:
            return {"ok": True, "ignored": "mensagem ja processada"}
        if marca:
            vistos.append(marca)
            lead["chatwoot_msgs_vistas"] = vistos[-80:]
        # Eco/entrega atrasada: se a resposta do cliente e MAIS NOVA que esta
        # mensagem do atendente, a ultima palavra e do cliente — o aviso verde
        # fica aceso (senao o eco da propria mensagem do CRM apagaria a
        # resposta que chegou logo depois dela).
        resp = lead.get("cliente_respondeu")
        t_msg = _dt_msg(payload)
        if resp and t_msg:
            try:
                t_resp = datetime.fromisoformat(str(resp).replace("Z", "+00:00"))
                if t_msg <= t_resp:
                    save_db()  # o id ja entrou na lista de vistos
                    return {"ok": True, "ignored": "anterior a resposta do cliente"}
            except Exception:
                pass
        lead["cliente_respondeu"] = None
        if lead.get("status") not in ("ganho", "perdido", "desistiu", "curioso"):
            lead["aguardando_resposta"] = now_iso()
        lead["updated_at"] = now_iso()
        save_db()
    return {"ok": True, "atendente": True, "id": lead["id"]}


def handle_chatwoot_event(payload, origem=""):
    """origem = "" (Chatwoot da venda de drones) | "curso" (Chatwoot do curso).
    Numeros de conversa das duas instancias podem COINCIDIR — todo casamento
    por conversa filtra tambem pela origem."""
    event = payload.get("event") if isinstance(payload, dict) else None
    if not event:
        return {"ok": False, "reason": "sem evento"}

    # Filtra os tipos de evento suportados; os demais sao confirmados e ignorados
    if event in CONVERSATION_EVENTS:
        conversation = payload
        conversation_id = payload.get("id")
    elif event in MESSAGE_EVENTS:
        # edicao/apagamento de mensagem nao e atividade nova: nao pode mexer no
        # lead (nem adiar o alerta de retorno via updated_at)
        if event == "message_updated":
            return {"ok": True, "ignored": "edicao de mensagem"}
        # mensagem do atendente/nota privada: so baixa os avisos do lead
        if not _is_incoming(payload):
            return _msg_atendente(payload, origem)
        conversation = payload.get("conversation") or {}
        # preferimos o display_id (o numero que aparece nas URLs do Chatwoot);
        # o id global fica como candidato para casar leads antigos
        conversation_id, conv_cands = _conv_ids_do_evento(payload)
    else:
        return {"ok": True, "ignored": event}

    # id de conversa só serve se for numérico (vai parar em URLs autenticadas)
    if conversation_id is not None and not isinstance(conversation_id, int):
        s = str(conversation_id).strip()
        conversation_id = int(s) if s.isdigit() else None
    if event in CONVERSATION_EVENTS:
        conv_cands = {conversation_id} if conversation_id is not None else set()

    meta = conversation.get("meta") or payload.get("meta") or {}
    sender = (
        meta.get("sender")
        or payload.get("sender")
        or (payload.get("contact_inbox") or {}).get("contact")
        or {}
    )

    add_attrs = conversation.get("additional_attributes") or payload.get("additional_attributes") or {}
    referer = add_attrs.get("referer") or add_attrs.get("referrer") or ""
    utms = dict(parse_utms(referer))
    custom = conversation.get("custom_attributes") or payload.get("custom_attributes") or {}
    for f in ("utm_source", "utm_medium", "utm_campaign", "utm_content", "utm_term"):
        if not utms.get(f) and custom.get(f):
            utms[f] = str(custom[f])
        if not utms.get(f) and add_attrs.get(f):
            utms[f] = str(add_attrs[f])
    canal = guess_canal(utms, referer)

    last_message = ""
    msgs = payload.get("messages")
    if isinstance(msgs, list) and msgs:
        incoming_msgs = [m for m in msgs if _is_incoming(m)]
        if incoming_msgs:
            last_message = incoming_msgs[-1].get("content") or ""
    elif payload.get("content"):
        last_message = payload.get("content")

    nome = sender.get("name") or sender.get("pushname") or ""
    telefone = sender.get("phone_number") or sender.get("phone") or ""
    email = sender.get("email") or ""
    contact_id = sender.get("id")

    sender_custom = sender.get("custom_attributes") or {}
    regiao = sender_custom.get("regiao") or sender_custom.get("region") or custom.get("regiao") or ""
    area = sender_custom.get("area_cultivada") or sender_custom.get("area") or custom.get("area_cultivada") or ""
    produto = sender_custom.get("produto") or custom.get("produto") or ""
    cargo = sender_custom.get("cargo") or custom.get("cargo") or ""
    decisor = sender_custom.get("decisor") or custom.get("decisor") or ""
    # padroniza grafia sem rejeitar (lead do webhook nunca pode ser perdido)
    regiao = canon_cidade(regiao) or regiao
    for p in PRODUTOS:
        if produto and str(produto).strip().lower() == p.lower():
            produto = p
            break

    # ---- Atribuicao de campanha ----
    # Rota 1: anuncio Meta clique-pro-WhatsApp (bloco "referral" no evento)
    referral = find_referral(payload)
    meta_ad_id = ""
    ctwa_clid = ""
    campanha_nome = utms.get("utm_campaign", "")
    if referral:
        canal = canal or "Meta"
        meta_ad_id = referral["source_id"]
        ctwa_clid = referral["ctwa_clid"]
        if not campanha_nome:
            campanha_nome = referral["headline"] or ("Anúncio " + meta_ad_id if meta_ad_id else "")

    # Rotas 2 e 3: campanha cadastrada casando com UTM ou com o texto da
    # mensagem (o titulo/texto do anuncio tambem entram na busca por
    # palavra-chave, para vincular leads de anuncio CTWA automaticamente)
    match_text = last_message
    if referral:
        match_text = " ".join(x for x in (last_message, referral["headline"], referral["body"]) if x)
    campanha_id = ""
    camp = match_campaign(utms, match_text)
    if camp:
        campanha_id = camp["id"]
        campanha_nome = camp["nome"]
        canal = canal or camp.get("canal", "")

    incoming = {
        "source": "chatwoot",
        "chatwoot_conversation_id": conversation_id,
        "chatwoot_contact_id": contact_id,
        "nome": nome,
        "telefone": telefone,
        "email": email,
        "regiao": regiao,
        "area_cultivada": area,
        "produto": produto,
        "itens": itens_de_produto(produto),
        "cargo": cargo,
        "decisor": decisor,
        "origem_canal": canal,
        "campanha": campanha_nome,
        "campanha_id": campanha_id,
        "meta_ad_id": meta_ad_id,
        "ctwa_clid": ctwa_clid,
        "last_message": last_message,
    }
    incoming.update(utms)

    with _lock:
        # A campanha casada pode ter sido excluida entre o match (fora do lock)
        # e agora; revalida para nao gravar vinculo orfao.
        if incoming.get("campanha_id") and not any(
                c["id"] == incoming["campanha_id"] for c in _db.get("campaigns", [])):
            incoming["campanha_id"] = ""

        lead = None
        if conv_cands:
            lead = next((l for l in _db["leads"]
                         if l.get("chatwoot_conversation_id") in conv_cands
                         and origem_do_lead(l) == origem), None)

        # Cliente conhecido abrindo conversa NOVA: reconhece pelo id do contato,
        # telefone ou e-mail e atualiza o lead existente em vez de duplicar
        # (adota a conversa nova para as proximas mensagens chegarem certo).
        if lead is None and (contact_id is not None or telefone or email):
            email_n = str(email or "").strip().lower()
            # id de CONTATO so casa dentro da MESMA instancia (o contato nº 41
            # dos drones e o nº 41 do curso podem ser pessoas diferentes);
            # telefone/e-mail identificam a pessoa em qualquer instancia
            candidatos = [l for l in _db["leads"] if (
                (contact_id is not None and l.get("chatwoot_contact_id") == contact_id
                 and origem_do_lead(l) == origem)
                or (telefone and same_phone(l.get("telefone"), telefone))
                or (email_n and str(l.get("email") or "").strip().lower() == email_n))]
            if candidatos:
                # prefere um lead em atendimento; so cai num encerrado se nao houver
                ativos = [l for l in candidatos if l.get("status") not in ("ganho", "perdido", "desistiu", "curioso")]
                lead = (ativos or candidatos)[0]
                # So ADOTA (troca conversa/origem) quando e uma conversa NOVA de
                # verdade (evento de conversa) ou quando o lead ainda nao tem
                # conversa. Mensagem avulsa da OUTRA instancia nao rouba o
                # vinculo — senao um cliente ativo nas duas instancias faria o
                # lead "quicar" de origem a cada mensagem, com spam no historico.
                trocou = (conversation_id is not None
                          and lead.get("chatwoot_conversation_id") != conversation_id
                          and (event in CONVERSATION_EVENTS
                               or not conv_id_valido(lead.get("chatwoot_conversation_id"))))
                if trocou:
                    lead["chatwoot_conversation_id"] = conversation_id
                    # a conversa adotada vive NESTA instancia: troca a origem e o
                    # contato junto (o contato antigo era da outra instancia)
                    lead["chatwoot_origem"] = origem
                    if contact_id is not None:
                        lead["chatwoot_contact_id"] = contact_id
                # cliente que estava dado como perdido/desistiu/curioso voltou:
                # reentra na triagem para a equipe enxergar
                if lead.get("status") in ("perdido", "desistiu", "curioso"):
                    lead["status"] = "novo"
                    lead["tipo"] = ""  # volta para a triagem do SDR
                    registra_hist(lead, "Chatwoot", ["🔄 Cliente voltou pelo Chatwoot — reaberto na triagem"])
                elif trocou:
                    registra_hist(lead, "Chatwoot", ["💬 Nova conversa no Chatwoot"])
                print("[webhook] conversa nova %s reconhecida -> lead %s" % (
                    conversation_id, lead.get("nome") or lead["id"]))

        if lead is None:
            if conversation_id is None and not telefone and not email:
                return {"ok": False, "reason": "evento sem dados de contato"}
            lead = make_lead(incoming)
            lead["chatwoot_origem"] = origem
            if origem == "curso":
                # cliente do Chatwoot do CURSO ja chega classificado e no painel
                lead["tipo"] = "curso"
                lead["em_curso"] = True
                lead["status_curso"] = CURSO_STAGES[0]
                lead["qualificado_em"] = lead.get("qualificado_em") or now_iso()
            # Rodizio: o lead novo ja cai para um SDR fazer o primeiro contato.
            sdr = next_sdr()
            if sdr:
                lead["sdr"] = sdr
                lead["responsavel"] = sdr
            registra_hist(lead, "Chatwoot", [
                "🆕 Lead recebido do Chatwoot" + (" (canal: %s)" % canal if canal else ""),
                "📞 SDR: %s (rodízio)" % sdr if sdr else "",
            ], tipo="novo")
            _db["leads"].append(lead)
            # primeira mensagem do cliente ja entra registrada no historico
            if event == "message_created":
                _registra_msg_cliente(lead, payload, origem)
            save_db()
            print("[webhook] novo lead: %s -> SDR %s (canal: %s)" % (
                nome or telefone or conversation_id, sdr or "-", canal or "-"))
            return {"ok": True, "created": True, "id": lead["id"], "sdr": sdr}

        # Se a campanha cadastrada foi identificada agora (ex.: o codigo veio na
        # mensagem seguinte), vincula e espelha o nome mesmo que o campo texto
        # ja tivesse algo generico (titulo do anuncio, utm solto). Lead ja GANHO
        # nao recebe vinculo novo: a venda fechada nao pode ser creditada a uma
        # campanha que nao a gerou (distorceria o relatorio).
        if incoming.get("campanha_id") and not lead.get("campanha_id") and lead.get("status") != "ganho":
            lead["campanha_id"] = incoming["campanha_id"]
            lead["campanha"] = incoming["campanha"]

        for key, value in incoming.items():
            if key == "last_message":
                if value:
                    lead["last_message"] = value
                continue
            if value and not lead.get(key):
                # nao preenche telefone/email que criaria duplicata com OUTRO lead
                if key in ("telefone", "email"):
                    d, _campo = find_duplicado(
                        value if key == "telefone" else None,
                        value if key == "email" else None,
                        exclude_id=lead["id"])
                    if d:
                        continue
                lead[key] = value
        # mantém o espelho `produto` (usado na busca) coerente com os itens do
        # pedido — o merge acima trata as chaves soltas e poderia divergir.
        if lead.get("itens"):
            lead["produto"] = resumo_produtos(lead["itens"])
        # o que o cliente respondeu entra sozinho no historico + aviso "responda!"
        if event == "message_created":
            _registra_msg_cliente(lead, payload, origem)
        lead["updated_at"] = now_iso()
        save_db()
        print("[webhook] lead atualizado: %s" % (lead.get("nome") or lead.get("telefone") or lead["id"]))
        return {"ok": True, "created": False, "id": lead["id"]}


# ---------------------------------------------------------------------------
# Servidor HTTP
# ---------------------------------------------------------------------------
class Handler(BaseHTTPRequestHandler):
    server_version = "AgroCRM/1.0"

    def log_message(self, fmt, *args):
        pass  # silencioso (evita poluir o terminal com cada request)

    # -- helpers de resposta --
    def headers_seguranca(self):
        """Cabecalhos de protecao do navegador, em TODA resposta:
        - nosniff: nada de adivinhar tipo de arquivo (bloqueia truques de upload)
        - X-Frame/frame-ancestors: o CRM nao pode ser embutido em site alheio
          (anti-clickjacking)
        - CSP: so scripts/estilos do proprio CRM rodam; imagens/midia tambem de
          https (fotos e audios do Chatwoot). Em localhost libera http (dev)."""
        host = (self.headers.get("Host") or "").split(":")[0]
        midia = "'self' data: blob: https:" + (" http:" if host in ("localhost", "127.0.0.1") else "")
        self.send_header("X-Content-Type-Options", "nosniff")
        self.send_header("X-Frame-Options", "DENY")
        self.send_header("Referrer-Policy", "same-origin")
        self.send_header("Content-Security-Policy",
                         "default-src 'self'; script-src 'self' 'unsafe-inline'; "
                         "style-src 'self' 'unsafe-inline'; img-src %s; media-src %s; "
                         "connect-src 'self'; frame-ancestors 'none'; base-uri 'self'; "
                         "form-action 'self'" % (midia, midia))

    def send_json(self, status, obj, headers=None):
        body = json.dumps(obj, ensure_ascii=False).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(body)))
        self.headers_seguranca()
        for k, v in (headers or {}).items():
            self.send_header(k, v)
        self.end_headers()
        self.wfile.write(body)

    def _cookie(self, nome):
        raw = self.headers.get("Cookie") or ""
        for parte in raw.split(";"):
            k, _, v = parte.strip().partition("=")
            if k == nome:
                return v
        return None

    def read_body(self):
        length = int(self.headers.get("Content-Length") or 0)
        if length <= 0:
            return {}
        if length > 15 * 1024 * 1024:  # fotos de visita chegam em base64
            raise ValueError("corpo muito grande")
        raw = self.rfile.read(length)
        if not raw:
            return {}
        data = json.loads(raw.decode("utf-8"))
        # As rotas assumem objeto; lista/string/numero viraria AttributeError
        if not isinstance(data, dict):
            raise ValueError("esperado um objeto JSON")
        return data

    # -- Atlas de prospeccao (embutido: consulta o data/atlas.db local) --
    def _atlas_escopo(self, qs, prefixo="f."):
        """WHERE conforme municipio, territorio, area minima e cultura."""
        cond, params = ["1=1"], []
        mun = (qs.get("municipio_id") or [""])[0]
        ter = (qs.get("territorio_id") or [""])[0]
        if ter.isdigit():
            cond.append(prefixo + "municipio_id IN (SELECT municipio_id FROM "
                        "territorio_municipios WHERE territorio_id = ?)")
            params.append(int(ter))
        elif mun.isdigit():
            cond.append(prefixo + "municipio_id = ?")
            params.append(int(mun))
        amin = (qs.get("area_min") or [""])[0]
        try:
            if amin:
                cond.append(prefixo + "area_total_ha >= ?")
                params.append(float(amin))
        except ValueError:
            pass
        cult = (qs.get("cultura") or [""])[0]
        if cult.isdigit():
            try:
                cmin = float((qs.get("cultura_min") or ["0"])[0] or 0)
            except ValueError:
                cmin = 0.0
            cond.append(prefixo + "id IN (SELECT fc.fazenda_id FROM fazenda_culturas fc "
                        "WHERE fc.cultura_id = ? AND fc.area_ha >= ? "
                        "AND fc.safra = (SELECT MAX(safra) FROM fazenda_culturas))")
            params += [int(cult), cmin]
        return " AND ".join(cond), params

    def _multipart_arquivo(self):
        """Extrai (nome, bytes) do primeiro arquivo de um POST multipart."""
        ctype = self.headers.get("Content-Type") or ""
        m = re.search(r'boundary="?([^";]+)"?', ctype)
        tam = int(self.headers.get("Content-Length") or 0)
        if not m or tam <= 0 or tam > 30 * 1024 * 1024:
            return None, None
        corpo = self.rfile.read(tam)
        for parte in corpo.split(b"--" + m.group(1).encode()):
            if b"filename=" not in parte:
                continue
            cab, _, dados = parte.partition(b"\r\n\r\n")
            mm = re.search(rb'filename="([^"]*)"', cab)
            nome = mm.group(1).decode("utf-8", "replace") if mm else "arquivo"
            if dados.endswith(b"\r\n"):
                dados = dados[:-2]
            return nome, dados
        return None, None

    def handle_atlas(self, method, path, qs, user, gestor):
        """API do Atlas de prospeccao (242 mil fazendas de Goias), servida do
        data/atlas.db. Portada do atlas-agro/app.py — mesmas rotas e formatos.
        Sem a tabela de contornos (versao compacta), o mapa cai para pontos."""
        if not os.path.exists(ATLAS_DB):
            return self.send_json(503, {"erro": "O Atlas ainda não está instalado neste servidor"})
        # Prospeccao e trabalho de venda: SDR nao acessa (a API guarda CPF/
        # telefone de produtores — mesmo modelo de acesso dos leads)
        if user["papel"] not in ("admin", "gerente", "vendedor"):
            return self.send_json(403, {"erro": "A Prospecção é para vendedores e gestores"})
        con = atlas_con()
        try:
            # contornos: tabela nova (delta, compacta) ou a antiga (geojson)
            tabs = {r["name"] for r in con.execute(
                "SELECT name FROM sqlite_master WHERE type='table'").fetchall()}
            contorno_delta = "fazenda_contorno_d" in tabs
            tem_contorno = contorno_delta or "fazenda_contorno" in tabs

            if path == "/atlas-api/pessoas" and method == "GET":
                # Quem são os donos: ranking por área, busca por nome/sobrenome
                # (ignorando acento) e o resumo do grupo — é o cruzamento de
                # "família": todo Romualdo/Francelin que tem fazenda.
                # Com uma cultura escolhida, traz também quanto cada produtor
                # tem dela (somando as fazendas dele) e o potencial em R$.
                busca = ((qs.get("busca") or [""])[0]).strip()
                ordenar = (qs.get("ordenar") or ["area"])[0]
                try:
                    pagina = max(1, int((qs.get("pagina") or ["1"])[0]))
                except ValueError:
                    pagina = 1
                w, params = self._atlas_escopo(qs)
                cond = [w]
                cat = (qs.get("categoria") or [""])[0]
                if cat:
                    cond.append("f.categoria = ?")
                    params = params + [cat]
                if busca:
                    # CPF/CNPJ só entra na busca quando o termo tem dígitos: um
                    # valor "vazio" no LIKE casaria com todo mundo sem documento
                    digitos = re.sub(r"\D", "", busca)
                    if digitos:
                        cond.append("(semacento(p.nome) LIKE semacento(?) OR REPLACE(REPLACE("
                                    "REPLACE(COALESCE(p.documento,''),'.',''),'-',''),'/','') LIKE ?)")
                        params = params + ["%" + busca + "%", "%" + digitos + "%"]
                    else:
                        cond.append("semacento(p.nome) LIKE semacento(?)")
                        params = params + ["%" + busca + "%"]
                wtudo = " AND ".join(cond)
                base = ("FROM pessoas p JOIN fazenda_pessoas fp ON fp.pessoa_id = p.id "
                        "JOIN fazendas f ON f.id = fp.fazenda_id "
                        "LEFT JOIN municipios m ON m.id = f.municipio_id WHERE " + wtudo)
                # área de UMA cultura na fazenda da linha (subconsulta por fazenda:
                # juntar fazenda_culturas direto multiplicaria as linhas e a área)
                col_cult = ("(SELECT COALESCE(SUM(fc.area_ha),0) FROM fazenda_culturas fc "
                            "WHERE fc.fazenda_id = f.id AND fc.cultura_id = ? "
                            "AND fc.safra = (SELECT MAX(safra) FROM fazenda_culturas))")
                # quantas pessoas estao ligadas a mesma fazenda: em espolio e
                # assentamento sao dezenas, e cada uma apareceria como dona da
                # area inteira — quem vende precisa enxergar isso
                col_soc = "(SELECT COUNT(*) FROM fazenda_pessoas x WHERE x.fazenda_id = f.id)"
                msoc = (qs.get("max_socios") or [""])[0]
                if msoc.isdigit() and int(msoc) > 0:
                    cond.append(col_soc + " <= ?")
                    params = params + [int(msoc)]
                    wtudo = " AND ".join(cond)
                    base = ("FROM pessoas p JOIN fazenda_pessoas fp ON fp.pessoa_id = p.id "
                            "JOIN fazendas f ON f.id = fp.fazenda_id "
                            "LEFT JOIN municipios m ON m.id = f.municipio_id WHERE " + wtudo)
                sel, psel = "", []
                cult = (qs.get("cultura") or [""])[0]
                if cult.isdigit():
                    sel += ", ROUND(SUM(" + col_cult + ")) AS area_cultura_ha"
                    psel.append(int(cult))
                regras = [r for r in potencial_regras() if r["ativa"]][:6]
                for i, r in enumerate(regras):
                    sel += ", SUM(" + col_cult + ") AS pot%d" % i
                    psel.append(r["cultura_id"])
                # mínimo da cultura somando TODAS as fazendas do produtor
                # (o filtro de cima é por fazenda; este é por pessoa)
                having = ""
                phav = []
                ctmin = (qs.get("cultura_total_min") or [""])[0]
                if cult.isdigit() and ctmin:
                    try:
                        having = " HAVING SUM(" + col_cult + ") >= ?"
                        phav = [int(cult), float(ctmin)]
                    except ValueError:
                        having, phav = "", []
                ordem = {"area": "SUM(f.area_total_ha) DESC",
                         "fazendas": "COUNT(DISTINCT f.id) DESC",
                         "nome": "p.nome COLLATE NOCASE"}.get(ordenar, "SUM(f.area_total_ha) DESC")
                if ordenar == "cultura" and cult.isdigit():
                    ordem = "area_cultura_ha DESC"
                rows = con.execute(
                    "SELECT p.id, p.nome, p.documento, p.telefone, p.email, "
                    "COUNT(DISTINCT f.id) AS fazendas, "
                    "ROUND(SUM(f.area_total_ha)) AS area_ha, "
                    "GROUP_CONCAT(DISTINCT m.nome) AS municipios, "
                    "MAX(" + col_soc + ") AS socios_max" + sel + " " + base +
                    " GROUP BY p.id" + having + " ORDER BY " + ordem + " LIMIT ? OFFSET ?",
                    psel + params + phav + [25, (pagina - 1) * 25]).fetchall()
                pessoas = []
                for r in rows:
                    d = dict(r)
                    total, itens = 0.0, []
                    for i, g in enumerate(regras):
                        un = int((d.pop("pot%d" % i, 0) or 0) // g["ha_por_unidade"])
                        if un > 0:
                            total += un * g["valor_unidade"]
                            itens.append({"produto": g["produto"], "unidades": un})
                    d["potencial_rs"] = round(total, 2)
                    d["potencial_itens"] = itens
                    pessoas.append(d)
                # Quando há mínimo por produtor, os totais só podem contar
                # quem passou no corte — senão o resumo mostraria a região toda.
                filtro_pes, pfil = "", []
                if having:
                    filtro_pes = " AND p.id IN (SELECT p.id " + base + " GROUP BY p.id" + having + ")"
                    pfil = params + phav
                # totais do grupo sobre fazendas DISTINTAS: uma fazenda com dois
                # donos apareceria duas vezes e dobraria a área da família
                res_p = con.execute("SELECT COUNT(DISTINCT p.id) AS pessoas " + base + filtro_pes,
                                    params + pfil).fetchone()
                colca = (", " + col_cult + " AS ca") if cult.isdigit() else ""
                res_f = con.execute(
                    "SELECT COUNT(*) AS fazendas, ROUND(SUM(area_total_ha)) AS area_ha"
                    + (", ROUND(SUM(ca)) AS area_cultura_ha" if cult.isdigit() else "") +
                    " FROM (SELECT DISTINCT f.id, f.area_total_ha" + colca + " " +
                    base + filtro_pes + ")",
                    (psel[:1] if cult.isdigit() else []) + params + pfil).fetchone()
                resumo = dict(res_p or {})
                resumo.update(dict(res_f or {}))
                cidades = con.execute(
                    "SELECT m.nome, COUNT(DISTINCT f.id) AS fazendas, "
                    "COUNT(DISTINCT p.id) AS pessoas " + base + filtro_pes +
                    " AND m.nome IS NOT NULL GROUP BY m.id "
                    "ORDER BY COUNT(DISTINCT f.id) DESC LIMIT 12", params + pfil).fetchall()
                # O Atlas so tem NOME de dono em parte dos municipios (o SICAR
                # nao publica proprietario). Sem esse aviso, o painel vazio
                # parece defeito quando na verdade e falta de dado na regiao.
                cobertura = None
                if not resumo.get("pessoas"):
                    cob = con.execute(
                        "SELECT COUNT(DISTINCT f.municipio_id) AS municipios, "
                        "COUNT(DISTINCT fp.pessoa_id) AS pessoas "
                        "FROM fazenda_pessoas fp JOIN fazendas f ON f.id = fp.fazenda_id").fetchone()
                    cobertura = dict(cob) if cob else None
                return self.send_json(200, {
                    "pessoas": pessoas,
                    "pagina": pagina, "por_pagina": 25,
                    "resumo": resumo,
                    "cidades": [dict(c) for c in cidades],
                    "cobertura": cobertura,
                    "busca": busca})

            mpes = re.match(r"^/atlas-api/pessoa/(\d+)$", path)
            if mpes and method == "GET":
                pid = int(mpes.group(1))
                p = con.execute("SELECT * FROM pessoas WHERE id = ?", [pid]).fetchone()
                if not p:
                    return self.send_json(404, {"erro": "não encontrada"})
                faz = con.execute(
                    "SELECT f.id, f.nome, f.codigo_car, f.area_total_ha, f.categoria, "
                    "f.latitude, f.longitude, m.nome AS municipio, fp.relacao_fundiaria "
                    "FROM fazenda_pessoas fp JOIN fazendas f ON f.id = fp.fazenda_id "
                    "LEFT JOIN municipios m ON m.id = f.municipio_id "
                    "WHERE fp.pessoa_id = ? ORDER BY f.area_total_ha DESC", [pid]).fetchall()
                # outras pessoas com o mesmo sobrenome (possíveis parentes)
                partes = [x for x in _sem_acento(p["nome"]).split() if len(x) > 3
                          and x not in ("DOS", "DAS", "DER", "JUNIOR", "NETO", "FILHO", "SOBRINHO")]
                sobrenome = partes[-1] if partes else ""
                parentes = []
                if sobrenome:
                    parentes = con.execute(
                        "SELECT p2.id, p2.nome, COUNT(DISTINCT f.id) AS fazendas, "
                        "ROUND(SUM(f.area_total_ha)) AS area_ha "
                        "FROM pessoas p2 JOIN fazenda_pessoas fp ON fp.pessoa_id = p2.id "
                        "JOIN fazendas f ON f.id = fp.fazenda_id "
                        "WHERE p2.id <> ? AND semacento(p2.nome) LIKE ? "
                        "GROUP BY p2.id ORDER BY SUM(f.area_total_ha) DESC LIMIT 25",
                        [pid, "%" + sobrenome + "%"]).fetchall()
                return self.send_json(200, {"pessoa": dict(p),
                                            "fazendas": [dict(f) for f in faz],
                                            "sobrenome": sobrenome,
                                            "parentes": [dict(x) for x in parentes]})

            if path == "/atlas-api/municipios" and method == "GET":
                rows = con.execute(
                    """SELECT m.id, m.nome, m.uf, COUNT(f.id) AS fazendas,
                              ROUND(COALESCE(SUM(f.area_total_ha),0),0) AS area_ha
                       FROM municipios m LEFT JOIN fazendas f ON f.municipio_id = m.id
                       GROUP BY m.id ORDER BY m.nome""").fetchall()
                return self.send_json(200, [dict(r) for r in rows])

            if path == "/atlas-api/territorios" and method == "GET":
                rows = con.execute(
                    """SELECT t.id, t.nome, COUNT(DISTINCT tm.municipio_id) AS municipios,
                              COUNT(f.id) AS fazendas,
                              ROUND(COALESCE(SUM(f.area_total_ha),0),0) AS area_ha
                       FROM territorios t
                       LEFT JOIN territorio_municipios tm ON tm.territorio_id = t.id
                       LEFT JOIN fazendas f ON f.municipio_id = tm.municipio_id
                       GROUP BY t.id ORDER BY t.nome""").fetchall()
                return self.send_json(200, [dict(r) for r in rows])

            if path == "/atlas-api/territorios" and method == "POST":
                if not gestor:
                    return self.send_json(403, {"erro": "Só gerente/administrador cria territórios"})
                d = self.read_body()
                nome = str(d.get("nome") or "").strip()[:80]
                muns = [int(m) for m in (d.get("municipios") or []) if str(m).isdigit()]
                if not nome or not muns:
                    return self.send_json(400, {"erro": "informe nome e ao menos um município"})
                cur = con.execute("INSERT INTO territorios (nome) VALUES (?)", [nome])
                con.executemany(
                    "INSERT OR IGNORE INTO territorio_municipios (territorio_id, municipio_id) VALUES (?,?)",
                    [(cur.lastrowid, m) for m in muns])
                con.commit()
                return self.send_json(200, {"ok": True, "id": cur.lastrowid})

            mter = re.match(r"^/atlas-api/territorios/(\d+)$", path)
            if mter and method == "DELETE":
                if not gestor:
                    return self.send_json(403, {"erro": "Só gerente/administrador exclui territórios"})
                con.execute("DELETE FROM territorio_municipios WHERE territorio_id = ?", [int(mter.group(1))])
                con.execute("DELETE FROM territorios WHERE id = ?", [int(mter.group(1))])
                con.commit()
                return self.send_json(200, {"ok": True})

            if path == "/atlas-api/pivos" and method == "GET":
                # Pivos centrais do recorte. No mapa cada um vira um circulo,
                # entao devolvemos centro + raio calculado da area.
                if "pivos" not in tabs:
                    return self.send_json(200, {"pivos": [], "sem_pivos": True,
                                                "total": 0, "area_ha": 0})
                cond, params = ["1=1"], []
                mun = (qs.get("municipio_id") or [""])[0]
                ter = (qs.get("territorio_id") or [""])[0]
                if ter.isdigit():
                    cond.append("municipio_id IN (SELECT municipio_id FROM "
                                "territorio_municipios WHERE territorio_id = ?)")
                    params.append(int(ter))
                elif mun.isdigit():
                    cond.append("municipio_id = ?")
                    params.append(int(mun))
                try:   # recorte do mapa (opcional)
                    n = float((qs.get("norte") or [""])[0]); s = float((qs.get("sul") or [""])[0])
                    le = float((qs.get("leste") or [""])[0]); o = float((qs.get("oeste") or [""])[0])
                    cond.append("lat BETWEEN ? AND ?")
                    cond.append("lng BETWEEN ? AND ?")
                    params += [s, n, o, le]
                except (ValueError, IndexError):
                    pass
                w = " AND ".join(cond)
                tot = con.execute("SELECT COUNT(*) AS n, ROUND(COALESCE(SUM(ha),0)) AS ha "
                                  "FROM pivos WHERE " + w, params).fetchone()
                rows = con.execute(
                    "SELECT p.id, p.lat, p.lng, p.ha, p.ano, m.nome AS municipio "
                    "FROM pivos p LEFT JOIN municipios m ON m.id = p.municipio_id "
                    "WHERE " + w + " ORDER BY p.ha DESC LIMIT 4000", params).fetchall()
                pivos = []
                for r in rows:
                    ha = float(r["ha"] or 0)
                    # circulo de area equivalente: raio = raiz(area / pi)
                    raio = int(math.sqrt(max(ha, 0.1) * 10000 / math.pi))
                    d = dict(r)
                    d["raio_m"] = raio
                    pivos.append(d)
                return self.send_json(200, {
                    "pivos": pivos, "total": tot["n"] or 0, "area_ha": tot["ha"] or 0,
                    "mostrando": len(pivos),
                    "fonte": "ANA — pivôs centrais mapeados até 2019"})

            if path == "/atlas-api/potencial" and method == "GET":
                # Potencial comercial do recorte: "a cada X ha da cultura Y
                # vendo 1 Z de R$ W". Sao DUAS leituras, e as duas importam:
                #  - imediato: a conta feita POR FAZENDA (quem sozinho ja
                #    justifica a compra) — é a lista de quem visitar hoje;
                #  - da regiao: a lavoura toda dividida pela regra — o tamanho
                #    do mercado, incluindo quem só fecha via prestador/vizinho.
                regras = [r for r in potencial_regras() if r["ativa"]]
                w, p = self._atlas_escopo(qs)
                safra = " AND fc.safra = (SELECT MAX(safra) FROM fazenda_culturas) AND "
                itens, total_rs, total_reg, por_mun = [], 0.0, 0.0, {}
                for r in regras:
                    ha_un = r["ha_por_unidade"]
                    row = con.execute(
                        "SELECT COUNT(DISTINCT f.id) AS fazendas, COALESCE(SUM(fc.area_ha),0) AS area_ha, "
                        "COALESCE(SUM(CAST(fc.area_ha / ? AS INTEGER)),0) AS unidades "
                        "FROM fazenda_culturas fc JOIN fazendas f ON f.id = fc.fazenda_id "
                        "WHERE fc.cultura_id = ?" + safra + w,
                        [ha_un, r["cultura_id"]] + p).fetchone()
                    un = int(row["unidades"] or 0)
                    area = float(row["area_ha"] or 0)
                    un_reg = int(area // ha_un)
                    total_rs += un * r["valor_unidade"]
                    total_reg += un_reg * r["valor_unidade"]
                    itens.append({"id": r["id"], "produto": r["produto"], "cultura": r["cultura"],
                                  "cultura_id": r["cultura_id"], "ha_por_unidade": ha_un,
                                  "valor_unidade": r["valor_unidade"],
                                  "unidades": un, "valor": un * r["valor_unidade"],
                                  "unidades_regiao": un_reg,
                                  "valor_regiao": un_reg * r["valor_unidade"],
                                  "area_ha": round(area), "fazendas": row["fazendas"] or 0})
                    for m in con.execute(
                            "SELECT m.id, m.nome, COALESCE(SUM(fc.area_ha),0) AS area, "
                            "COALESCE(SUM(CAST(fc.area_ha / ? AS INTEGER)),0) AS un "
                            "FROM fazenda_culturas fc JOIN fazendas f ON f.id = fc.fazenda_id "
                            "JOIN municipios m ON m.id = f.municipio_id "
                            "WHERE fc.cultura_id = ?" + safra + w + " GROUP BY m.id",
                            [ha_un, r["cultura_id"]] + p).fetchall():
                        u, ureg = int(m["un"] or 0), int(float(m["area"] or 0) // ha_un)
                        if u <= 0 and ureg <= 0:
                            continue
                        d = por_mun.setdefault(m["id"], {"id": m["id"], "nome": m["nome"],
                                                         "valor": 0.0, "valor_regiao": 0.0,
                                                         "itens": []})
                        d["valor"] += u * r["valor_unidade"]
                        d["valor_regiao"] += ureg * r["valor_unidade"]
                        d["itens"].append({"produto": r["produto"], "unidades": u,
                                           "unidades_regiao": ureg})
                municipios = sorted(por_mun.values(),
                                    key=lambda x: x["valor_regiao"], reverse=True)[:15]
                return self.send_json(200, {"regras": itens,
                                            "total_rs": round(total_rs, 2),
                                            "total_regiao_rs": round(total_reg, 2),
                                            "municipios": municipios,
                                            "sem_regras": not regras})

            if path == "/atlas-api/culturas" and method == "GET":
                w, p = self._atlas_escopo(qs)
                rows = con.execute(
                    "SELECT c.id, c.nome, COUNT(*) AS fazendas, ROUND(SUM(fc.area_ha)) AS area_ha "
                    "FROM fazenda_culturas fc JOIN culturas c ON c.id = fc.cultura_id "
                    "JOIN fazendas f ON f.id = fc.fazenda_id "
                    "WHERE fc.safra = (SELECT MAX(safra) FROM fazenda_culturas) AND " + w +
                    " GROUP BY c.id ORDER BY SUM(fc.area_ha) DESC", p).fetchall()
                total = con.execute("SELECT COUNT(*) c FROM fazendas f WHERE " + w, p).fetchone()["c"]
                return self.send_json(200, {"total_fazendas": total,
                                            "culturas": [dict(r) for r in rows]})

            if path == "/atlas-api/resumo" and method == "GET":
                w, p = self._atlas_escopo(qs, "")
                tot = con.execute(
                    "SELECT COUNT(*) AS fazendas, ROUND(COALESCE(SUM(area_total_ha),0),0) AS area_ha, "
                    "ROUND(COALESCE(AVG(area_total_ha),0),1) AS tamanho_medio FROM fazendas WHERE " + w,
                    p).fetchone()
                cats = con.execute(
                    "SELECT categoria, COUNT(*) qtde FROM fazendas WHERE " + w + " GROUP BY categoria",
                    p).fetchall()
                return self.send_json(200, {"totais": dict(tot),
                                            "categorias": [dict(c) for c in cats]})

            if path == "/atlas-api/fazendas" and method == "GET":
                busca = ((qs.get("busca") or [""])[0]).strip()
                cat = (qs.get("categoria") or [""])[0]
                try:
                    pagina = max(1, int((qs.get("pagina") or ["1"])[0]))
                except ValueError:
                    pagina = 1
                w, params = self._atlas_escopo(qs)
                where = [w]
                if cat:
                    where.append("f.categoria = ?")
                    params.append(cat)
                if busca:
                    where.append("(f.nome LIKE ? OR f.codigo_car LIKE ?)")
                    params += ["%" + busca + "%", "%" + busca + "%"]
                wtudo = " AND ".join(where)
                total = con.execute("SELECT COUNT(*) c FROM fazendas f WHERE " + wtudo, params).fetchone()["c"]
                rows = con.execute(
                    "SELECT f.id, f.nome, f.codigo_car, f.categoria, f.area_total_ha, "
                    "f.perimetro_km, f.status_car, f.latitude, f.longitude, m.nome AS municipio, "
                    "(SELECT p.nome FROM fazenda_pessoas fp JOIN pessoas p ON p.id = fp.pessoa_id "
                    " WHERE fp.fazenda_id = f.id LIMIT 1) AS dono "
                    "FROM fazendas f LEFT JOIN municipios m ON m.id = f.municipio_id "
                    "WHERE " + wtudo + " ORDER BY f.area_total_ha DESC LIMIT ? OFFSET ?",
                    params + [20, (pagina - 1) * 20]).fetchall()
                return self.send_json(200, {"total": total, "pagina": pagina, "por_pagina": 20,
                                            "fazendas": [dict(r) for r in rows]})

            if path == "/atlas-api/pontos" and method == "GET":
                w, params = self._atlas_escopo(qs)   # prefixo f. (a consulta usa alias)
                cat_p = (qs.get("categoria") or [""])[0]
                if cat_p:                      # o mapa tem de respeitar o filtro
                    w += " AND f.categoria = ?"
                    params = params + [cat_p]
                rows = con.execute(
                    "SELECT f.id, f.nome, f.latitude, f.longitude, f.area_total_ha, f.categoria, "
                    "(SELECT p.nome FROM fazenda_pessoas fp JOIN pessoas p ON p.id = fp.pessoa_id "
                    " WHERE fp.fazenda_id = f.id LIMIT 1) AS dono "
                    "FROM fazendas f WHERE f.latitude IS NOT NULL AND " + w +
                    " ORDER BY f.area_total_ha DESC LIMIT 3000", params).fetchall()
                return self.send_json(200, [dict(r) for r in rows])

            if path == "/atlas-api/limite" and method == "GET":
                # contorno (limite) dos municipios do recorte atual — a linha
                # que delimita a cidade selecionada ou o territorio inteiro
                if not con.execute("SELECT 1 FROM sqlite_master WHERE type='table' "
                                   "AND name='municipio_contorno'").fetchone():
                    return self.send_json(200, {"limites": [], "sem_limites": True})
                mun = (qs.get("municipio_id") or [""])[0]
                ter = (qs.get("territorio_id") or [""])[0]
                if ter.isdigit():
                    rows = con.execute(
                        "SELECT m.nome, c.geojson FROM municipio_contorno c "
                        "JOIN municipios m ON m.id = c.municipio_id "
                        "WHERE c.municipio_id IN (SELECT municipio_id FROM "
                        "territorio_municipios WHERE territorio_id = ?)", [int(ter)]).fetchall()
                elif mun.isdigit():
                    rows = con.execute(
                        "SELECT m.nome, c.geojson FROM municipio_contorno c "
                        "JOIN municipios m ON m.id = c.municipio_id "
                        "WHERE c.municipio_id = ?", [int(mun)]).fetchall()
                else:
                    return self.send_json(200, {"limites": []})   # Goiás inteiro: sem linha
                saida = []
                for r in rows:
                    try:
                        saida.append({"nome": r["nome"], "aneis": json.loads(r["geojson"])})
                    except Exception:
                        continue
                return self.send_json(200, {"limites": saida})

            if path == "/atlas-api/mapa-municipios" and method == "GET":
                # visao de longe: UMA bolha por municipio com a contagem —
                # assim o mapa mostra TODAS as fazendas, nao so as 3.000 maiores
                w, params = self._atlas_escopo(qs)
                cat = (qs.get("categoria") or [""])[0]
                if cat:
                    w += " AND f.categoria = ?"
                    params = params + [cat]
                rows = con.execute(
                    "SELECT m.id, m.nome, COUNT(*) AS fazendas, "
                    "ROUND(SUM(f.area_total_ha)) AS area_ha, "
                    "AVG(f.latitude) AS lat, AVG(f.longitude) AS lng "
                    "FROM fazendas f JOIN municipios m ON m.id = f.municipio_id "
                    "WHERE f.latitude IS NOT NULL AND " + w +
                    " GROUP BY m.id ORDER BY COUNT(*) DESC", params).fetchall()
                return self.send_json(200, [dict(r) for r in rows])

            if path == "/atlas-api/contornos" and method == "GET":
                if not tem_contorno:
                    return self.send_json(200, {"fazendas": [], "total_na_area": 0,
                                                "mostrando": 0, "sem_contornos": True})
                try:
                    n = float((qs.get("norte") or [""])[0]); s = float((qs.get("sul") or [""])[0])
                    le = float((qs.get("leste") or [""])[0]); o = float((qs.get("oeste") or [""])[0])
                except ValueError:
                    return self.send_json(400, {"fazendas": [], "erro": "área do mapa não informada"})
                w, params = self._atlas_escopo(qs)
                cond = [w, "f.latitude BETWEEN ? AND ?", "f.longitude BETWEEN ? AND ?"]
                params = params + [s, n, o, le]
                cat = (qs.get("categoria") or [""])[0]
                if cat:
                    cond.append("f.categoria = ?")
                    params.append(cat)
                tab, col = ("fazenda_contorno_d", "d") if contorno_delta else ("fazenda_contorno", "geojson")
                rows = con.execute(
                    "SELECT f.id, f.nome, f.area_total_ha, f.perimetro_km, f.categoria, c.%s AS geo, "
                    "(SELECT p.nome FROM fazenda_pessoas fp JOIN pessoas p ON p.id = fp.pessoa_id "
                    " WHERE fp.fazenda_id = f.id LIMIT 1) AS dono "
                    "FROM fazendas f JOIN %s c ON c.fazenda_id = f.id "
                    "WHERE " % (col, tab) + " AND ".join(cond) +
                    " ORDER BY f.area_total_ha DESC LIMIT 1200", params).fetchall()
                total = con.execute(
                    "SELECT COUNT(*) c FROM fazendas f WHERE " + " AND ".join(cond), params).fetchone()["c"]
                saida = []
                for r in rows:
                    try:
                        if contorno_delta:
                            aneis = _atlas_decodifica(r["geo"])
                        else:
                            aneis = [_atlas_simplifica(a) for a in json.loads(r["geo"])]
                    except Exception:
                        continue
                    saida.append({"id": r["id"], "nome": r["nome"],
                                  "area_total_ha": r["area_total_ha"],
                                  "perimetro_km": r["perimetro_km"],
                                  "categoria": r["categoria"], "dono": r["dono"],
                                  "aneis": aneis})
                return self.send_json(200, {"fazendas": saida, "total_na_area": total,
                                            "mostrando": len(saida)})

            mfz = re.match(r"^/atlas-api/fazenda/(\d+)(?:/(\w+))?$", path)
            if mfz:
                fid = int(mfz.group(1))
                sub = mfz.group(2) or ""
                if not con.execute("SELECT 1 FROM fazendas WHERE id = ?", [fid]).fetchone():
                    return self.send_json(404, {"erro": "não encontrada"})
                if not sub and method == "GET":
                    f = con.execute(
                        "SELECT f.*, m.nome AS municipio, m.uf FROM fazendas f "
                        "LEFT JOIN municipios m ON m.id = f.municipio_id WHERE f.id = ?",
                        [fid]).fetchone()
                    if not f:
                        return self.send_json(404, {"erro": "não encontrada"})
                    pessoas = con.execute(
                        "SELECT p.*, fp.relacao_fundiaria, fp.relacao_comercial "
                        "FROM fazenda_pessoas fp JOIN pessoas p ON p.id = fp.pessoa_id "
                        "WHERE fp.fazenda_id = ?", [fid]).fetchall()
                    registros = con.execute(
                        "SELECT * FROM registros WHERE fazenda_id = ? ORDER BY criado_em DESC",
                        [fid]).fetchall()
                    culturas = con.execute(
                        "SELECT c.nome, fc.safra, fc.area_ha FROM fazenda_culturas fc "
                        "JOIN culturas c ON c.id = fc.cultura_id WHERE fc.fazenda_id = ? "
                        "ORDER BY fc.safra DESC, fc.area_ha DESC", [fid]).fetchall()
                    return self.send_json(200, {"fazenda": dict(f),
                                                "pessoas": [dict(p) for p in pessoas],
                                                "registros": [dict(r) for r in registros],
                                                "culturas": [dict(c) for c in culturas]})
                if sub == "contorno" and method == "GET":
                    if contorno_delta:
                        r = con.execute("SELECT d FROM fazenda_contorno_d WHERE fazenda_id = ?",
                                        [fid]).fetchone()
                        if r:
                            return self.send_json(200, {"aneis": _atlas_decodifica(r["d"])})
                    elif tem_contorno:
                        r = con.execute("SELECT geojson FROM fazenda_contorno WHERE fazenda_id = ?",
                                        [fid]).fetchone()
                        if r:
                            try:
                                return self.send_json(200, {"aneis": json.loads(r["geojson"])})
                            except Exception:
                                pass
                    return self.send_json(200, {"aneis": []})
                if sub == "categoria" and method == "POST":
                    cat = (self.read_body().get("categoria") or "")
                    if cat not in ("cliente", "lead", "descartada", "sem_categoria"):
                        return self.send_json(400, {"erro": "categoria inválida"})
                    con.execute("UPDATE fazendas SET categoria = ? WHERE id = ?", [cat, fid])
                    con.commit()
                    return self.send_json(200, {"ok": True})
                if sub == "registro" and method == "POST":
                    texto = str(self.read_body().get("texto") or "").strip()[:2000]
                    if not texto:
                        return self.send_json(400, {"erro": "texto vazio"})
                    con.execute("INSERT INTO registros (fazenda_id, autor, texto) VALUES (?,?,?)",
                                [fid, user["nome"], texto])
                    con.commit()
                    return self.send_json(200, {"ok": True})
                if sub == "proprietario" and method == "POST":
                    d = self.read_body()
                    nome = str(d.get("nome") or "").strip()[:120]
                    if not nome:
                        return self.send_json(400, {"erro": "nome vazio"})
                    doc = str(d.get("documento") or "").strip()[:20] or None
                    tipo = "juridica" if doc and len(re.sub(r"\D", "", doc)) == 14 else "fisica"
                    cur = con.execute(
                        "INSERT INTO pessoas (nome, tipo, documento, telefone, email) VALUES (?,?,?,?,?)",
                        [nome, tipo, doc, str(d.get("telefone") or "").strip()[:40] or None,
                         str(d.get("email") or "").strip()[:120] or None])
                    con.execute(
                        "INSERT OR IGNORE INTO fazenda_pessoas "
                        "(fazenda_id, pessoa_id, relacao_fundiaria, relacao_comercial) VALUES (?,?,?,?)",
                        [fid, cur.lastrowid, str(d.get("relacao_fundiaria") or "proprietario")[:30],
                         str(d.get("relacao_comercial") or "nao_definida")[:30]])
                    con.commit()
                    return self.send_json(200, {"ok": True})

            if path == "/atlas-api/importar/analisar" and method == "POST":
                if not gestor:
                    return self.send_json(403, {"erro": "Só gerente/administrador importa planilhas"})
                nomearq, dados = self._multipart_arquivo()
                if not nomearq or dados is None:
                    return self.send_json(400, {"erro": "nenhum arquivo enviado"})
                ext = os.path.splitext(nomearq)[1].lower()
                if ext not in (".csv", ".xlsx", ".xlsm"):
                    return self.send_json(400, {"erro": "formato não suportado — use .xlsx ou .csv"})
                try:
                    abas, cab, linhas = _atlas_le_planilha(nomearq, dados)
                except Exception as e:
                    return self.send_json(400, {"erro": "não consegui ler o arquivo: %s" % e})
                token = secrets.token_hex(12)
                # guarda em memoria (max 3 analises simultaneas; limpa as velhas)
                while len(_atlas_imports) >= 3:
                    _atlas_imports.pop(next(iter(_atlas_imports)))
                _atlas_imports[token] = {"linhas": linhas}
                amostra = [[("" if v is None else str(v))[:40] for v in l] for l in linhas[:5]]
                return self.send_json(200, {"token": token, "arquivo": nomearq, "abas": abas,
                                            "colunas": cab, "total_linhas": len(linhas),
                                            "amostra": amostra})

            if path == "/atlas-api/importar/executar" and method == "POST":
                if not gestor:
                    return self.send_json(403, {"erro": "Só gerente/administrador importa planilhas"})
                d = self.read_body()
                reg = _atlas_imports.pop(str(d.get("token") or ""), None)
                if not reg:
                    return self.send_json(400, {"erro": "sessão de importação expirada — envie o arquivo de novo"})
                res = _atlas_importa(con, reg["linhas"], d.get("mapa") or {},
                                     d.get("categoria") or None,
                                     (str(d.get("origem") or "").strip() or "Importação de planilha"),
                                     tem_contorno)
                con.commit()
                return self.send_json(200, res)
        except sqlite3.Error as e:
            print("Erro no atlas:", e)
            return self.send_json(500, {"erro": "Erro no banco do Atlas"})
        finally:
            con.close()
        return self.send_json(404, {"erro": "Não encontrado"})

    # -- roteamento --
    def do_GET(self):
        parsed = urlparse(self.path)
        path = parsed.path
        if path.startswith("/api/") or path.startswith("/atlas-api/"):
            return self.handle_api("GET", parsed)
        if path == "/prospeccao.html":
            u = usuario_da_sessao(self._cookie("sessao"))
            if not u:
                self.send_response(302)
                self.send_header("Location", "/login.html")
                self.end_headers()
                return
            if u["papel"] not in ("admin", "gerente", "vendedor"):
                self.send_response(302)
                self.send_header("Location", "/")
                self.end_headers()
                return
        return self.serve_static(path)

    def do_POST(self):
        parsed = urlparse(self.path)
        path = parsed.path
        if path == "/webhook/chatwoot":
            return self.handle_webhook(parsed)
        if path.startswith("/api/") or path.startswith("/atlas-api/"):
            return self.handle_api("POST", parsed)
        self.send_json(404, {"error": "Nao encontrado"})

    def do_PATCH(self):
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/"):
            return self.handle_api("PATCH", parsed)
        self.send_json(404, {"error": "Nao encontrado"})

    def do_PUT(self):
        self.do_PATCH()

    def do_DELETE(self):
        parsed = urlparse(self.path)
        if parsed.path.startswith("/api/") or parsed.path.startswith("/atlas-api/"):
            return self.handle_api("DELETE", parsed)
        self.send_json(404, {"error": "Nao encontrado"})

    # -- webhook --
    def handle_webhook(self, parsed):
        qs = parse_qs(parsed.query)
        token = (qs.get("token") or [None])[0] or self.headers.get("X-Webhook-Token")
        if token != WEBHOOK_TOKEN:
            return self.send_json(401, {"error": "Token invalido"})
        # origem=curso -> eventos do Chatwoot do CURSO (instancia separada)
        origem = "curso" if (qs.get("origem") or [""])[0] == "curso" else ""
        # marca que o Chatwoot está entregando eventos (indicador no painel)
        global _webhook_ultimo
        _webhook_ultimo = now_iso()
        try:
            payload = self.read_body()
        except Exception:
            return self.send_json(400, {"error": "Corpo invalido"})
        try:
            result = handle_chatwoot_event(payload, origem)
        except Exception as e:
            print("Erro no webhook:", e)
            result = {"ok": False, "reason": "erro interno"}
        # Responde 200 sempre pro Chatwoot nao reenviar em loop.
        self.send_json(200, result)

    # -- API --
    def handle_api(self, method, parsed):
        # Qualquer erro nao previsto vira 500 com resposta valida, em vez de
        # derrubar a conexao sem status.
        try:
            return self._handle_api(method, parsed)
        except Exception as e:
            print("Erro na API %s %s: %r" % (method, parsed.path, e))
            if not self.wfile.closed:
                try:
                    return self.send_json(500, {"error": "Erro interno"})
                except Exception:
                    pass

    def _handle_api(self, method, parsed):
        path = parsed.path
        qs = parse_qs(parsed.query)

        # ---- Status (aberto): so diz se esta no ar e o que ja foi publicado.
        # Serve para conferir se um deploy chegou sem precisar entrar no CRM.
        # Nao expoe dado nenhum de cliente, usuario ou configuracao. ----
        if path == "/api/status" and method == "GET":
            return self.send_json(200, {
                "ok": True,
                "no_ar_desde": INICIO_ISO,
                "recursos": RECURSOS})

        # ---- Login (unica rota aberta sem sessao) ----
        if path == "/api/login" and method == "POST":
            try:
                body = self.read_body()
            except Exception:
                return self.send_json(400, {"error": "Corpo invalido"})
            login = str(body.get("login") or "").strip().lower()
            senha = str(body.get("senha") or "")
            agora_t = time.time()
            with _lock:
                # forca-bruta: login bloqueado por excesso de tentativas?
                reg = _login_falhas.get(login)
                if reg and reg.get("ate", 0) > agora_t:
                    espera = int((reg["ate"] - agora_t) / 60) + 1
                    return self.send_json(429, {"error":
                        "Muitas tentativas — este login está travado por %d minuto(s). "
                        "Tente de novo depois." % espera})
                u = next((x for x in _db["users"] if x.get("login") == login and x.get("ativo", True)), None)
            # A verificacao PBKDF2 e o atraso ocorrem FORA do lock (senao logins
            # errados em rajada congelariam o CRM inteiro). Quando o login nao
            # existe, roda um hash "fantasma" para o tempo de resposta nao
            # denunciar quais logins sao validos.
            if u and verifica_senha(u, senha):
                with _lock:
                    _login_falhas.pop(login, None)  # acertou: zera o contador
                    # limpeza: sessoes vencidas nao ficam acumulando no banco
                    _db["sessions"] = {t: s for t, s in _db["sessions"].items()
                                       if s.get("exp", 0) > agora_t}
                    token = cria_sessao(u["id"])
                # atras do HTTPS o cookie ganha "Secure" (nunca viaja sem criptografia)
                seguro = "; Secure" if self.headers.get("X-Forwarded-Proto") == "https" else ""
                return self.send_json(200, {"user": user_publico(u)}, headers={
                    "Set-Cookie": "sessao=%s; Path=/; HttpOnly; SameSite=Lax; Max-Age=%d%s" % (
                        token, SESSAO_DIAS * 86400, seguro)})
            if not u:
                hash_senha(senha, "00" * 16)  # equaliza o tempo (login inexistente)
            with _lock:
                reg = _login_falhas.setdefault(login, {"n": 0, "ate": 0})
                reg["n"] += 1
                if reg["n"] >= LOGIN_MAX_FALHAS:
                    reg["ate"] = agora_t + LOGIN_BLOQUEIO_S
                    reg["n"] = 0
                    print("[seguranca] login '%s' travado por %ds (forca-bruta)" % (login, LOGIN_BLOQUEIO_S))
                if len(_login_falhas) > 5000:  # nao crescer sem limite
                    _login_falhas.clear()
            time.sleep(0.4)  # freia adivinhacao de senha, sem segurar o lock
            return self.send_json(401, {"error": "Login ou senha incorretos"})

        # ---- Daqui em diante, toda rota exige sessao valida ----
        with _lock:
            user = usuario_da_sessao(self._cookie("sessao"))
            if user is not None:
                _online[user["id"]] = now_iso()  # marca presenca (qualquer requisicao)
        if user is None:
            return self.send_json(401, {"error": "Sessão expirada — faça login novamente"})
        gestor = user["papel"] in ("admin", "gerente")

        if path == "/api/me" and method == "GET":
            return self.send_json(200, {"user": user_publico(user)})

        # ---- Atlas de prospeccao (fazendas de Goias) ----
        if path.startswith("/atlas-api/"):
            return self.handle_atlas(method, path, qs, user, gestor)

        # ---- Heartbeat: mantem a presenca viva mesmo com um modal aberto ----
        if path == "/api/heartbeat" and method == "GET":
            return self.send_json(200, {"ok": True})

        # ---- Quem esta online + ultimas movimentacoes (so gestor) ----
        if path == "/api/online" and method == "GET":
            if not gestor:
                return self.send_json(403, {"error": "Disponível só para gerente/administrador"})
            agora = datetime.now(timezone.utc)
            out = []
            with _lock:
                usuarios = [u for u in _db["users"] if u.get("ativo", True)]
                for u in usuarios:
                    ls = _online.get(u["id"])
                    seg = None
                    if ls:
                        try:
                            seg = int((agora - datetime.fromisoformat(ls)).total_seconds())
                        except (ValueError, TypeError):
                            seg = None
                    out.append({
                        "nome": u["nome"], "papel": u["papel"],
                        "online": seg is not None and seg <= ONLINE_LIMIAR_S,
                        "segundos": seg,
                    })
            out.sort(key=lambda x: (not x["online"], x["segundos"] if x["segundos"] is not None else 10 ** 12))
            return self.send_json(200, {"usuarios": out, "limiar": ONLINE_LIMIAR_S})

        if path == "/api/atividades" and method == "GET":
            if not gestor:
                return self.send_json(403, {"error": "Disponível só para gerente/administrador"})
            try:
                limite = min(int((qs.get("limite") or ["60"])[0]), 200)
            except ValueError:
                limite = 60
            # Segura o _lock (global) o MINIMO possivel: so coleta tuplas leves das
            # ultimas 'limite' entradas de CADA lead (as unicas que poderiam entrar
            # no top-N global). A ordenacao/montagem pesada fica FORA do lock.
            parciais = []
            with _lock:
                for l in _db["leads"]:
                    nome = l.get("nome") or "(sem nome)"
                    lid = l.get("id")
                    for h in (l.get("historico") or [])[-limite:]:
                        parciais.append((h.get("data") or "", h.get("autor"), h.get("papel"),
                                         h.get("tipo"), list(h.get("itens") or []), lid, nome))
            top = heapq.nlargest(limite, parciais, key=lambda t: t[0])
            ev = [{"data": t[0], "autor": t[1], "papel": t[2], "tipo": t[3],
                   "itens": t[4], "lead_id": t[5], "lead_nome": t[6]} for t in top]
            return self.send_json(200, {"atividades": ev})

        # ---- Foto de visita (serve o arquivo; exige sessao) ----
        mf = re.match(r"^/api/foto/([A-Za-z0-9_.-]+)$", path)
        if mf and method == "GET":
            nome = mf.group(1)
            caminho = os.path.join(FOTOS_DIR, nome)
            if ".." in nome or not os.path.abspath(caminho).startswith(os.path.abspath(FOTOS_DIR)) \
                    or not os.path.isfile(caminho):
                return self.send_json(404, {"error": "Foto não encontrada"})
            with open(caminho, "rb") as f:
                data = f.read()
            ext = os.path.splitext(nome)[1].lower()
            ct = {".png": "image/png", ".webp": "image/webp"}.get(ext, "image/jpeg")
            self.send_response(200)
            self.send_header("Content-Type", ct)
            self.send_header("Content-Length", str(len(data)))
            self.send_header("Cache-Control", "private, max-age=86400")
            self.end_headers()
            self.wfile.write(data)
            return

        if path == "/api/logout" and method == "POST":
            with _lock:
                _db.get("sessions", {}).pop(self._cookie("sessao"), None)
                save_db()
            return self.send_json(200, {"ok": True}, headers={
                "Set-Cookie": "sessao=; Path=/; HttpOnly; Max-Age=0"})

        # ---- Usuarios e niveis de acesso (so admin) ----
        if path.startswith("/api/users"):
            if user["papel"] != "admin":
                return self.send_json(403, {"error": "Só o administrador gerencia usuários"})
            if path == "/api/users" and method == "GET":
                with _lock:
                    return self.send_json(200, {"users": [user_publico(u) for u in _db["users"]]})
            if path == "/api/users" and method == "POST":
                try:
                    body = self.read_body()
                except Exception:
                    return self.send_json(400, {"error": "Corpo invalido"})
                nome = str(body.get("nome") or "").strip()
                papel = body.get("papel")
                senha = str(body.get("senha") or "")
                if not nome:
                    return self.send_json(400, {"error": "Informe o nome"})
                if papel not in PAPEIS_USUARIO:
                    return self.send_json(400, {"error": "Nível de acesso inválido"})
                if senha and len(senha) < 6:
                    return self.send_json(400, {"error": "Senha muito curta (mínimo 6 caracteres)"})
                with _lock:
                    login = str(body.get("login") or "").strip().lower() or _slug_login(nome)
                    if any(u.get("login") == login for u in _db["users"]):
                        return self.send_json(400, {"error": "Já existe usuário com esse login"})
                    if nome_em_uso(nome):
                        return self.send_json(400, {"error": "Já existe usuário com esse nome — use um nome diferente"})
                    salt, h = hash_senha(senha) if senha else ("", "")
                    novo = {"id": new_id(), "nome": nome, "login": login, "salt": salt,
                            "senha_hash": h, "papel": papel, "ativo": True}
                    _db["users"].append(novo)
                    save_db()
                    return self.send_json(201, {"user": user_publico(novo)})
            mu = re.match(r"^/api/users/([^/]+)$", path)
            if mu:
                body = None
                if method in ("PATCH", "PUT"):
                    try:
                        body = self.read_body()
                    except Exception:
                        return self.send_json(400, {"error": "Corpo invalido"})
                with _lock:
                    alvo = next((u for u in _db["users"] if u["id"] == mu.group(1)), None)
                    if not alvo:
                        return self.send_json(404, {"error": "Usuário não encontrado"})
                    if method in ("PATCH", "PUT"):
                        if "nome" in body and str(body["nome"]).strip():
                            novo_nome = str(body["nome"]).strip()
                            if nome_em_uso(novo_nome, exclude_id=alvo["id"]):
                                return self.send_json(400, {"error": "Já existe usuário com esse nome"})
                            renomeia_dono_leads(alvo["nome"], novo_nome)  # leads seguem o dono
                            alvo["nome"] = novo_nome
                        if body.get("papel") in PAPEIS_USUARIO:
                            if alvo["id"] == user["id"] and body["papel"] != "admin":
                                return self.send_json(400, {"error": "Você não pode rebaixar o próprio acesso"})
                            alvo["papel"] = body["papel"]
                        if "ativo" in body:
                            if alvo["id"] == user["id"] and not body["ativo"]:
                                return self.send_json(400, {"error": "Você não pode desativar a si mesmo"})
                            alvo["ativo"] = bool(body["ativo"])
                        if "acesso_recuperacao" in body:
                            alvo["acesso_recuperacao"] = bool(body["acesso_recuperacao"])
                        if "recebe_leads" in body:
                            alvo["recebe_leads"] = bool(body["recebe_leads"])
                        if "pode_mover" in body:
                            alvo["pode_mover"] = bool(body["pode_mover"])
                        if body.get("senha"):
                            if len(str(body["senha"])) < 6:
                                return self.send_json(400, {"error": "Senha muito curta (mínimo 6 caracteres)"})
                            alvo["salt"], alvo["senha_hash"] = hash_senha(str(body["senha"]))
                            alvo.pop("senha_padrao", None)  # deixou de ser a senha padrao
                            # troca de senha derruba sessoes antigas desse usuario
                            _db["sessions"] = {t: s for t, s in _db["sessions"].items()
                                               if s.get("user_id") != alvo["id"]}
                        save_db()
                        return self.send_json(200, {"user": user_publico(alvo)})
                    if method == "DELETE":
                        if alvo["id"] == user["id"]:
                            return self.send_json(400, {"error": "Você não pode excluir a si mesmo"})
                        _db["users"] = [u for u in _db["users"] if u["id"] != alvo["id"]]
                        _db["sessions"] = {t: s for t, s in _db["sessions"].items()
                                           if s.get("user_id") != alvo["id"]}
                        save_db()
                        return self.send_json(200, {"ok": True})

        # Estatisticas (calculadas sobre os leads que ESTE usuario pode ver)
        if path == "/api/stats" and method == "GET":
            escopo = (qs.get("escopo") or ["atuais"])[0]
            with _lock:
                todos_visiveis = [l for l in _db["leads"] if pode_ver_lead(user, l)]
                # contagens dos lotes (para os botoes Atuais/Recuperacao/Servicos).
                # Recuperacao so conta/aparece para quem tem acesso liberado.
                pode_rec = pode_recuperacao(user)
                n_recuperacao = sum(1 for l in todos_visiveis
                                    if l.get("recuperacao") and l.get("tipo") != "curso") if pode_rec else 0
                n_servicos = sum(1 for l in todos_visiveis if l.get("em_servicos")) \
                    if user["papel"] != "sdr" else 0
                n_curso = sum(1 for l in todos_visiveis if l.get("em_curso")
                              and (not l.get("recuperacao") or pode_rec
                                   or user["nome"] in (l.get("vendedor"), l.get("responsavel")))) \
                    if user["papel"] != "sdr" else 0
                n_atuais = sum(1 for l in todos_visiveis if not l.get("recuperacao"))
                if escopo == "servicos":
                    visiveis = [l for l in todos_visiveis if l.get("em_servicos")] \
                        if user["papel"] != "sdr" else []
                elif escopo == "curso":
                    visiveis = [l for l in todos_visiveis if l.get("em_curso")
                                and (not l.get("recuperacao") or pode_rec
                                     or user["nome"] in (l.get("vendedor"), l.get("responsavel")))] \
                        if user["papel"] != "sdr" else []
                elif escopo == "recuperacao":
                    # recuperacao do CURSO vive so no painel 🎓 — aqui e a dos drones
                    visiveis = [l for l in todos_visiveis
                                if l.get("recuperacao") and l.get("tipo") != "curso"] if pode_rec else []
                else:
                    visiveis = [l for l in todos_visiveis if not l.get("recuperacao")]
                por_status = {s: {"count": 0, "valor": 0} for s in STAGES}
                total_valor = 0
                for l in visiveis:
                    s = l.get("status") if l.get("status") in STAGES else "novo"
                    por_status[s]["count"] += 1
                    por_status[s]["valor"] += float(l.get("valor") or 0)
                    if l.get("status") not in ("perdido", "desistiu", "curioso"):
                        total_valor += float(l.get("valor") or 0)
                produtores = sum(1 for l in visiveis if l.get("tipo") == "produtor")
                prestadores = sum(1 for l in visiveis if l.get("tipo") == "prestador")
                pecuaristas = sum(1 for l in visiveis if l.get("tipo") == "pecuarista")
                cursos = sum(1 for l in visiveis if l.get("tipo") == "curso")
                # Cada lead conta UMA vez no 🔔, na mesma prioridade dos cards:
                # respondeu > tarefa vencendo > registrar resposta > retorno
                _cad = cadencia_dias_cfg()
                respondeu = sum(1 for l in visiveis if l.get("cliente_respondeu"))
                tarefas_cobrando = sum(1 for l in visiveis if tarefa_cobrando(l)
                                       and not l.get("cliente_respondeu"))
                aguardando = sum(1 for l in visiveis if l.get("aguardando_resposta")
                                 and not tarefa_cobrando(l))
                retornos = sum(1 for l in visiveis
                               if precisa_retorno(l, _cad) and not tarefa_cobrando(l))
                # cidades presentes nos leads visiveis (para o filtro de cidade)
                cidades = sorted({str(l.get("regiao") or "").strip()
                                  for l in visiveis if str(l.get("regiao") or "").strip()})
                return self.send_json(200, {
                    "total": len(visiveis),
                    "valor_pipeline": total_valor,
                    "produtores": produtores,
                    "prestadores": prestadores,
                    "pecuaristas": pecuaristas,
                    "cursos": cursos,
                    "aguardando_resposta": aguardando,
                    "cliente_respondeu": respondeu,
                    "retornos": retornos,
                    "tarefas_cobrando": tarefas_cobrando,
                    "alertas": aguardando + retornos + respondeu + tarefas_cobrando,
                    "atuais_total": n_atuais,
                    "recuperacao_total": n_recuperacao,
                    "servicos_total": n_servicos,
                    "curso_total": n_curso,
                    "etapas": etapas_publico(),
                    "cadencia_dias": _cad,
                    "resposta_horas": resposta_horas_cfg(),
                    "chatwoot_url": str(_db.get("settings", {}).get("chatwoot_url") or ""),
                    "chatwoot_account_id": str(_db.get("settings", {}).get("chatwoot_account_id") or ""),
                    "curso_chatwoot_url": str(_db.get("settings", {}).get("curso_chatwoot_url") or ""),
                    "curso_chatwoot_account_id": str(_db.get("settings", {}).get("curso_chatwoot_account_id") or ""),
                    "cidades": cidades,
                    "mesorregioes": MESORREGIOES,
                    "por_status": por_status,
                    "stages": STAGES,
                })

        # Campanhas
        if path == "/api/campaigns" and method == "GET":
            with _lock:
                return self.send_json(200, {
                    "campaigns": list(_db.get("campaigns", [])),
                    "settings": settings_publico(),
                })

        if path == "/api/campaigns" and method == "POST":
            if not gestor:
                return self.send_json(403, {"error": "Sem permissão para gerenciar campanhas"})
            try:
                body = self.read_body()
            except Exception:
                return self.send_json(400, {"error": "Corpo invalido"})
            nome = (body.get("nome") or "").strip()
            if not nome:
                return self.send_json(400, {"error": "Informe o nome da campanha"})
            canal = body.get("canal") if body.get("canal") in CANAIS else "Meta"
            with _lock:
                codigo = (body.get("codigo") or "").strip().upper()
                codigo = re.sub(r"[^A-Z0-9]", "", codigo)
                if codigo:
                    existing = {(c.get("codigo") or "").upper() for c in _db.get("campaigns", [])}
                    if codigo in existing:
                        return self.send_json(400, {"error": "Ja existe campanha com esse codigo"})
                else:
                    codigo = gen_codigo(nome)
                camp = {
                    "id": new_id(),
                    "nome": nome,
                    "canal": canal,
                    "codigo": codigo,
                    "keyword": (body.get("keyword") or "").strip(),
                    "utm_campaign": (body.get("utm_campaign") or "").strip(),
                    "ativo": True,
                    "created_at": now_iso(),
                }
                _db.setdefault("campaigns", []).append(camp)
                save_db()
                return self.send_json(201, {"campaign": camp})

        mc = re.match(r"^/api/campaigns/([^/]+)$", path)
        if mc:
            if not gestor:
                return self.send_json(403, {"error": "Sem permissão para gerenciar campanhas"})
            camp_id = mc.group(1)
            body = None
            if method in ("PATCH", "PUT"):
                # le o corpo ANTES do lock: rfile.read e I/O de rede bloqueante
                try:
                    body = self.read_body()
                except Exception:
                    return self.send_json(400, {"error": "Corpo invalido"})
            with _lock:
                camp = next((c for c in _db.get("campaigns", []) if c["id"] == camp_id), None)
                if not camp:
                    return self.send_json(404, {"error": "Campanha nao encontrada"})
                if method in ("PATCH", "PUT"):
                    if "nome" in body and str(body["nome"]).strip():
                        camp["nome"] = str(body["nome"]).strip()
                    if body.get("canal") in CANAIS:
                        camp["canal"] = body["canal"]
                    if "keyword" in body:
                        camp["keyword"] = str(body["keyword"]).strip()
                    if "utm_campaign" in body:
                        camp["utm_campaign"] = str(body["utm_campaign"]).strip()
                    if "ativo" in body:
                        camp["ativo"] = bool(body["ativo"])
                    save_db()
                    return self.send_json(200, {"campaign": camp})
                if method == "DELETE":
                    _db["campaigns"] = [c for c in _db.get("campaigns", []) if c["id"] != camp_id]
                    # leads mantem o nome da campanha em texto; so desfaz o vinculo
                    for l in _db["leads"]:
                        if l.get("campanha_id") == camp_id:
                            l["campanha_id"] = ""
                    save_db()
                    return self.send_json(200, {"ok": True})

        # Configuracoes (numero do WhatsApp para gerar links de anuncio)
        if path == "/api/settings" and method in ("PATCH", "PUT", "POST"):
            if not gestor:
                return self.send_json(403, {"error": "Sem permissão para alterar configurações"})
            try:
                body = self.read_body()
            except Exception:
                return self.send_json(400, {"error": "Corpo invalido"})
            with _lock:
                st = _db.setdefault("settings", {})
                if "whatsapp_number" in body:
                    st["whatsapp_number"] = re.sub(r"[^0-9]", "", str(body["whatsapp_number"]))
                if "cadencia_dias" in body:
                    try:
                        st["cadencia_dias"] = max(1, min(30, int(body["cadencia_dias"])))
                    except (TypeError, ValueError):
                        return self.send_json(400, {"error": "Prazo de retorno inválido (use 1 a 30 dias)"})
                if "resposta_horas" in body:
                    try:
                        st["resposta_horas"] = max(1, min(168, int(body["resposta_horas"])))
                    except (TypeError, ValueError):
                        return self.send_json(400, {"error": "Prazo da resposta inválido (use 1 a 168 horas)"})
                # Integracao com o Chatwoot (atender no canal oficial)
                if "chatwoot_url" in body:
                    # aceita o endereço colado do navegador (com /app/...): guarda só a raiz
                    st["chatwoot_url"] = chatwoot_base_url(body["chatwoot_url"])
                if "chatwoot_account_id" in body:
                    st["chatwoot_account_id"] = re.sub(r"[^0-9]", "", str(body["chatwoot_account_id"]))
                if "chatwoot_inbox_id" in body:
                    st["chatwoot_inbox_id"] = re.sub(r"[^0-9]", "", str(body["chatwoot_inbox_id"]))
                if "chatwoot_token" in body:
                    st["chatwoot_token"] = str(body["chatwoot_token"] or "").strip()
                if "chatwoot_saudacao" in body:
                    st["chatwoot_saudacao"] = str(body["chatwoot_saudacao"] or "").strip()[:2000]
                # Chatwoot separado do CURSO (mesmos campos, prefixo proprio)
                if "curso_chatwoot_url" in body:
                    st["curso_chatwoot_url"] = chatwoot_base_url(body["curso_chatwoot_url"])
                if "curso_chatwoot_account_id" in body:
                    st["curso_chatwoot_account_id"] = re.sub(r"[^0-9]", "", str(body["curso_chatwoot_account_id"]))
                if "curso_chatwoot_inbox_id" in body:
                    st["curso_chatwoot_inbox_id"] = re.sub(r"[^0-9]", "", str(body["curso_chatwoot_inbox_id"]))
                if "curso_chatwoot_token" in body:
                    st["curso_chatwoot_token"] = str(body["curso_chatwoot_token"] or "").strip()
                if "curso_chatwoot_saudacao" in body:
                    st["curso_chatwoot_saudacao"] = str(body["curso_chatwoot_saudacao"] or "").strip()[:2000]
                save_db()
                return self.send_json(200, {"settings": settings_publico()})

        # ---- Regras de potencial comercial (a cada X ha de tal cultura = 1 venda) ----
        if path == "/api/potencial" and method == "GET":
            return self.send_json(200, {"regras": potencial_regras()})

        if path == "/api/potencial" and method == "POST":
            if not gestor:
                return self.send_json(403, {"error": "Só gerente/administrador define o potencial"})
            try:
                body = self.read_body()
            except Exception:
                return self.send_json(400, {"error": "Corpo invalido"})
            entradas = body.get("regras")
            if not isinstance(entradas, list) or len(entradas) > 40:
                return self.send_json(400, {"error": "Envie até 40 regras"})
            limpas = []
            for r in entradas:
                if not isinstance(r, dict):
                    continue
                try:
                    ha = float(r.get("ha_por_unidade") or 0)
                    valor = float(r.get("valor_unidade") or 0)
                except (TypeError, ValueError):
                    return self.send_json(400, {"error": "Área e valor precisam ser números"})
                cid = str(r.get("cultura_id") or "")
                if not cid.isdigit() or ha <= 0:
                    return self.send_json(400, {"error": "Escolha a cultura e uma área maior que zero"})
                if valor < 0 or valor > 1e12 or ha > 1e7:
                    return self.send_json(400, {"error": "Valores fora do razoável"})
                limpas.append({"id": str(r.get("id") or "").strip()[:40] or ("p" + secrets.token_hex(4)),
                               "cultura_id": int(cid),
                               "cultura": str(r.get("cultura") or "").strip()[:60],
                               "ha_por_unidade": ha,
                               "produto": str(r.get("produto") or "").strip()[:60] or "Produto",
                               "valor_unidade": valor,
                               "ativa": r.get("ativa") is not False})
            with _lock:
                _db.setdefault("settings", {})["potencial_regras"] = limpas
                save_db()
            return self.send_json(200, {"regras": potencial_regras()})

        # ---- Etapas do funil: renomear qualquer coluna, criar/excluir (gestor) ----
        if path == "/api/etapas" and method == "POST":
            if not gestor:
                return self.send_json(403, {"error": "Só gerente/administrador edita as etapas"})
            try:
                body = self.read_body()
            except Exception:
                return self.send_json(400, {"error": "Corpo invalido"})
            funil = str(body.get("funil") or "")
            acao = str(body.get("acao") or "")
            key = str(body.get("key") or "").strip()
            label = str(body.get("label") or "").strip()[:40]
            if funil not in ("vendas", "servicos", "curso") or acao not in ("criar", "renomear", "excluir"):
                return self.send_json(400, {"error": "Pedido inválido"})
            campo_lead = {"vendas": "status", "servicos": "status_servico", "curso": "status_curso"}[funil]
            lista_atual = {"vendas": STAGES, "servicos": SERVICO_STAGES, "curso": CURSO_STAGES}[funil]
            with _lock:
                st = _db.setdefault("settings", {})
                custom = st.setdefault("etapas_custom", {})
                minhas = custom.setdefault(funil, [])
                rem = st.setdefault("etapas_removidas", [])
                rot = st.setdefault("rotulos", {})
                if acao == "criar":
                    if not label:
                        return self.send_json(400, {"error": "Dê um nome para a nova etapa"})
                    if len(minhas) >= MAX_ETAPAS_CUSTOM:
                        return self.send_json(400, {"error": "Limite de %d etapas criadas neste funil" % MAX_ETAPAS_CUSTOM})
                    minhas.append({"key": "x" + secrets.token_hex(4), "label": label})
                elif acao == "renomear":
                    if not label:
                        return self.send_json(400, {"error": "Dê um nome para a etapa"})
                    if key not in lista_atual:
                        return self.send_json(400, {"error": "Etapa não encontrada"})
                    proprio = next((e for e in minhas if e.get("key") == key), None)
                    if proprio:
                        proprio["label"] = label
                    else:
                        rot[key] = label
                else:  # excluir
                    if key not in lista_atual:
                        return self.send_json(400, {"error": "Etapa não encontrada"})
                    removiveis = set(vendas_meio()) if funil == "vendas" else set(lista_atual)
                    if key not in removiveis:
                        return self.send_json(400, {"error": "Esta etapa é fixa do sistema — dá para renomear, mas não excluir"})
                    if funil != "vendas" and len(lista_atual) <= 1:
                        return self.send_json(400, {"error": "O funil precisa de ao menos uma etapa"})
                    n_leads = sum(1 for l in _db["leads"] if l.get(campo_lead) == key)
                    if n_leads:
                        return self.send_json(400, {"error":
                            "Há %d lead(s) nessa etapa — mova todos para outra coluna antes de excluir" % n_leads})
                    if any(e.get("key") == key for e in minhas):
                        custom[funil] = [e for e in minhas if e.get("key") != key]
                    else:
                        rem.append(key)
                    rot.pop(key, None)
                save_db()
                recalcula_etapas()
                return self.send_json(200, {"ok": True, "etapas": etapas_publico()})

        # Testa a conexao com o Chatwoot e devolve um veredito em portugues claro
        # (token errado? conta errada? servidor fora?). So gestor. NUNCA inclui o
        # token nas mensagens.
        # Endereço do webhook (a "ponte" Chatwoot -> CRM) + status de recebimento.
        # Só gestor vê — a URL carrega o token secreto do webhook.
        if path == "/api/webhook-info" and method == "GET":
            if not gestor:
                return self.send_json(403, {"error": "Disponível só para gerente/administrador"})
            host = self.headers.get("Host") or ("localhost:%d" % PORT)
            proto = self.headers.get("X-Forwarded-Proto") or (
                "http" if host.split(":")[0] in ("localhost", "127.0.0.1") else "https")
            url = "%s://%s/webhook/chatwoot?token=%s" % (proto, host, WEBHOOK_TOKEN or ensure_webhook_token())
            with _lock:
                cw = [l for l in _db["leads"] if l.get("source") == "chatwoot"]
                ultimo_lead = max((str(l.get("created_at") or "") for l in cw), default="") or None
            return self.send_json(200, {"url": url, "url_curso": url + "&origem=curso",
                                        "total_chatwoot": len(cw),
                                        "ultimo_lead": ultimo_lead,
                                        "ultimo_evento": _webhook_ultimo})

        # Conectar TODOS os leads da recuperacao (lote, em segundo plano).
        # POST inicia; GET acompanha o progresso. So vincula — nao manda mensagem.
        # Importar leads ANTIGOS do Chatwoot do CURSO (lote de recuperacao).
        # POST inicia; GET acompanha. So cria/vincula — NAO envia mensagem.
        if path == "/api/chatwoot/importar-curso" and method == "GET":
            if not gestor:
                return self.send_json(403, {"error": "Disponível só para gerente/administrador"})
            return self.send_json(200, dict(_importar_curso))

        if path == "/api/chatwoot/importar-curso" and method == "POST":
            if not gestor:
                return self.send_json(403, {"error": "Disponível só para gerente/administrador"})
            base, acc, token, _s = chatwoot_cfg("curso")
            if not (base and acc and token):
                return self.send_json(400, {"error": "Configure o Chatwoot do curso acima (endereço, conta e token) antes"})
            with _lock:
                if _importar_curso["rodando"]:
                    return self.send_json(200, dict(_importar_curso, iniciado=False, ja_rodando=True))
                _importar_curso.update({"rodando": True, "paginas": 0, "vistos": 0,
                                        "criados": 0, "ja_no_crm": 0, "sem_contato": 0,
                                        "falhas": 0, "ultimo_erro": None, "terminado_em": None})
            threading.Thread(target=_roda_importar_curso, args=(base, acc, token), daemon=True).start()
            return self.send_json(200, dict(_importar_curso, iniciado=True))

        if path == "/api/chatwoot/conectar-todos" and method == "GET":
            if not gestor:
                return self.send_json(403, {"error": "Disponível só para gerente/administrador"})
            return self.send_json(200, dict(_conectar_lote))
        if path == "/api/chatwoot/conectar-todos" and method == "POST":
            if not gestor:
                return self.send_json(403, {"error": "Disponível só para gerente/administrador"})
            base, acc, token, _s = chatwoot_cfg()
            if not (base and acc and token):
                return self.send_json(400, {"error": "Configure o Chatwoot acima (endereço, conta e token) antes"})
            inbox = str(_db.get("settings", {}).get("chatwoot_inbox_id") or "").strip()
            with _lock:
                if _conectar_lote["rodando"]:
                    return self.send_json(200, dict(_conectar_lote, iniciado=False, ja_rodando=True))
                _conectar_lote.update({"rodando": True, "total": 0, "feitos": 0, "conectados": 0,
                                       "ja_tinham": 0, "falhas": 0, "ultimo_erro": None, "terminado_em": None})
            threading.Thread(target=_roda_conectar_lote, args=(base, acc, token, inbox), daemon=True).start()
            return self.send_json(200, {"iniciado": True})

        if path == "/api/chatwoot/teste" and method == "POST":
            if not gestor:
                return self.send_json(403, {"error": "Teste disponível só para gerente/administrador"})
            try:
                corpo_teste = self.read_body()
            except Exception:
                corpo_teste = {}
            origem_teste = "curso" if corpo_teste.get("origem") == "curso" else ""
            base, acc, token, saud = chatwoot_cfg(origem_teste)
            faltas = [n for n, v in (("o endereço", base), ("o nº da conta", acc), ("o token", token)) if not v]
            if faltas:
                return self.send_json(200, {"ok": False, "mensagem":
                    "Falta preencher %s — preencha, salve e teste de novo." % " e ".join(faltas)})

            def _get(url):
                req = urllib.request.Request(url, headers=chatwoot_headers(token))
                with _ABRIDOR_SEM_REDIRECT.open(req, timeout=10):
                    pass

            try:
                _get("%s/api/v1/profile" % base)
            except urllib.error.HTTPError as e:
                if e.code in (401, 403):
                    return self.send_json(200, {"ok": False, "mensagem":
                        "❌ O Chatwoot recusou o token (HTTP %d). Copie de novo em: "
                        "Chatwoot → sua foto (canto inferior esquerdo) → Configurações de perfil → "
                        "Token de acesso — e cole aqui por cima." % e.code})
                return self.send_json(200, {"ok": False, "mensagem":
                    "❌ O endereço %s respondeu HTTP %d — parece não ser a raiz do Chatwoot. "
                    "Use só o começo do endereço (ex.: https://chat.novaeradrones.com.br)." % (base, e.code)})
            except Exception as e:
                return self.send_json(200, {"ok": False, "mensagem":
                    "❌ Não consegui falar com %s (%s) — confira o endereço e se o servidor "
                    "do CRM tem acesso à internet." % (base, type(e).__name__)})
            try:
                _get("%s/api/v1/accounts/%s/labels" % (base, acc))
            except urllib.error.HTTPError as e:
                return self.send_json(200, {"ok": False, "mensagem":
                    "⚠️ O token é válido, mas a conta Nº %s foi recusada (HTTP %d). Confira o "
                    "número: ele aparece no endereço do Chatwoot, em /app/accounts/NÚMERO/." % (acc, e.code)})
            except Exception as e:
                return self.send_json(200, {"ok": False, "mensagem":
                    "❌ Não consegui falar com o Chatwoot (%s) — tente de novo." % type(e).__name__})
            # lista as caixas de entrada (o gestor precisa do Nº p/ o CRM criar
            # conversas novas — leads da recuperação sem conversa antiga)
            caixas_txt = ""
            try:
                res = _cw_req(base, "/api/v1/accounts/%s/inboxes" % acc, token)
                caixas = _cw_payload(res)
                if isinstance(caixas, list) and caixas:
                    nomes = ["Nº %s = %s" % (c.get("id"), c.get("name") or "?")
                             for c in caixas if isinstance(c, dict)]
                    caixas_txt = " Caixas de entrada: " + " · ".join(nomes) + "."
            except Exception:
                pass
            inbox_cfg = str(_db.get("settings", {}).get("chatwoot_inbox_id") or "").strip()
            aviso_inbox = "" if inbox_cfg else (" ⚠️ Preencha o Nº da caixa de entrada para o CRM "
                                                "poder CRIAR conversa para cliente que ainda não tem.")
            if not saud:
                return self.send_json(200, {"ok": True, "mensagem":
                    ("✅ Conexão OK (token e conta Nº %s válidos) — mas a saudação está vazia. "
                     "Escreva a mensagem e salve.%s%s") % (acc, caixas_txt, aviso_inbox)})
            return self.send_json(200, {"ok": True, "mensagem":
                ("✅ Tudo certo! Token válido e conta Nº %s acessível — o botão "
                 "\"Atender no Chatwoot\" vai enviar a saudação automática.%s%s") % (acc, caixas_txt, aviso_inbox)})

        # Relatorio por campanha (quantos leads/produtores/ganhos e R$ cada uma gerou)
        if path == "/api/report/campanhas" and method == "GET":
            if not gestor:
                return self.send_json(403, {"error": "Relatório disponível só para gerente/administrador"})
            with _lock:
                camps = list(_db.get("campaigns", []))
                rows = {c["id"]: {
                    "id": c["id"], "nome": c["nome"], "canal": c.get("canal", ""),
                    "codigo": c.get("codigo", ""), "ativo": c.get("ativo", True),
                    "leads": 0, "produtores": 0, "ganhos": 0,
                    "valor_ganho": 0.0, "valor_aberto": 0.0,
                } for c in camps}
                sem = {"id": "", "nome": "Sem campanha identificada", "canal": "",
                       "codigo": "", "ativo": True, "leads": 0, "produtores": 0,
                       "ganhos": 0, "valor_ganho": 0.0, "valor_aberto": 0.0}
                for l in _db["leads"]:
                    if l.get("recuperacao"):
                        continue  # relatorio = pipeline atual, sem o lote de recuperacao
                    row = rows.get(l.get("campanha_id") or "", sem)
                    row["leads"] += 1
                    if l.get("tipo") == "produtor":
                        row["produtores"] += 1
                    valor = float(l.get("valor") or 0)
                    if l.get("status") == "ganho":
                        row["ganhos"] += 1
                        row["valor_ganho"] += valor
                    elif l.get("status") not in ("perdido", "desistiu", "curioso"):
                        row["valor_aberto"] += valor
                out = sorted(rows.values(), key=lambda r: r["leads"], reverse=True)
                if sem["leads"]:
                    out.append(sem)
                return self.send_json(200, {"report": out})

        # Relatorio diario: quantos leads chegaram e quantos foram qualificados
        if path == "/api/report/diario" and method == "GET":
            if not gestor:
                return self.send_json(403, {"error": "Relatório disponível só para gerente/administrador"})
            try:
                dias = min(int((qs.get("dias") or ["30"])[0]), 365)
            except ValueError:
                dias = 30
            # agrupar = dia (padrao) | semana (a linha e a segunda-feira) | mes
            agrupar = (qs.get("agrupar") or ["dia"])[0]
            if agrupar not in ("dia", "semana", "mes"):
                agrupar = "dia"

            def _data_ok(s):
                try:
                    datetime.strptime(s, "%Y-%m-%d")
                    return True
                except (ValueError, TypeError):
                    return False
            # Janela por DATA (ultimos N dias de verdade), nao por numero de
            # linhas. Alternativas: dias<=0 = TEMPO TOTAL (desde o inicio);
            # de/ate = PERIODO ESPECIFICO escolhido pelo gestor (datas BRT,
            # inclusivas; pode vir so uma das pontas).
            de = (qs.get("de") or [""])[0]
            ate = (qs.get("ate") or [""])[0]
            teto = "9999-12-31"
            if _data_ok(de) or _data_ok(ate):
                corte = de if _data_ok(de) else "0000-01-01"
                if _data_ok(ate):
                    teto = ate
            elif dias <= 0:
                corte = "0000-01-01"
            else:
                corte = dia_brt((datetime.now(timezone.utc) - timedelta(days=dias)).isoformat())

            def chave(d):  # d = "AAAA-MM-DD" ja em horario de Brasilia
                if agrupar == "mes":
                    return d[:7]  # "AAAA-MM"
                if agrupar == "semana":
                    dt = datetime.strptime(d, "%Y-%m-%d")
                    return (dt - timedelta(days=dt.weekday())).strftime("%Y-%m-%d")
                return d
            with _lock:
                por_dia = {}

                def bucket(d):
                    return por_dia.setdefault(chave(d), {
                        "dia": chave(d), "recebidos": 0, "recebidos_chatwoot": 0,
                        "qualificados": 0, "produtores": 0, "prestadores": 0, "pecuaristas": 0,
                        "cursos": 0, "ganhos": 0, "perdidos": 0, "desistidos": 0})

                for l in _db["leads"]:
                    if l.get("recuperacao"):
                        continue  # relatorio = leads NOVOS, sem o lote de recuperacao
                    d = dia_brt(l.get("created_at"))
                    if d and corte <= d <= teto:
                        b = bucket(d)
                        b["recebidos"] += 1
                        if l.get("source") == "chatwoot":
                            b["recebidos_chatwoot"] += 1
                    dq = dia_brt(l.get("qualificado_em"))
                    if dq and corte <= dq <= teto:
                        b = bucket(dq)
                        b["qualificados"] += 1
                        if l.get("tipo") == "produtor":
                            b["produtores"] += 1
                        elif l.get("tipo") == "prestador":
                            b["prestadores"] += 1
                        elif l.get("tipo") == "pecuarista":
                            b["pecuaristas"] += 1
                        elif l.get("tipo") == "curso":
                            b["cursos"] += 1
                    if l.get("status") == "ganho":
                        dg = dia_brt(l.get("ganho_em") or l.get("updated_at"))
                        if dg and corte <= dg <= teto:
                            bucket(dg)["ganhos"] += 1
                    elif l.get("status") == "perdido":
                        dp = dia_brt(l.get("perdido_em") or l.get("updated_at"))
                        if dp and corte <= dp <= teto:
                            bucket(dp)["perdidos"] += 1
                    elif l.get("status") == "desistiu":
                        dd = dia_brt(l.get("desistiu_em") or l.get("updated_at"))
                        if dd and corte <= dd <= teto:
                            bucket(dd)["desistidos"] += 1

                linhas = sorted(por_dia.values(), key=lambda r: r["dia"], reverse=True)
                totais = {"recebidos": sum(r["recebidos"] for r in linhas),
                          "recebidos_chatwoot": sum(r["recebidos_chatwoot"] for r in linhas),
                          "qualificados": sum(r["qualificados"] for r in linhas),
                          "cursos": sum(r["cursos"] for r in linhas),
                          "ganhos": sum(r["ganhos"] for r in linhas)}
                return self.send_json(200, {"report": linhas, "totais": totais, "agrupar": agrupar})

        # Equipe (visao dos usuarios com papel de raia: SDRs e vendedores).
        # Leitura para todos (nomes das raias); criacao/edicao so gestores —
        # a gestao completa (login/senha/nivel) fica em /api/users (admin).
        if path == "/api/members" and method == "GET":
            with _lock:
                equipe = [user_publico(u) for u in _db["users"] if u.get("papel") in PAPEIS]
                return self.send_json(200, {"members": equipe})

        if path == "/api/members" and method == "POST":
            if not gestor:
                return self.send_json(403, {"error": "Sem permissão para alterar a equipe"})
            try:
                body = self.read_body()
            except Exception:
                return self.send_json(400, {"error": "Corpo invalido"})
            nome = (body.get("nome") or "").strip()
            papel = body.get("papel")
            if not nome:
                return self.send_json(400, {"error": "Informe o nome"})
            if papel not in PAPEIS:
                return self.send_json(400, {"error": "Papel invalido"})
            with _lock:
                if nome_em_uso(nome):
                    return self.send_json(400, {"error": "Já existe alguém com esse nome — use um nome diferente"})
                login = _slug_login(nome)
                if any(u.get("login") == login for u in _db["users"]):
                    login = login + secrets.token_hex(2)
                novo = {"id": new_id(), "nome": nome, "login": login, "salt": "",
                        "senha_hash": "", "papel": papel, "ativo": True}
                _db["users"].append(novo)
                save_db()
                return self.send_json(201, {"member": user_publico(novo)})

        mm = re.match(r"^/api/members/([^/]+)$", path)
        if mm:
            if not gestor:
                return self.send_json(403, {"error": "Sem permissão para alterar a equipe"})
            member_id = mm.group(1)
            body = None
            if method in ("PATCH", "PUT"):
                try:
                    body = self.read_body()
                except Exception:
                    return self.send_json(400, {"error": "Corpo invalido"})
            with _lock:
                member = next((x for x in _db["users"] if x["id"] == member_id and x.get("papel") in PAPEIS), None)
                if not member:
                    return self.send_json(404, {"error": "Membro nao encontrado"})
                if method in ("PATCH", "PUT"):
                    if "nome" in body and str(body["nome"]).strip():
                        member["nome"] = str(body["nome"]).strip()
                    if body.get("papel") in PAPEIS:
                        member["papel"] = body["papel"]
                    if "ativo" in body:
                        member["ativo"] = bool(body["ativo"])
                    save_db()
                    return self.send_json(200, {"member": user_publico(member)})
                if method == "DELETE":
                    _db["users"] = [x for x in _db["users"] if x["id"] != member_id]
                    _db["sessions"] = {t: s for t, s in _db["sessions"].items()
                                       if s.get("user_id") != member_id}
                    save_db()
                    return self.send_json(200, {"ok": True})

        # Listar (cada papel enxerga so o que lhe cabe). Todos os filtros se
        # combinam (E logico): busca + canal + pagamento + produto + cidade + hectare.
        if path == "/api/leads" and method == "GET":
            q = (qs.get("q") or [""])[0].lower().strip()
            canal = (qs.get("canal") or [""])[0]
            pagamento = (qs.get("pagamento") or [""])[0]
            produto = (qs.get("produto") or [""])[0]
            cidade = (qs.get("cidade") or [""])[0]
            mesorregiao = (qs.get("mesorregiao") or [""])[0]
            f_vendedor = (qs.get("vendedor") or [""])[0]
            f_sdr = (qs.get("sdr") or [""])[0]
            # escopo: "atuais" (padrao, funil dos leads NOVOS) x "recuperacao"
            escopo = (qs.get("escopo") or ["atuais"])[0]

            def _f(k):
                try:
                    return float((qs.get(k) or [""])[0])
                except ValueError:
                    return None
            ha_min, ha_max = _f("ha_min"), _f("ha_max")

            with _lock:
                leads = [l for l in _db["leads"] if pode_ver_lead(user, l)]
            # separa os lotes: recuperacao (antigos), servicos (pos-venda, em
            # paralelo) e atuais (funil de drones dos leads novos).
            if escopo == "servicos":
                # painel de Servicos e trabalho de vendedor — SDR nao acessa
                leads = [l for l in leads if l.get("em_servicos")] if user["papel"] != "sdr" else []
            elif escopo == "curso":
                # painel do Curso idem — venda e trabalho de vendedor. Leads de
                # RECUPERACAO do curso respeitam a liberacao por pessoa — mas o
                # DONO do lead sempre ve o proprio lead (senao o gestor atribui
                # e o responsavel fica cego para ele).
                pode_rec_c = pode_recuperacao(user)
                leads = [l for l in leads if l.get("em_curso")
                         and (not l.get("recuperacao") or pode_rec_c
                              or user["nome"] in (l.get("vendedor"), l.get("responsavel")))] \
                    if user["papel"] != "sdr" else []
            elif escopo == "recuperacao":
                # só quem tem acesso liberado vê a Recuperação. Leads 🔄 do CURSO
                # nao entram aqui — ficam SO no painel 🎓 (senao bagunca a
                # recuperacao dos drones)
                leads = [l for l in leads
                         if l.get("recuperacao") and l.get("tipo") != "curso"] \
                    if pode_recuperacao(user) else []
            else:
                leads = [l for l in leads if not l.get("recuperacao")]
            if q:
                def match(l):
                    blob = " ".join(str(l.get(k) or "") for k in
                                    ("nome", "telefone", "email", "regiao", "produto", "campanha", "vendedor", "cargo", "decisor")).lower()
                    return q in blob
                leads = [l for l in leads if match(l)]
            if canal:
                leads = [l for l in leads if l.get("origem_canal") == canal]
            if pagamento:
                leads = [l for l in leads
                         if any(fp.get("tipo") == pagamento for fp in (l.get("formas_pagamento") or []))]
            if produto:
                # casa se QUALQUER drone do pedido for esse modelo (pedido com vários)
                leads = [l for l in leads
                         if any(it.get("produto") == produto for it in (l.get("itens") or []))
                         or l.get("produto") == produto]
            if cidade:
                leads = [l for l in leads if str(l.get("regiao") or "").strip() == cidade]
            if mesorregiao:
                leads = [l for l in leads if meso_da_regiao(l.get("regiao")) == mesorregiao]
            if f_vendedor:
                if f_vendedor == "__none__":
                    leads = [l for l in leads if not str(l.get("vendedor") or "").strip()]
                else:
                    leads = [l for l in leads if l.get("vendedor") == f_vendedor]
            if f_sdr:
                if f_sdr == "__none__":
                    leads = [l for l in leads if not str(l.get("sdr") or "").strip()]
                else:
                    leads = [l for l in leads if l.get("sdr") == f_sdr]
            if ha_min is not None or ha_max is not None:
                def na_faixa(l):
                    h = parse_hectares(l.get("area_cultivada"))
                    if h is None:
                        return False  # sem area informada nao entra numa faixa
                    if ha_min is not None and h < ha_min:
                        return False
                    if ha_max is not None and h > ha_max:
                        return False
                    return True
                leads = [l for l in leads if na_faixa(l)]
            leads.sort(key=lambda l: l.get("updated_at") or "", reverse=True)
            return self.send_json(200, {"leads": leads, "stages": STAGES})

        # ---- Acao em massa: atribuir vendedor / classificar varios leads ----
        # (ANTES da rota por id, senao 'bulk' casaria no regex)
        if path == "/api/leads/bulk" and method == "POST":
            if not gestor:
                return self.send_json(403, {"error": "Só gerente/administrador pode alterar vários leads de uma vez"})
            try:
                body = self.read_body()
            except Exception:
                return self.send_json(400, {"error": "Corpo invalido"})
            ids = body.get("ids")
            if not isinstance(ids, list) or not ids:
                return self.send_json(400, {"error": "Nenhum lead selecionado"})
            if len(ids) > 2000:
                return self.send_json(400, {"error": "Seleção grande demais (máximo 2000)"})
            updates = {}
            vend = str(body.get("vendedor") or "").strip()
            if vend:
                updates["vendedor"] = vend
            if body.get("tipo") in ("produtor", "prestador", "pecuarista", "curso"):
                updates["tipo"] = body["tipo"]
            if body.get("qualificar"):
                updates["status"] = "qualificado"
            if not updates:
                return self.send_json(400, {"error": "Escolha o que alterar (vendedor e/ou classificação)"})
            alvo = set(str(i) for i in ids)
            ok, sem_mudanca, fechados, falhas = 0, 0, 0, []
            with _lock:
                for lead in _db["leads"]:
                    if lead["id"] not in alvo or not pode_ver_lead(user, lead):
                        continue
                    # Caso ja encerrado/arquivado (ganho/perdido/desistiu/curioso)
                    # NAO e tocado pela acao em massa.
                    if lead.get("status") in ("ganho", "perdido", "desistiu", "curioso"):
                        fechados += 1
                        continue
                    u = dict(updates)
                    # "qualificar" so vale para quem ainda esta na triagem: nunca
                    # puxa de volta quem ja avancou (negociacao/proposta/...)
                    if lead.get("status") not in ("novo", "triagem"):
                        u.pop("status", None)
                    if not u:
                        sem_mudanca += 1
                        continue
                    tentativa = dict(lead)
                    try:
                        apply_updates(tentativa, u)
                    except ValueError as e:
                        falhas.append({"nome": lead.get("nome") or lead["id"], "motivo": str(e)})
                        continue
                    itens = descreve_mudancas(lead, tentativa, list(u.keys()))
                    if not itens:
                        # nada mudou de verdade: nao carimba updated_at nem gera
                        # entrada no historico (senao "queima" o selo de atividade)
                        sem_mudanca += 1
                        continue
                    lead.update(tentativa)
                    registra_hist(lead, user["nome"], itens, papel=user["papel"], tipo="massa")
                    ok += 1
                if ok:
                    save_db()
            print("[massa] %d atualizados, %d sem mudanca, %d fechados ignorados (por %s)"
                  % (ok, sem_mudanca, fechados, user["nome"]))
            return self.send_json(200, {"atualizados": ok, "sem_alteracao": sem_mudanca,
                                        "fechados_ignorados": fechados, "falhas": falhas})

        # Importacao em massa (ANTES da rota por id: 'import' casaria no regex)
        if path == "/api/leads/import" and method == "POST":
            if not gestor:
                return self.send_json(403, {"error": "Importação disponível só para gerente/administrador"})
            try:
                body = self.read_body()
            except Exception:
                return self.send_json(400, {"error": "Corpo invalido"})
            csv_texto = body.get("csv")
            if not isinstance(csv_texto, str) or not csv_texto.strip():
                return self.send_json(400, {"error": "Envie o conteúdo do arquivo CSV"})
            try:
                criados, rejeitados = importar_csv(csv_texto)
            except ValueError as e:
                return self.send_json(400, {"error": str(e)})
            print("[importacao] %d lead(s) criados, %d rejeitados" % (criados, len(rejeitados)))
            return self.send_json(200, {"criados": criados, "rejeitados": rejeitados})

        # ---- Visitas de campo de um lead ----
        mv = re.match(r"^/api/leads/([^/]+)/visitas(?:/([^/]+))?$", path)
        if mv:
            lead_id, visita_id = mv.group(1), mv.group(2)
            if user["papel"] == "sdr" and method == "POST":
                return self.send_json(403, {"error": "Registrar visita é trabalho do vendedor"})
            if method == "POST" and not visita_id:
                try:
                    body = self.read_body()
                except Exception:
                    return self.send_json(400, {"error": "Corpo invalido"})
                resultado = str(body.get("resultado") or "").strip()
                if resultado and resultado not in RESULTADOS_VISITA:
                    return self.send_json(400, {"error": "Resultado de visita inválido"})
                # GPS obrigatorio: toda visita tem que registrar a localizacao
                try:
                    la, lo = float(body.get("lat")), float(body.get("lng"))
                except (TypeError, ValueError):
                    return self.send_json(400, {"error": "É obrigatório registrar a localização (GPS) da visita — permita o acesso à localização"})
                if not (math.isfinite(la) and math.isfinite(lo) and abs(la) <= 90 and abs(lo) <= 180) \
                        or (abs(la) < 0.0001 and abs(lo) < 0.0001):  # 0,0 = leitura invalida
                    return self.send_json(400, {"error": "Localização (GPS) inválida"})
                la, lo = round(la, 6), round(lo, 6)
                try:
                    acc = float(body.get("acc"))  # precisao em metros (se enviada)
                except (TypeError, ValueError):
                    acc = None
                op_id = str(body.get("op_id") or "").strip()[:64]
                vid = new_id()
                with _lock:
                    lead = next((l for l in _db["leads"] if l["id"] == lead_id), None)
                    if not lead or not pode_ver_lead(user, lead):
                        return self.send_json(404, {"error": "Lead nao encontrado"})
                    # idempotencia: visita registrada offline pode ser reenviada — se
                    # o op_id ja existe, devolve a visita existente (nao duplica foto).
                    if op_id:
                        ja = next((v for v in lead.get("visitas", []) if v.get("op_id") == op_id), None)
                        if ja:
                            return self.send_json(201, {"visita": ja, "total": len(lead["visitas"])})
                    # foto so e gravada em disco APOS confirmar o lead (evita arquivo orfao)
                    try:
                        foto = salva_foto_visita(body.get("foto"), vid)
                    except ValueError as e:
                        return self.send_json(400, {"error": str(e)})
                    quando = ts_offline(body.get("criado_em"))  # hora da visita (offline) ou agora
                    visita = {
                        "id": vid,
                        "data": quando,
                        "visitante": user["nome"],
                        "resultado": resultado,
                        "obs": str(body.get("obs") or "").strip()[:2000],
                        "foto": foto,
                        "lat": la, "lng": lo,
                    }
                    if op_id:
                        visita["op_id"] = op_id
                    lead.setdefault("visitas", []).append(visita)
                    # So atualiza a localizacao da fazenda se ela ainda nao foi
                    # ajustada, ou se a leitura for precisa (<=150 m). Assim uma
                    # posicao ruim (Wi-Fi/desktop) nao apaga um pino ja acertado.
                    sem_local = lead.get("lat") is None or lead.get("lng") is None
                    if sem_local or (acc is not None and acc <= 150):
                        lead["lat"], lead["lng"] = la, lo
                    registra_hist(lead, user["nome"], ["🚗 Visita registrada" + (": " + resultado if resultado else "")], papel=user["papel"], op_id=op_id or None, quando=quando)
                    lead["updated_at"] = now_iso()
                    save_db()
                    return self.send_json(201, {"visita": visita, "total": len(lead["visitas"])})

            if method == "DELETE" and visita_id:
                with _lock:
                    lead = next((l for l in _db["leads"] if l["id"] == lead_id), None)
                    if not lead or not pode_ver_lead(user, lead):
                        return self.send_json(404, {"error": "Lead nao encontrado"})
                    v = next((x for x in lead.get("visitas", []) if x["id"] == visita_id), None)
                    if not v:
                        return self.send_json(404, {"error": "Visita nao encontrada"})
                    if not gestor and v.get("visitante") != user["nome"]:
                        return self.send_json(403, {"error": "Só quem registrou (ou um gestor) pode excluir a visita"})
                    if v.get("foto"):
                        try:
                            os.remove(os.path.join(FOTOS_DIR, v["foto"]))
                        except OSError:
                            pass
                    lead["visitas"] = [x for x in lead["visitas"] if x["id"] != visita_id]
                    save_db()
                    return self.send_json(200, {"ok": True, "total": len(lead["visitas"])})

        # ---- Nota manual na linha do tempo (vendedor/gerente escrevem updates) ----
        # ---- Tarefas do cliente (criar / concluir-reabrir / excluir) ----
        mtar = re.match(r"^/api/leads/([^/]+)/tarefas(?:/([^/]+))?$", path)
        if mtar and method in ("POST", "PATCH", "DELETE"):
            lead_id, tarefa_id = mtar.group(1), mtar.group(2)
            try:
                body = self.read_body()
            except Exception:
                return self.send_json(400, {"error": "Corpo invalido"})
            with _lock:
                lead = next((l for l in _db["leads"] if l["id"] == lead_id), None)
                if not lead or not pode_ver_lead(user, lead):
                    return self.send_json(404, {"error": "Lead nao encontrado"})
                tarefas = lead.setdefault("tarefas", [])
                if method == "POST" and not tarefa_id:
                    texto = str(body.get("texto") or "").strip()[:300]
                    if not texto:
                        return self.send_json(400, {"error": "Escreva a tarefa antes de salvar"})
                    prazo = str(body.get("prazo") or "").strip()[:16]
                    if prazo and not re.match(r"^\d{4}-\d{2}-\d{2}", prazo):
                        return self.send_json(400, {"error": "Prazo inválido"})
                    if len(tarefas) >= 100:
                        return self.send_json(400, {"error": "Este cliente já tem 100 tarefas — conclua ou exclua antigas"})
                    t = {"id": secrets.token_hex(6), "texto": texto, "prazo": prazo,
                         "criada_por": user["nome"], "criada_em": now_iso(),
                         "feita": False, "feita_em": None, "feita_por": ""}
                    tarefas.append(t)
                    registra_hist(lead, user["nome"],
                                  ['📋 Tarefa criada: "%s"%s' % (texto, (" (até %s)" % prazo[:10]) if prazo else "")],
                                  papel=user["papel"])
                    lead["updated_at"] = now_iso()
                    save_db()
                    return self.send_json(201, {"tarefa": t, "tarefas": tarefas})
                t = next((x for x in tarefas if x.get("id") == tarefa_id), None)
                if not t:
                    return self.send_json(404, {"error": "Tarefa não encontrada"})
                if method == "PATCH":
                    if "feita" in body:
                        t["feita"] = bool(body["feita"])
                        t["feita_em"] = now_iso() if t["feita"] else None
                        t["feita_por"] = user["nome"] if t["feita"] else ""
                        if t["feita"]:
                            registra_hist(lead, user["nome"],
                                          ['✅ Tarefa concluída: "%s"' % t["texto"]], papel=user["papel"])
                    lead["updated_at"] = now_iso()
                    save_db()
                    return self.send_json(200, {"tarefa": t, "tarefas": tarefas})
                # DELETE
                lead["tarefas"] = [x for x in tarefas if x.get("id") != tarefa_id]
                registra_hist(lead, user["nome"],
                              ['🗑️ Tarefa excluída: "%s"' % t["texto"]], papel=user["papel"])
                lead["updated_at"] = now_iso()
                save_db()
                return self.send_json(200, {"ok": True, "tarefas": lead["tarefas"]})

        mn = re.match(r"^/api/leads/([^/]+)/notas$", path)
        if mn and method == "POST":
            lead_id = mn.group(1)
            try:
                body = self.read_body()
            except Exception:
                return self.send_json(400, {"error": "Corpo invalido"})
            texto = str(body.get("texto") or "").strip()[:2000]
            if not texto:
                return self.send_json(400, {"error": "Escreva a atualização antes de salvar"})
            op_id = str(body.get("op_id") or "").strip()[:64]
            with _lock:
                lead = next((l for l in _db["leads"] if l["id"] == lead_id), None)
                if not lead or not pode_ver_lead(user, lead):
                    return self.send_json(404, {"error": "Lead nao encontrado"})
                # idempotencia: nota registrada offline pode ser reenviada — se o
                # op_id ja existe, devolve a entrada existente sem duplicar.
                if op_id:
                    ja = next((h for h in lead.get("historico", []) if h.get("op_id") == op_id), None)
                    if ja:
                        return self.send_json(201, {"entrada": ja,
                                                    "aguardando_resposta": lead.get("aguardando_resposta")})
                registra_hist(lead, user["nome"], ["💬 " + texto], papel=user["papel"], tipo="nota", op_id=op_id or None, quando=ts_offline(body.get("criado_em")))
                # A atualizacao escrita E o registro da resposta: some o alerta
                # (e o aviso "cliente respondeu" tambem — foi tratado).
                lead["aguardando_resposta"] = None
                lead["cliente_respondeu"] = None
                lead["updated_at"] = now_iso()
                save_db()
                return self.send_json(201, {"entrada": lead["historico"][-1],
                                            "aguardando_resposta": lead["aguardando_resposta"],
                                            "cliente_respondeu": None})

        # Contato por WhatsApp: marca "aguardando o vendedor registrar a resposta"
        # e anota no historico. O alerta some quando alguem escreve uma nota.
        mc = re.match(r"^/api/leads/([^/]+)/contato-whatsapp$", path)
        if mc and method == "POST":
            lead_id = mc.group(1)
            with _lock:
                lead = next((l for l in _db["leads"] if l["id"] == lead_id), None)
                if not lead or not pode_ver_lead(user, lead):
                    return self.send_json(404, {"error": "Lead nao encontrado"})
                # Lead encerrado nao gera cobranca de registro.
                if lead.get("status") in ("ganho", "perdido", "desistiu", "curioso"):
                    return self.send_json(200, {"lead": lead})
                ja_pendente = bool(lead.get("aguardando_resposta"))
                ultimo = lead["historico"][-1] if lead.get("historico") else None
                # Anota o contato — menos em cliques repetidos (ja aguardando E a
                # ultima entrada ja foi um contato), para nao poluir o historico.
                if not (ja_pendente and ultimo and ultimo.get("tipo") == "contato"):
                    registra_hist(lead, user["nome"], ["📱 Contato por WhatsApp"],
                                  papel=user["papel"], tipo="contato")
                    lead["updated_at"] = now_iso()
                # So carimba o horario no PRIMEIRO contato sem resposta registrada:
                # reabrir a conversa nao pode zerar o cronometro (senao nunca fica
                # vermelho). O relogio conta desde o contato que ainda espera nota.
                if not ja_pendente:
                    lead["aguardando_resposta"] = now_iso()
                # foi responder pelo WhatsApp: os dois avisos nao coexistem
                # (o contador e a central contam LEADS, nao avisos)
                lead["cliente_respondeu"] = None
                save_db()
                return self.send_json(200, {"lead": lead})

        # Atender no Chatwoot: manda a saudacao automatica na conversa e devolve a
        # URL da conversa para o app abrir. (Canal OFICIAL — evita o WhatsApp pessoal.)
        # A saudacao sai 1x por ciclo de contato (mesmo debounce do historico) —
        # reclicar/reabrir NAO reenvia a mensagem ao cliente.
        mch = re.match(r"^/api/leads/([^/]+)/chatwoot$", path)
        if mch and method == "POST":
            lead_id = mch.group(1)
            with _lock:
                lead = next((l for l in _db["leads"] if l["id"] == lead_id), None)
                if not lead or not pode_ver_lead(user, lead):
                    return self.send_json(404, {"error": "Lead nao encontrado"})
                origem = origem_do_lead(lead)
                base, acc, token, saud = chatwoot_cfg(origem)
                configurado = bool(base and acc and token)
                conv = conv_id_valido(lead.get("chatwoot_conversation_id"))
                conv_url = chatwoot_conversa_url(base, acc, conv)
                nome = str(lead.get("nome") or "").strip()
                telefone = lead.get("telefone") or ""
                inbox = str(_db.get("settings", {}).get(inbox_cfg_key(origem)) or "").strip()
                autor, papel = user["nome"], user["papel"]
                ja_pendente = bool(lead.get("aguardando_resposta"))
                ultimo = lead["historico"][-1] if lead.get("historico") else None
                encerrado = lead.get("status") in ("ganho", "perdido", "desistiu", "curioso")
                # clique repetido no mesmo ciclo (ja aguardando + ultimo = contato)
                # nao reenvia; lead encerrado tambem nao recebe saudacao
                deve_enviar = bool(conv) and configurado and bool(saud) and not encerrado \
                    and not (ja_pendente and ultimo and ultimo.get("tipo") == "contato")
                # trava de duplo-clique: uma requisicao em voo por lead
                if lead_id in _cw_em_voo:
                    return self.send_json(200, {"conversa_url": conv_url, "enviada": False,
                                                "erro": None if conv else
                                                "Já estamos conectando este lead no Chatwoot — aguarde uns segundos",
                                                "tem_conversa": bool(conv),
                                                "configurado": configurado,
                                                "chatwoot_url": base, "chatwoot_account_id": acc})
                _cw_em_voo.add(lead_id)
            # respostas do cliente que chegarem DEPOIS deste instante (durante a
            # chamada externa) nao podem ser apagadas pelo pos-envio
            t_envio = now_iso()
            enviada, erro = False, None
            conectada = False
            try:
                # Lead SEM conversa (recuperacao/importado): CONECTA no Chatwoot —
                # acha o contato pelo telefone, REUSA a conversa antiga (com o
                # historico) ou cria uma nova na caixa de entrada. FORA do lock.
                if not conv and configurado and not encerrado:
                    try:
                        conv_novo, contato_id = chatwoot_conectar(base, acc, token, inbox, nome, telefone)
                        with _lock:
                            l2 = next((l for l in _db["leads"] if l["id"] == lead_id), None)
                            if l2:
                                ja_conv = conv_id_valido(l2.get("chatwoot_conversation_id"))
                                if ja_conv:  # outro clique conectou primeiro: respeita
                                    conv = ja_conv
                                else:
                                    l2["chatwoot_conversation_id"] = conv_novo
                                    if contato_id and not l2.get("chatwoot_contact_id"):
                                        l2["chatwoot_contact_id"] = contato_id
                                    registra_hist(l2, autor, ["🔗 Conversa conectada no Chatwoot"], papel=papel)
                                    l2["updated_at"] = now_iso()
                                    save_db()
                                    conv = conv_novo
                                    conectada = True
                        conv_url = chatwoot_conversa_url(base, acc, conv)
                        deve_enviar = bool(conv) and bool(saud) and \
                            not (ja_pendente and ultimo and ultimo.get("tipo") == "contato")
                    except ChatwootErro as e:
                        erro = str(e)
                    except Exception as e:
                        erro = "Não consegui conectar no Chatwoot (%s)" % type(e).__name__
                # envia a saudacao FORA do lock (chamada externa pode demorar)
                if deve_enviar:
                    primeiro = nome.split()[0] if nome else ""
                    if primeiro:
                        texto = saud.replace("{nome}", primeiro).replace("{primeiro_nome}", primeiro)
                    else:  # sem nome: tira o {nome} (e o espaco) p/ nao sair "Olá !"
                        texto = re.sub(r"\s*\{(nome|primeiro_nome)\}", "", saud)
                    try:
                        enviada = chatwoot_envia_mensagem(base, acc, conv, token, texto)
                    except urllib.error.HTTPError as e:
                        erro = "O Chatwoot recusou (HTTP %s) — confira o token e a conta" % e.code
                    except Exception as e:
                        erro = "Não consegui falar com o Chatwoot (%s)" % type(e).__name__
                # registra o contato e marca "aguardando resposta" — mas so quando
                # algo aconteceu de fato (saudacao enviada e/ou conversa aberta)
                if conv and not encerrado and (enviada or conv_url):
                    with _lock:
                        lead = next((l for l in _db["leads"] if l["id"] == lead_id), None)
                        if lead:
                            # RELE o estado (outro clique pode ter gravado durante a
                            # chamada externa) — snapshot velho duplicaria o historico
                            ja2 = bool(lead.get("aguardando_resposta"))
                            ult2 = lead["historico"][-1] if lead.get("historico") else None
                            if not (ja2 and ult2 and ult2.get("tipo") == "contato"):
                                msg = "📨 Saudação enviada pelo Chatwoot" if enviada else "💬 Atendimento aberto no Chatwoot"
                                registra_hist(lead, autor, [msg], papel=papel, tipo="contato")
                                lead["updated_at"] = now_iso()
                            # saudacao realmente enviada = contato novo: reinicia o
                            # relogio; so ABRIR a conversa (sem enviar) nao zera.
                            # Resposta do cliente chegada DURANTE o envio fica.
                            resp = lead.get("cliente_respondeu")
                            chegou_durante = bool(resp and str(resp) > t_envio)
                            if (enviada or not ja2) and not chegou_durante:
                                lead["aguardando_resposta"] = now_iso()
                                # exclusividade dos avisos: armar a cobranca de
                                # registro apaga o verde (foi responder por la)
                                lead["cliente_respondeu"] = None
                            save_db()
            finally:
                with _lock:
                    _cw_em_voo.discard(lead_id)
            with _lock:
                lead_atual = next((l for l in _db["leads"] if l["id"] == lead_id), None)
            return self.send_json(200, {"conversa_url": conv_url, "enviada": enviada,
                                        "conectada": conectada,
                                        "erro": erro, "tem_conversa": bool(conv),
                                        "configurado": configurado, "lead": lead_atual,
                                        "chatwoot_url": base, "chatwoot_account_id": acc})

        # Conversa do Chatwoot dentro do CRM (leitura): as mensagens do cliente
        # e da equipe, com anexos. O token NUNCA vai ao navegador — o CRM busca
        # e devolve so os campos mapeados.
        mchat = re.match(r"^/api/leads/([^/]+)/chatwoot-chat$", path)
        if mchat and method == "GET":
            with _lock:
                lead = next((l for l in _db["leads"] if l["id"] == mchat.group(1)), None)
                if not lead or not pode_ver_lead(user, lead):
                    return self.send_json(404, {"error": "Lead nao encontrado"})
                base, acc, token, _s = chatwoot_cfg(origem_do_lead(lead))
                if not (base and acc and token):
                    return self.send_json(400, {"error": "Chatwoot não configurado"})
                conv = conv_id_valido(lead.get("chatwoot_conversation_id"))
            if not conv:
                return self.send_json(200, {"mensagens": [], "sem_conversa": True})
            try:
                res = _cw_req(base, "/api/v1/accounts/%s/conversations/%s/messages" % (acc, conv), token)
            except urllib.error.HTTPError as e:
                return self.send_json(502, {"error": "O Chatwoot recusou (HTTP %s)" % e.code})
            except Exception as e:
                return self.send_json(502, {"error": "Não consegui falar com o Chatwoot (%s)" % type(e).__name__})
            itens = _cw_payload(res)
            mensagens = []
            if isinstance(itens, list):
                for m in itens:
                    if not isinstance(m, dict) or m.get("private"):
                        continue
                    mt = m.get("message_type")
                    if mt not in (0, 1, "incoming", "outgoing"):
                        continue  # 2 = atividade do sistema (resolvida, etiqueta...) — fora
                    anexos = []
                    for a in (m.get("attachments") or []):
                        if isinstance(a, dict) and a.get("data_url"):
                            anexos.append({"tipo": str(a.get("file_type") or "file"),
                                           "url": str(a.get("data_url"))})
                    texto = str(m.get("content") or "").strip()
                    if not texto and not anexos:
                        continue
                    snd = m.get("sender") if isinstance(m.get("sender"), dict) else {}
                    mensagens.append({
                        "de": "cliente" if mt in (0, "incoming") else "equipe",
                        "texto": texto[:4000],
                        "anexos": anexos[:5],
                        "autor": str(snd.get("name") or snd.get("available_name") or "")[:60],
                        "data": _cw_data_msg(m.get("created_at")),
                    })
            mensagens.sort(key=lambda x: str(x.get("data") or ""))
            return self.send_json(200, {"mensagens": mensagens[-100:]})

        # Mensagem escrita NO CRM e enviada pelo Chatwoot (canal oficial). Se o
        # lead ainda nao tem conversa, conecta primeiro (mesmo fluxo do botao).
        # Mensagem manual e deliberada: vale tambem para lead encerrado (pos-venda).
        mmsg = re.match(r"^/api/leads/([^/]+)/chatwoot-mensagem$", path)
        if mmsg and method == "POST":
            lead_id = mmsg.group(1)
            try:
                body = self.read_body()
            except Exception:
                return self.send_json(400, {"error": "Corpo invalido"})
            texto = str(body.get("texto") or "").strip()[:2000]
            # anexo opcional: {nome, mime, dados(base64)} — imagem/audio/video
            anexo = body.get("anexo") if isinstance(body.get("anexo"), dict) else None
            anexo_bytes, anexo_nome, anexo_mime = None, "", ""
            if anexo:
                anexo_mime = str(anexo.get("mime") or "").lower().split(";")[0].strip()
                anexo_nome = str(anexo.get("nome") or "arquivo")
                if anexo_mime not in ANEXO_MIMES:
                    return self.send_json(400, {"error": "Tipo de arquivo não aceito — envie imagem, áudio ou vídeo"})
                try:
                    anexo_bytes = base64.b64decode(str(anexo.get("dados") or ""), validate=True)
                except Exception:
                    return self.send_json(400, {"error": "Arquivo inválido — tente anexar de novo"})
                if not anexo_bytes:
                    return self.send_json(400, {"error": "Arquivo vazio — tente anexar de novo"})
                if len(anexo_bytes) > ANEXO_MAX_BYTES:
                    return self.send_json(400, {"error": "Arquivo muito grande — o limite é 10 MB"})
            if not texto and not anexo_bytes:
                return self.send_json(400, {"error": "Escreva a mensagem antes de enviar"})
            with _lock:
                lead = next((l for l in _db["leads"] if l["id"] == lead_id), None)
                if not lead or not pode_ver_lead(user, lead):
                    return self.send_json(404, {"error": "Lead nao encontrado"})
                origem = origem_do_lead(lead)
                base, acc, token, _saud = chatwoot_cfg(origem)
                if not (base and acc and token):
                    return self.send_json(400, {"error": "Chatwoot não configurado — peça ao gestor "
                                                "para configurar em Gerenciar → Campanhas"})
                conv = conv_id_valido(lead.get("chatwoot_conversation_id"))
                nome = str(lead.get("nome") or "").strip()
                telefone = lead.get("telefone") or ""
                inbox = str(_db.get("settings", {}).get(inbox_cfg_key(origem)) or "").strip()
                autor, papel = user["nome"], user["papel"]
                if lead_id in _cw_em_voo:
                    return self.send_json(409, {"error": "Já tem um envio em andamento para este lead — aguarde uns segundos"})
                _cw_em_voo.add(lead_id)
            # respostas do cliente que chegarem DEPOIS deste instante (durante a
            # chamada externa) nao podem ser apagadas pelo pos-envio
            t_envio = now_iso()
            conectada, enviada, erro = False, False, None
            try:
                # conecta se preciso (FORA do lock; mesmo padrao do botao)
                if not conv:
                    try:
                        conv_novo, contato_id = chatwoot_conectar(base, acc, token, inbox, nome, telefone)
                        with _lock:
                            l2 = next((l for l in _db["leads"] if l["id"] == lead_id), None)
                            if l2:
                                ja_conv = conv_id_valido(l2.get("chatwoot_conversation_id"))
                                if ja_conv:
                                    conv = ja_conv
                                else:
                                    l2["chatwoot_conversation_id"] = conv_novo
                                    if contato_id and not l2.get("chatwoot_contact_id"):
                                        l2["chatwoot_contact_id"] = contato_id
                                    registra_hist(l2, autor, ["🔗 Conversa conectada no Chatwoot"], papel=papel)
                                    l2["updated_at"] = now_iso()
                                    save_db()
                                    conv = conv_novo
                                    conectada = True
                    except ChatwootErro as e:
                        erro = str(e)
                    except Exception as e:
                        erro = "Não consegui conectar no Chatwoot (%s)" % type(e).__name__
                if conv and not erro:
                    try:
                        if anexo_bytes:
                            enviada = chatwoot_envia_anexo(base, acc, conv, token, texto,
                                                           anexo_nome, anexo_mime, anexo_bytes)
                        else:
                            enviada = chatwoot_envia_mensagem(base, acc, conv, token, texto)
                        if not enviada:
                            erro = "Não consegui enviar — recarregue e tente de novo"
                    except urllib.error.HTTPError as e:
                        erro = "O Chatwoot recusou o envio (HTTP %s)" % e.code
                    except Exception as e:
                        erro = "Não consegui falar com o Chatwoot (%s)" % type(e).__name__
                if enviada:
                    with _lock:
                        l3 = next((l for l in _db["leads"] if l["id"] == lead_id), None)
                        if l3:
                            if anexo_mime.startswith("image/"):
                                rotulo = "🖼 imagem"
                            elif anexo_mime.startswith("audio/"):
                                rotulo = "🎤 áudio"
                            elif anexo_mime.startswith("video/"):
                                rotulo = "🎬 vídeo"
                            else:
                                rotulo = ""
                            resumo = texto if len(texto) <= 200 else texto[:200] + "…"
                            item = '📤 Chatwoot: %s%s' % (rotulo, (' "%s"' % resumo) if resumo else "") \
                                if rotulo else ('📤 Chatwoot: "%s"' % resumo)
                            registra_hist(l3, autor, [item], papel=papel, tipo="contato")
                            # Mensagem DE FATO enviada reinicia o relogio da resposta
                            # pendente: o card cobra "ha X" desde o ULTIMO contato real
                            # (clique no botao de WhatsApp continua sem zerar, pois
                            # abrir a conversa nao prova que algo foi enviado).
                            # Se o cliente respondeu DURANTE o envio (webhook na
                            # janela da chamada externa), o aviso verde dele fica.
                            resp = l3.get("cliente_respondeu")
                            if not (resp and str(resp) > t_envio):
                                l3["cliente_respondeu"] = None  # respondemos: baixa o aviso
                                if l3.get("status") not in ("ganho", "perdido", "desistiu", "curioso"):
                                    l3["aguardando_resposta"] = now_iso()
                            l3["updated_at"] = now_iso()
                            save_db()
            finally:
                with _lock:
                    _cw_em_voo.discard(lead_id)
            if erro and not enviada:
                return self.send_json(502, {"error": erro})
            with _lock:
                lead_atual = next((l for l in _db["leads"] if l["id"] == lead_id), None)
            return self.send_json(200, {"ok": True, "enviada": enviada, "conectada": conectada,
                                        "lead": lead_atual,
                                        "conversa_url": chatwoot_conversa_url(base, acc, conv)})

        # Criar
        if path == "/api/leads" and method == "POST":
            try:
                body = self.read_body()
            except Exception:
                return self.send_json(400, {"error": "Corpo invalido"})
            # Regras por papel valem TAMBEM na criacao (o PATCH ja barra; sem
            # isto o SDR entraria nas etapas/paineis de venda por este caminho).
            # Valores efetivos dao 403; sobras vazias do formulario (checkbox
            # desmarcado, campo em branco) sao apenas descartadas.
            if user["papel"] == "sdr":
                if body.get("status") in ETAPAS_DE_VENDA:
                    return self.send_json(403, {"error": "Etapas de venda são trabalho dos vendedores — o SDR vai até a qualificação"})
                if body.get("em_servicos") or body.get("status_servico") or body.get("valor_servico"):
                    return self.send_json(403, {"error": "O painel de Serviços é trabalho dos vendedores"})
                if body.get("em_curso") or body.get("status_curso") or body.get("valor_curso"):
                    return self.send_json(403, {"error": "O painel do Curso é trabalho dos vendedores"})
                for k in ("em_servicos", "status_servico", "valor_servico",
                          "em_curso", "status_curso", "valor_curso"):
                    body.pop(k, None)
            # "mover etapas" desligado vale tambem na criacao: nao cadastra lead
            # ja em etapa avancada nem dentro dos paineis (senao o 403 do PATCH
            # seria contornavel criando o lead direto na etapa desejada)
            if user["papel"] in ("sdr", "vendedor") and not user.get("pode_mover", True):
                if body.get("status") not in (None, "", "novo", "triagem"):
                    return self.send_json(403, {"error": "Mover leads de etapa está bloqueado para você — cadastre o lead e peça a um colega para avançá-lo"})
                if body.get("em_servicos") or body.get("status_servico") \
                        or body.get("em_curso") or body.get("status_curso") \
                        or body.get("tipo") == "curso":
                    return self.send_json(403, {"error": "Mover leads de etapa está bloqueado para você — fale com o administrador"})
                for k in ("em_servicos", "status_servico", "em_curso", "status_curso"):
                    body.pop(k, None)
            with _lock:
                lead = make_lead({"source": "manual"})
                try:
                    apply_updates(lead, body)
                except ValueError as e:
                    return self.send_json(400, {"error": str(e)})
                # SDR cadastra para si; vendedor cadastra ja no proprio funil
                if user["papel"] == "sdr":
                    lead["sdr"] = user["nome"]
                    lead["responsavel"] = user["nome"]
                elif user["papel"] == "vendedor":
                    lead["vendedor"] = user["nome"]
                    lead["responsavel"] = user["nome"]
                    lead["atendido_em"] = now_iso()
                    if lead.get("status") not in VENDAS_STATUSES:
                        lead["status"] = "qualificado"
                        if not lead.get("tipo"):
                            lead["tipo"] = "produtor"
                        lead["qualificado_em"] = now_iso()
                if not str(lead.get("telefone") or "").strip() or not str(lead.get("email") or "").strip():
                    return self.send_json(400, {"error": "Telefone e e-mail são obrigatórios"})
                dup, campo = find_duplicado(lead["telefone"], lead["email"])
                if dup:
                    return self.send_json(400, {"error": "Já existe um lead com esse %s: %s" % (
                        campo, dup.get("nome") or dup.get("telefone") or "(sem nome)")})
                registra_hist(lead, user["nome"], ["🆕 Lead criado (cadastro manual)"], papel=user["papel"], tipo="novo")
                _db["leads"].append(lead)
                save_db()
                return self.send_json(201, {"lead": lead})

        # Editar / excluir por id
        m = re.match(r"^/api/leads/([^/]+)$", path)
        if m:
            lead_id = m.group(1)
            body = None
            if method in ("PATCH", "PUT"):
                try:
                    body = self.read_body()
                except Exception:
                    return self.send_json(400, {"error": "Corpo invalido"})
            with _lock:
                lead = next((l for l in _db["leads"] if l["id"] == lead_id), None)
                if not lead or not pode_ver_lead(user, lead):
                    # invisivel para este papel = como se nao existisse
                    return self.send_json(404, {"error": "Lead nao encontrado"})
                if method in ("PATCH", "PUT"):
                    # Regras por papel: cada um so mexe no que e seu
                    if user["papel"] == "sdr" and "sdr" in body and body["sdr"] != user["nome"]:
                        return self.send_json(403, {"error": "SDR não pode transferir o lead para outro SDR"})
                    # SDR nao faz trabalho de vendedor: nem etapas de venda,
                    # nem o painel de Servicos (pos-venda)
                    if user["papel"] == "sdr" and body.get("status") in ETAPAS_DE_VENDA:
                        return self.send_json(403, {"error": "Etapas de venda são trabalho dos vendedores — o SDR vai até a qualificação"})
                    # trava por PRESENCA da chave (nao por valor truthy): senao o
                    # SDR tiraria o lead do painel com em_servicos/em_curso=false,
                    # rebaixaria a etapa com "" ou editaria o valor
                    if user["papel"] == "sdr" and any(
                            k in body for k in ("em_servicos", "status_servico", "valor_servico")):
                        return self.send_json(403, {"error": "O painel de Serviços é trabalho dos vendedores"})
                    if user["papel"] == "sdr" and any(
                            k in body for k in ("em_curso", "status_curso", "valor_curso")):
                        return self.send_json(403, {"error": "O painel do Curso é trabalho dos vendedores"})
                    # permissao individual: quem esta com "mover etapas" desligado
                    # so consulta e registra notas — nao muda etapa de NENHUM funil,
                    # nem por caminhos indiretos (entrar/sair de painel move o lead
                    # de coluna; mudar a classificacao muda o funil dele)
                    if user["papel"] in ("sdr", "vendedor") and not user.get("pode_mover", True) \
                            and any(k in body for k in ("status", "status_servico", "status_curso",
                                                        "em_servicos", "em_curso", "tipo")):
                        return self.send_json(403, {"error": "Mover leads de etapa está bloqueado para você — fale com o administrador"})
                    if user["papel"] == "vendedor" and "vendedor" in body and body["vendedor"] not in ("", user["nome"]):
                        return self.send_json(403, {"error": "Vendedor só pode assumir o lead para si"})
                    # aplica numa copia: se uma regra barrar no meio, o lead
                    # original nao fica meio-editado na memoria
                    tentativa = dict(lead)
                    try:
                        apply_updates(tentativa, body)
                    except ValueError as e:
                        return self.send_json(400, {"error": str(e)})
                    if "telefone" in body or "email" in body:
                        # checa SO o campo alterado: um duplicado pre-existente
                        # no OUTRO campo nao pode travar esta edicao
                        dup, campo = find_duplicado(
                            tentativa.get("telefone") if "telefone" in body else None,
                            tentativa.get("email") if "email" in body else None,
                            exclude_id=lead["id"])
                        if dup:
                            return self.send_json(400, {"error": "Outro lead já usa esse %s: %s" % (
                                campo, dup.get("nome") or dup.get("telefone") or "(sem nome)")})
                    # registra no historico o que mudou (antes de sobrescrever)
                    itens = descreve_mudancas(lead, tentativa, list(body.keys()))
                    lead.update(tentativa)
                    registra_hist(lead, user["nome"], itens, papel=user["papel"])
                    save_db()
                    return self.send_json(200, {"lead": lead})
                if method == "DELETE":
                    if not gestor:
                        return self.send_json(403, {"error": "Só gerente/administrador pode excluir leads"})
                    _db["leads"] = [l for l in _db["leads"] if l["id"] != lead_id]
                    save_db()
                    return self.send_json(200, {"ok": True})

        return self.send_json(404, {"error": "Rota nao encontrada"})

    # -- arquivos estaticos --
    def serve_static(self, path):
        rel = "/index.html" if path == "/" else path
        # normaliza e impede path traversal
        rel = os.path.normpath(rel).lstrip("/\\")
        file_path = os.path.join(PUBLIC_DIR, rel)
        if not os.path.abspath(file_path).startswith(PUBLIC_DIR):
            self.send_response(403)
            self.end_headers()
            self.wfile.write(b"Forbidden")
            return
        if not os.path.isfile(file_path):
            self.send_response(404)
            self.send_header("Content-Type", "text/plain; charset=utf-8")
            self.end_headers()
            self.wfile.write("Nao encontrado".encode("utf-8"))
            return
        ext = os.path.splitext(file_path)[1].lower()
        with open(file_path, "rb") as f:
            data = f.read()
        self.send_response(200)
        self.send_header("Content-Type", MIME.get(ext, "application/octet-stream"))
        self.send_header("Content-Length", str(len(data)))
        # sem cache: apos atualizar o CRM, um recarregar simples ja traz a
        # versao nova (evita interface velha presa no navegador da equipe)
        self.send_header("Cache-Control", "no-cache")
        self.headers_seguranca()
        self.end_headers()
        self.wfile.write(data)


# ---------------------------------------------------------------------------
# Backup automatico diario do banco (a empresa ja perdeu dados uma vez por
# erro operacional — esta e a rede de protecao). 1 copia por dia, 14 guardadas.
# ---------------------------------------------------------------------------
BACKUPS_DIR = os.path.join(DATA_DIR, "backups")
BACKUPS_MANTER = 14


def faz_backup_diario():
    try:
        os.makedirs(BACKUPS_DIR, exist_ok=True)
        hoje = datetime.now(timezone.utc).astimezone(BRT).strftime("%Y-%m-%d")
        alvo = os.path.join(BACKUPS_DIR, "leads-%s.json" % hoje)
        if not os.path.exists(alvo) and os.path.exists(DB_FILE):
            with _lock:  # copia consistente (nao no meio de uma gravacao)
                shutil.copyfile(DB_FILE, alvo)
            print("[backup] copia diaria salva: %s" % alvo)
        # poda: guarda so as ultimas BACKUPS_MANTER copias
        arqs = sorted(f for f in os.listdir(BACKUPS_DIR)
                      if f.startswith("leads-") and f.endswith(".json"))
        for velho in arqs[:-BACKUPS_MANTER]:
            os.remove(os.path.join(BACKUPS_DIR, velho))
    except Exception as e:
        print("AVISO: backup diario falhou:", e)


def _laco_backup():
    while True:
        faz_backup_diario()
        time.sleep(3600)  # confere de hora em hora; grava 1 por dia


def main():
    load_db()
    recalcula_etapas()  # aplica as etapas personalizadas salvas nas configuracoes
    try:
        atlas_boot()    # descompacta/prepara o banco do Atlas de prospeccao
    except Exception as e:
        # o Atlas e um recurso SECUNDARIO: falha nele (disco cheio, gz ruim)
        # nao pode derrubar o CRM inteiro — a Prospeccao responde 503
        print("[atlas] AVISO: nao consegui preparar o Atlas (%s) — o CRM segue sem a Prospecção" % e)
    load_cidades()
    senha_admin = ensure_admin()
    ensure_webhook_token()
    threading.Thread(target=_laco_backup, daemon=True).start()  # backup diario
    httpd = ThreadingHTTPServer(("0.0.0.0", PORT), Handler)
    print("")
    print("  CRM Nova Era Drones rodando")
    if senha_admin:
        print("  -----------------------------------------------")
        print("  PRIMEIRO ACESSO -> login: admin | senha: %s" % senha_admin)
        print("  (troque a senha no painel de Usuarios)")
    print("  -----------------------------------------------")
    print("  Painel:   http://localhost:%d" % PORT)
    print("  Webhook:  http://localhost:%d/webhook/chatwoot?token=%s" % (PORT, WEBHOOK_TOKEN))
    print("  Leads salvos em: %s" % DB_FILE)
    print("  -----------------------------------------------")
    print("  (Ctrl+C para parar)")
    print("")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\n  Encerrando...")
        httpd.shutdown()


if __name__ == "__main__":
    main()
